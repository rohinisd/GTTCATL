import calendar
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.leave import LeaveRequest, Holiday, LeaveStatusEnum, PayrollAdjustment
from app.models.attendance import TrainerAttendance
from app.models.hierarchy import School, SchoolTrainer
from app.models.users import User, RoleEnum
from app.core.security import require_any, get_current_user
from app.core.timezone import now_ist, today_ist

router = APIRouter()

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def _parse_hhmm(t_str) -> Optional[int]:
    """Return total minutes from midnight for an 'HH:MM' string, or None."""
    if not t_str:
        return None
    try:
        h, m = str(t_str).split(':')[:2]
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return None


# ── Scope helpers ────────────────────────────────────────────────────────────

def _scoped_trainer_ids(db: Session, reviewer) -> Optional[set]:
    """User-ids of trainers the reviewer is allowed to manage.
    Returns None for state_admin (means: everyone)."""
    if reviewer.role == RoleEnum.state_admin:
        return None
    if reviewer.role == RoleEnum.division_master:
        rows = db.query(User.id).filter(
            User.role == RoleEnum.atl_trainer,
            User.division_id == reviewer.division_id,
        ).all()
        return {r[0] for r in rows}
    if reviewer.role == RoleEnum.master_trainer:
        if not reviewer.division_id:
            return set()
        school_ids = [s.id for s in db.query(School.id).filter(
            School.division_id == reviewer.division_id
        ).all()]
        rows = db.query(SchoolTrainer.user_id).filter(
            SchoolTrainer.school_id.in_(school_ids or [-1]),
            SchoolTrainer.is_current == True,
        ).all()
        return {r[0] for r in rows}
    return set()


def _fmt_leave(lv: LeaveRequest) -> dict:
    days = (lv.to_date - lv.from_date).days + 1
    return {
        "id":          lv.id,
        "user_id":     lv.user_id,
        "user_name":   lv.user.name if lv.user else None,
        "from_date":   lv.from_date.isoformat(),
        "to_date":     lv.to_date.isoformat(),
        "days":        days,
        "reason":      lv.reason,
        "status":      lv.status.value,
        "review_note": lv.review_note,
        "reviewed_by": lv.reviewer.name if lv.reviewer else None,
        "reviewed_at": lv.reviewed_at.isoformat() if lv.reviewed_at else None,
        "created_at":  lv.created_at.isoformat() if lv.created_at else None,
    }


def _fmt_holiday(h: Holiday) -> dict:
    return {
        "id":            h.id,
        "date":          h.date.isoformat(),
        "name":          h.name,
        "division_id":   h.division_id,
        "division_name": h.division.name if h.division else "All Divisions",
        "created_at":    h.created_at.isoformat() if h.created_at else None,
    }


# ── Leave: trainer self-service ──────────────────────────────────────────────

@router.post("/request")
def create_leave(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_any),
):
    if current_user.role != RoleEnum.atl_trainer:
        raise HTTPException(403, "Only trainers can request leave")
    try:
        from_date = date.fromisoformat(body["from_date"])
        to_date   = date.fromisoformat(body["to_date"])
    except (KeyError, ValueError):
        raise HTTPException(400, "from_date and to_date (YYYY-MM-DD) are required")
    if to_date < from_date:
        raise HTTPException(400, "to_date cannot be before from_date")

    # Block overlapping pending/approved requests
    overlap = db.query(LeaveRequest).filter(
        LeaveRequest.user_id == current_user.id,
        LeaveRequest.status.in_([LeaveStatusEnum.pending, LeaveStatusEnum.approved]),
        LeaveRequest.from_date <= to_date,
        LeaveRequest.to_date   >= from_date,
    ).first()
    if overlap:
        raise HTTPException(400, "You already have a leave request overlapping these dates")

    lv = LeaveRequest(
        user_id   = current_user.id,
        from_date = from_date,
        to_date   = to_date,
        reason    = body.get("reason", ""),
        status    = LeaveStatusEnum.pending,
    )
    db.add(lv)
    db.commit()
    db.refresh(lv)
    return _fmt_leave(lv)


@router.get("/my")
def my_leaves(
    db: Session = Depends(get_db),
    current_user=Depends(require_any),
):
    rows = db.query(LeaveRequest).filter(
        LeaveRequest.user_id == current_user.id
    ).order_by(LeaveRequest.from_date.desc()).all()
    return [_fmt_leave(r) for r in rows]


@router.delete("/request/{leave_id}")
def cancel_leave(
    leave_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(require_any),
):
    lv = db.query(LeaveRequest).filter(LeaveRequest.id == leave_id).first()
    if not lv or lv.user_id != current_user.id:
        raise HTTPException(404, "Leave request not found")
    if lv.status != LeaveStatusEnum.pending:
        raise HTTPException(400, "Only pending requests can be cancelled")
    db.delete(lv)
    db.commit()
    return {"ok": True}


# ── Leave: reviewer (master_trainer / division_master / admin) ────────────────

@router.get("/list")
def list_leaves(
    status:    Optional[str] = Query(None),
    user_id:   Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    month:     Optional[int] = Query(None, ge=1, le=12),
    year:      Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")
    q = db.query(LeaveRequest)
    scope = _scoped_trainer_ids(db, current_user)
    if scope is not None:
        q = q.filter(LeaveRequest.user_id.in_(scope or {-1}))
    if status:
        q = q.filter(LeaveRequest.status == status)
    if user_id:
        q = q.filter(LeaveRequest.user_id == user_id)

    # Filter by school: find trainers currently assigned to that school
    if school_id:
        trainer_ids = [
            r[0] for r in db.query(SchoolTrainer.user_id).filter(
                SchoolTrainer.school_id == school_id,
                SchoolTrainer.is_current == True,
            ).all()
        ]
        q = q.filter(LeaveRequest.user_id.in_(trainer_ids or [-1]))

    # Filter by month/year: leaves that overlap the given period
    if year and month:
        period_first = date(year, month, 1)
        period_last  = date(year, month, calendar.monthrange(year, month)[1])
        q = q.filter(LeaveRequest.from_date <= period_last,
                     LeaveRequest.to_date   >= period_first)
    elif year:
        q = q.filter(LeaveRequest.from_date >= date(year, 1, 1),
                     LeaveRequest.to_date   <= date(year, 12, 31))

    rows = q.order_by(LeaveRequest.from_date.desc()).limit(500).all()

    # Batch-load school names per trainer to avoid N+1
    uid_set = {lv.user_id for lv in rows}
    links   = db.query(SchoolTrainer).filter(
        SchoolTrainer.user_id.in_(uid_set), SchoolTrainer.is_current == True
    ).all() if uid_set else []
    sid_set = {lnk.school_id for lnk in links}
    snames  = {s.id: s.name for s in db.query(School).filter(School.id.in_(sid_set)).all()} if sid_set else {}
    trainer_school: dict = {}
    for lnk in links:
        trainer_school.setdefault(lnk.user_id, []).append(snames.get(lnk.school_id, ""))

    result = []
    for lv in rows:
        d = _fmt_leave(lv)
        d["school_name"] = ", ".join(filter(None, trainer_school.get(lv.user_id, []))) or None
        result.append(d)
    return result


@router.patch("/{leave_id}/review")
def review_leave(
    leave_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")
    lv = db.query(LeaveRequest).filter(LeaveRequest.id == leave_id).first()
    if not lv:
        raise HTTPException(404, "Leave request not found")
    scope = _scoped_trainer_ids(db, current_user)
    if scope is not None and lv.user_id not in scope:
        raise HTTPException(403, "This trainer is outside your scope")

    action = (body.get("action") or "").lower()
    if action not in ("approve", "reject"):
        raise HTTPException(400, "action must be 'approve' or 'reject'")
    lv.status      = LeaveStatusEnum.approved if action == "approve" else LeaveStatusEnum.rejected
    lv.review_note = body.get("note", "")
    lv.reviewed_by = current_user.id
    lv.reviewed_at = now_ist()
    db.commit()
    db.refresh(lv)
    return _fmt_leave(lv)


# ── Holidays ──────────────────────────────────────────────────────────────────

@router.get("/holidays")
def list_holidays(
    year:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_any),
):
    q = db.query(Holiday)
    # Trainers only see holidays relevant to their division (or state-wide)
    if current_user.role == RoleEnum.atl_trainer and current_user.division_id:
        q = q.filter((Holiday.division_id == None) | (Holiday.division_id == current_user.division_id))
    if year:
        q = q.filter(Holiday.date >= date(year, 1, 1), Holiday.date <= date(year, 12, 31))
    rows = q.order_by(Holiday.date.asc()).all()
    return [_fmt_holiday(h) for h in rows]


@router.post("/holidays")
def add_holiday(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised to add holidays")
    try:
        d = date.fromisoformat(body["date"])
    except (KeyError, ValueError):
        raise HTTPException(400, "date (YYYY-MM-DD) is required")
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Holiday name is required")

    # Scope: only state_admin may create state-wide (all divisions) holidays.
    division_id = body.get("division_id")
    if current_user.role != RoleEnum.state_admin:
        # master_trainer / division_master pin to their own division
        division_id = current_user.division_id

    dup = db.query(Holiday).filter(
        Holiday.date == d,
        Holiday.division_id == division_id,
    ).first()
    if dup:
        raise HTTPException(400, "A holiday already exists for this date/division")

    h = Holiday(date=d, name=name, division_id=division_id, created_by=current_user.id)
    db.add(h)
    db.commit()
    db.refresh(h)
    return _fmt_holiday(h)


@router.patch("/holidays/{holiday_id}")
def update_holiday(
    holiday_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")
    h = db.query(Holiday).filter(Holiday.id == holiday_id).first()
    if not h:
        raise HTTPException(404, "Holiday not found")
    if current_user.role != RoleEnum.state_admin and h.division_id != current_user.division_id:
        raise HTTPException(403, "Outside your scope")
    if body.get("name"):
        h.name = body["name"].strip()
    if body.get("date"):
        try:
            h.date = date.fromisoformat(body["date"])
        except ValueError:
            raise HTTPException(400, "Invalid date format (use YYYY-MM-DD)")
    db.commit()
    db.refresh(h)
    return _fmt_holiday(h)


@router.delete("/holidays/{holiday_id}")
def delete_holiday(
    holiday_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")
    h = db.query(Holiday).filter(Holiday.id == holiday_id).first()
    if not h:
        raise HTTPException(404, "Holiday not found")
    if current_user.role != RoleEnum.state_admin and h.division_id != current_user.division_id:
        raise HTTPException(403, "Outside your scope")
    db.delete(h)
    db.commit()
    return {"ok": True}


# ── Payroll summary ───────────────────────────────────────────────────────────

def _build_payroll_payload(db: Session, year: int, month: int, school_id: Optional[int], reviewer) -> dict:
    days_in_month = calendar.monthrange(year, month)[1]
    first = date(year, month, 1)
    last  = date(year, month, days_in_month)

    # Pre-compute month-level values (reused per trainer)
    sundays = sum(1 for d in range(1, days_in_month + 1)
                  if date(year, month, d).weekday() == 6)
    mon_sat_count = days_in_month - sundays   # all Mon–Sat days in month

    # Which trainers are in scope?
    scope = _scoped_trainer_ids(db, reviewer)
    tq = db.query(User).filter(User.role == RoleEnum.atl_trainer, User.is_active == True)
    if scope is not None:
        tq = tq.filter(User.id.in_(scope or {-1}))
    trainers = tq.all()

    # Pre-load all payroll adjustments for this month (keyed by user_id)
    adj_rows = db.query(PayrollAdjustment).filter(
        PayrollAdjustment.year  == year,
        PayrollAdjustment.month == month,
    ).all()
    adjustments = {a.user_id: a for a in adj_rows}

    results = []
    for t in trainers:
        # All current school assignments for this trainer
        links = db.query(SchoolTrainer).filter(
            SchoolTrainer.user_id == t.id, SchoolTrainer.is_current == True
        ).all()
        schools = [
            s for s in (
                db.query(School).filter(School.id == lnk.school_id).first()
                for lnk in links
            ) if s
        ]
        school_count = len(schools)

        # Apply school_id filter (trainer must be linked to that school)
        if school_id and not any(s.id == school_id for s in schools):
            continue

        # Use first school's division for holiday scoping
        primary = schools[0] if schools else None

        # Holidays for this trainer's division (non-Sundays only)
        hq = db.query(Holiday).filter(Holiday.date >= first, Holiday.date <= last)
        if primary:
            hq = hq.filter((Holiday.division_id == None) | (Holiday.division_id == primary.division_id))
        else:
            hq = hq.filter(Holiday.division_id == None)
        holiday_dates = {h.date for h in hq.all() if h.date.weekday() != 6}
        holidays = len(holiday_dates)

        # ── Working days ────────────────────────────────────────────────────
        # 2 schools → 6 days/week: all Mon–Sat minus holidays
        # 1 school  → 3 days/week: half of Mon–Sat minus holidays
        #   (trainer chooses any 3 days; we proportionally halve the pool)
        available = mon_sat_count - holidays   # workable Mon–Sat days this month
        if school_count >= 2:
            working_days = available
        else:
            working_days = round(available / 2)

        # ── Present days ────────────────────────────────────────────────────
        present_rows = db.query(TrainerAttendance.date).filter(
            TrainerAttendance.user_id == t.id,
            TrainerAttendance.date >= first,
            TrainerAttendance.date <= last,
            TrainerAttendance.check_in_at != None,
        ).all()
        present_dates = {r[0] for r in present_rows}
        present = len(present_dates)

        # ── Leave policy ────────────────────────────────────────────────────
        # Approved leaves this month, sorted earliest first.
        # • 1st leave request → casual leave → PAID (no deduction)
        # • All subsequent requests → treated as absent → UNPAID (deducted)
        leaves = db.query(LeaveRequest).filter(
            LeaveRequest.user_id == t.id,
            LeaveRequest.status  == LeaveStatusEnum.approved,
            LeaveRequest.from_date <= last,
            LeaveRequest.to_date   >= first,
        ).order_by(LeaveRequest.from_date.asc()).all()

        casual_dates = set()   # first leave → paid
        unpaid_dates = set()   # subsequent leaves → absent

        for idx, lv in enumerate(leaves):
            d   = max(lv.from_date, first)
            end = min(lv.to_date,   last)
            while d <= end:
                if d.weekday() != 6 and d not in holiday_dates:
                    if idx == 0:
                        casual_dates.add(d)
                    else:
                        unpaid_dates.add(d)
                d = date.fromordinal(d.toordinal() + 1)

        # Days already marked present don't count as leave
        casual_dates -= present_dates
        unpaid_dates -= present_dates

        casual_leave = len(casual_dates)   # paid (first leave)
        unpaid_leave = len(unpaid_dates)   # deducted like absents

        # Unexplained absents = working days not covered by present, casual, or unpaid leave
        absent = max(working_days - present - casual_leave - unpaid_leave, 0)

        # ── Salary ────────────────────────────────────────────────────
        salary  = t.salary or 0
        per_day = salary / working_days if working_days else 0
        # Deduct unpaid leave + unexplained absent; casual leave is paid
        payable = round(salary - (unpaid_leave + absent) * per_day, 2) if salary else 0

        school_names = ", ".join(s.name for s in schools) if schools else None

        # ── Late hours ───────────────────────────────────────────────────────
        # Late check-in (arrived after school_start_time) +
        # Early check-out (left before school_end_time), summed across all
        # present days, expressed in decimal hours.
        late_minutes = 0
        attn_records = db.query(TrainerAttendance).filter(
            TrainerAttendance.user_id == t.id,
            TrainerAttendance.date >= first,
            TrainerAttendance.date <= last,
            TrainerAttendance.check_in_at != None,
        ).all()
        for rec in attn_records:
            rec_school = next((s for s in schools if s.id == rec.school_id), None)
            if not rec_school:
                rec_school = db.query(School).filter(School.id == rec.school_id).first()
            if not rec_school:
                continue
            start_mins = _parse_hhmm(rec_school.school_start_time)
            if start_mins is not None:
                ci_mins = rec.check_in_at.hour * 60 + rec.check_in_at.minute
                if ci_mins > start_mins:
                    late_minutes += ci_mins - start_mins
            end_mins = _parse_hhmm(rec_school.school_end_time)
            if end_mins is not None and rec.check_out_at:
                co_mins = rec.check_out_at.hour * 60 + rec.check_out_at.minute
                if co_mins < end_mins:
                    late_minutes += end_mins - co_mins
        late_hours = round(late_minutes / 60, 2) if late_minutes > 0 else 0

        adj_obj      = adjustments.get(t.id)
        adjustment   = round(adj_obj.adjustment, 2) if adj_obj else 0
        adj_remarks  = adj_obj.remarks if adj_obj else ""
        adj_by       = adj_obj.adjuster.name if (adj_obj and adj_obj.adjuster) else None
        final_payable = round(payable + adjustment, 2)

        results.append({
            "user_id":       t.id,
            "user_name":     t.name,
            "school_id":     primary.id   if primary else None,
            "school_name":   school_names,
            "school_count":  school_count,
            "salary":        salary,
            "days_in_month": days_in_month,
            "sundays":       sundays,
            "holidays":      holidays,
            "working_days":  working_days,
            "present":       present,
            "casual_leave":  casual_leave,
            "unpaid_leave":  unpaid_leave,
            "on_leave":      casual_leave + unpaid_leave,
            "absent":        absent,
            "late_hours":    late_hours,
            "payable":       payable,          # calculated, before adjustment
            "adjustment":    adjustment,       # manual override
            "adj_remarks":   adj_remarks,
            "adj_by":        adj_by,
            "final_payable": final_payable,    # what is actually paid
        })

    results.sort(key=lambda r: r["user_name"] or "")
    return {
        "year": year, "month": month,
        "days_in_month": days_in_month, "sundays": sundays,
        "rows": results,
    }


@router.get("/payroll")
def payroll(
    year:  int = Query(...),
    month: int = Query(..., ge=1, le=12),
    school_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")
    return _build_payroll_payload(db, year, month, school_id, current_user)


@router.get("/payroll-excel")
def export_payroll_excel(
    year:  int = Query(...),
    month: int = Query(..., ge=1, le=12),
    school_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")

    payload = _build_payroll_payload(db, year, month, school_id, current_user)
    rows = payload["rows"]
    if not rows:
        raise HTTPException(404, "No payroll rows found for the selected month")

    from pathlib import Path
    from copy import copy
    import openpyxl

    template_path = Path(__file__).resolve().parents[1] / "static" / "Payroll.xlsx"
    wb = openpyxl.load_workbook(template_path)
    if "Attendance Register" not in wb.sheetnames:
        raise HTTPException(500, "Payroll template sheet not found")

    ws = wb["Attendance Register"]
    ws["B3"] = f"Full Time Guest Lecturer Honorarium Attendance Register - Month: {MONTH_NAMES[month]} - {year}"
    ws["B4"] = "Centre: Karnataka - GTTC/ATL Portal"

    data_start = 7
    footer_row = 20
    existing_data_rows = max(0, footer_row - data_start)
    extra_rows = max(0, len(rows) - existing_data_rows)

    for rng in ("B20:K20", "F23:G23", "I23:K23"):
        if rng in {str(r) for r in ws.merged_cells.ranges}:
            ws.unmerge_cells(rng)

    if extra_rows:
        ws.insert_rows(footer_row, amount=extra_rows)

    note_row = footer_row + extra_rows
    prepared_row = footer_row + extra_rows + 3
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=11)
    ws.cell(note_row, 2).value = "Note: Leave Days = Trainer applied approved leave days only."
    ws.merge_cells(start_row=prepared_row, start_column=6, end_row=prepared_row, end_column=7)
    ws.merge_cells(start_row=prepared_row, start_column=9, end_row=prepared_row, end_column=11)

    for rr in range(data_start, note_row):
        for cc in range(2, 12):
            ws.cell(rr, cc).value = None

    for idx, r in enumerate(rows):
        row_num = data_start + idx
        leave_days = (r.get("casual_leave") or 0) + (r.get("unpaid_leave") or 0)
        non_payable_days = (r.get("unpaid_leave") or 0) + (r.get("absent") or 0)
        remarks = (r.get("adj_remarks") or "").strip()
        if not remarks and (r.get("adjustment") or 0) != 0:
            remarks = f"Adjustment ₹{(r.get('adjustment') or 0):,.2f}"
        if not remarks and ((r.get("casual_leave") or 0) + (r.get("unpaid_leave") or 0)) > 0:
            remarks = f"Casual {r.get('casual_leave') or 0}; Unpaid {r.get('unpaid_leave') or 0}"

        values = [
            idx + 1,
            r.get("user_name") or "",
            r.get("salary") or 0,
            payload["days_in_month"],
            r.get("working_days") or 0,
            leave_days,
            r.get("present") or 0,
            non_payable_days,
            r.get("late_hours") or 0,
            remarks,
        ]
        for offset, value in enumerate(values, 2):
            cell = ws.cell(row_num, offset, value=value)
            cell.alignment = copy(ws.cell(7, offset).alignment)
            if offset in (4, 10):
                cell.number_format = '#,##0.00'
            elif offset in (2, 5, 6, 7, 8, 9):
                cell.number_format = '0'
        ws.row_dimensions[row_num].height = 18

    fname = f"payroll_{year}_{str(month).zfill(2)}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/payroll-adjustment")
def save_payroll_adjustment(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master, RoleEnum.master_trainer):
        raise HTTPException(403, "Not authorised")

    user_id    = body.get("user_id")
    year       = body.get("year")
    month      = body.get("month")
    adjustment = float(body.get("adjustment", 0))
    remarks    = (body.get("remarks") or "").strip()

    if not user_id or not year or not month:
        raise HTTPException(400, "user_id, year, and month are required")
    if not remarks:
        raise HTTPException(400, "Remarks are required for a payroll adjustment")

    existing = db.query(PayrollAdjustment).filter(
        PayrollAdjustment.user_id == user_id,
        PayrollAdjustment.year   == year,
        PayrollAdjustment.month  == month,
    ).first()

    if existing:
        existing.adjustment  = adjustment
        existing.remarks     = remarks
        existing.adjusted_by = current_user.id
    else:
        db.add(PayrollAdjustment(
            user_id=user_id, year=year, month=month,
            adjustment=adjustment, remarks=remarks,
            adjusted_by=current_user.id,
        ))

    db.commit()
    return {"ok": True}
