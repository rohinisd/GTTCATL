from datetime import datetime
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.reports import (
    MonthlyReport, ReportStatusEnum,
    Project, ProjectItem, ReportProjectUsage,
    EquipmentInventory, ReportUsedItem, ReportBrokenItem,
)
from app.models.hierarchy import School
from app.core.security import require_any, get_download_user
from app.services.aggregator import get_scoped_reports

router = APIRouter()

MONTH_NAMES = ["","January","February","March","April","May","June",
               "July","August","September","October","November","December"]


@router.get("/excel")
def export_excel(
    academic_year: str = Query("2024-25"),
    division_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.status == ReportStatusEnum.submitted,
    )
    if division_id: q = q.filter(School.division_id == division_id)  # School already joined in get_scoped_reports
    reports = q.order_by(MonthlyReport.report_year, MonthlyReport.report_month).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active; ws.title = "Monthly Reports"
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))
    alt_fill  = PatternFill("solid", fgColor="EEF4FF")

    ws.merge_cells("A1:S1")
    ws["A1"] = f"GTTC Robotics Portal — Monthly Report ({academic_year})"
    ws["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws["A1"].alignment = ctr_align
    ws.row_dimensions[1].height = 25

    headers = [
        "School Name","UDISE Code","District","Division","Lab Type",
        "Month","Year","Students (School)","Students (Community)","Girls",
        "Workshops","Mentoring","Projects","Patents","Copyrights",
        "ATL Won","Other Won","Status","Submitted On",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = ctr_align; cell.border = thin
    ws.row_dimensions[2].height = 35

    for ri, r in enumerate(reports, 3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        row_data = [
            r.school.name if r.school else "",
            r.school.udise_code if r.school else "",
            r.school.district if r.school else "",
            r.school.division.name if r.school and r.school.division else "",
            r.school.lab_type.value.upper() if r.school and r.school.lab_type else "",
            MONTH_NAMES[r.report_month] if r.report_month else "",
            r.report_year,
            r.students_school, r.students_community, r.students_girls,
            r.workshops_count, r.mentoring_sessions, r.innovation_projects,
            r.patents_filed, r.copyrights_filed,
            r.atl_competitions_won, r.other_competitions_won,
            r.status.value.capitalize(),
            str(r.submitted_at.date()) if r.submitted_at else "",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin; cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for ci in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, len(reports) + 3)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"; ws2["B1"] = "Value"
    ws2["A1"].font = Font(bold=True); ws2["B1"].font = Font(bold=True)
    summary = [
        ("Schools Reporting",          len(set(r.school_id for r in reports))),
        ("Total Students (School)",    sum(r.students_school or 0 for r in reports)),
        ("Total Students (Community)", sum(r.students_community or 0 for r in reports)),
        ("Total Girls",                sum(r.students_girls or 0 for r in reports)),
        ("Total Workshops",            sum(r.workshops_count or 0 for r in reports)),
        ("Total Mentoring",            sum(r.mentoring_sessions or 0 for r in reports)),
        ("Total Projects",             sum(r.innovation_projects or 0 for r in reports)),
        ("Total Patents",              sum(r.patents_filed or 0 for r in reports)),
        ("ATL Competitions Won",       sum(r.atl_competitions_won or 0 for r in reports)),
        ("Other Competitions Won",     sum(r.other_competitions_won or 0 for r in reports)),
    ]
    for i, (m, v) in enumerate(summary, 2):
        ws2[f"A{i}"] = m; ws2[f"B{i}"] = v
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"gttc_{academic_year.replace('-','_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/pdf")
def export_pdf(
    academic_year: str = Query("2024-25"),
    division_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    q = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.status == ReportStatusEnum.submitted,
    )
    if division_id: q = q.filter(School.division_id == division_id)  # School already joined in get_scoped_reports
    reports = q.all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Title"], fontSize=14,
                              textColor=colors.HexColor("#1E3A5F"))
    sub_s   = ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                              textColor=colors.grey, alignment=TA_CENTER)
    cell_s  = ParagraphStyle("c", parent=styles["Normal"], fontSize=7)

    story = [
        Paragraph("GTTC Robotics Training Portal", title_s),
        Paragraph(f"Monthly Report — {academic_year}", sub_s),
        Spacer(1, 0.3*cm),
    ]
    headers = ["School","UDISE","District","Month","Students","Community","Workshops","Projects","Won","Status"]
    data = [headers] + [
        [
            Paragraph(r.school.name[:28] if r.school else "", cell_s),
            r.school.udise_code if r.school else "",
            r.school.district if r.school else "",
            f"{MONTH_NAMES[r.report_month][:3]} {r.report_year}" if r.report_month else "",
            str(r.students_school), str(r.students_community),
            str(r.workshops_count), str(r.innovation_projects),
            str(r.total_won),
            r.status.value.capitalize(),
        ]
        for r in reports
    ]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),  (-1,0),  colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",     (0,0),  (-1,0),  colors.white),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,0),  8),
        ("FONTSIZE",      (0,1),  (-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1),  (-1,-1), [colors.white, colors.HexColor("#EEF4FF")]),
        ("GRID",          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ("ALIGN",         (0,0),  (-1,-1), "CENTER"),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),  (-1,-1), 3),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Total Schools: {len(set(r.school_id for r in reports))} | "
        f"Total Students: {sum(r.students_school or 0 for r in reports):,} | "
        f"Total Projects: {sum(r.innovation_projects or 0 for r in reports):,} | "
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_s,
    ))
    doc.build(story)
    buf.seek(0)
    fname = f"gttc_{academic_year.replace('-','_')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ─── helper ────────────────────────────────────────────────────────────────
def _xl_response(wb, fname: str) -> StreamingResponse:
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _make_wb_styles():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    alt_fill  = PatternFill("solid", fgColor="EEF4FF")
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))
    return hdr_font, hdr_fill, alt_fill, ctr, left, thin


# ═══════════════════════════════════════════════════════════════════════════
#  ITEM USAGE COUNT EXPORT  (Projects & Item Usage → Item Usage Count tab)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/item-stats")
def export_item_stats(
    academic_year: Optional[str] = Query(None),
    month_from: Optional[int]    = Query(None),
    month_to:   Optional[int]    = Query(None),
    project_id: Optional[int]    = Query(None),
    status:     Optional[str]    = Query(None),
    division_id: Optional[int]   = Query(None),
    district:   Optional[str]    = Query(None),
    school_id:  Optional[int]    = Query(None),
    _token:     Optional[str]    = Query(None),
    db: Session = Depends(get_db), current_user=Depends(get_download_user),
):
    """Excel export of per-school item usage and broken data, honouring all filters."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    hdr_font, hdr_fill, alt_fill, ctr, left, thin = _make_wb_styles()

    # ── Fetch reports (base scope) ──
    rq = db.query(MonthlyReport).join(School, School.id == MonthlyReport.school_id)
    if school_id:    rq = rq.filter(MonthlyReport.school_id == school_id)
    if district:     rq = rq.filter(School.district == district)
    if division_id:  rq = rq.filter(School.division_id == division_id)
    if academic_year: rq = rq.filter(MonthlyReport.academic_year == academic_year)
    if month_from:   rq = rq.filter(MonthlyReport.report_month >= month_from)
    if month_to:     rq = rq.filter(MonthlyReport.report_month <= month_to)
    all_reports = rq.all()
    all_ids = [r.id for r in all_reports]

    item_ids = all_ids
    if project_id:
        linked = {r.report_id for r in db.query(ReportProjectUsage.report_id).filter(
            ReportProjectUsage.project_id == project_id).all()}
        item_ids = [rid for rid in all_ids if rid in linked]

    # ── Per-school aggregation ──
    sch_used   = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "usage_count": 0}))
    sch_broken = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "reasons": []}))
    try:
        if status != "broken":
            for row in db.query(ReportUsedItem).filter(ReportUsedItem.report_id.in_(item_ids or [-1])).all():
                k = row.item_name.strip()
                sch_used[row.school_id][k]["quantity"]    += row.quantity or 1
                sch_used[row.school_id][k]["usage_count"] += row.usage_count or 1
        if status != "working":
            for row in db.query(ReportBrokenItem).filter(ReportBrokenItem.report_id.in_(item_ids or [-1])).all():
                k = row.item_name.strip()
                sch_broken[row.school_id][k]["quantity"] += row.quantity or 1
                if row.reason and row.reason not in sch_broken[row.school_id][k]["reasons"]:
                    sch_broken[row.school_id][k]["reasons"].append(row.reason)
    except Exception:
        pass

    all_sids = set(sch_used.keys()) | set(sch_broken.keys())
    sch_objs = {s.id: s for s in db.query(School).filter(School.id.in_(all_sids or [-1])).all()}

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Item Usage Count"

    title_row = f"GTTC — Item Usage Count{' · '+academic_year if academic_year else ''}{' · '+district if district else ''}"
    ws.merge_cells("A1:G1")
    ws["A1"] = title_row
    ws["A1"].font = Font(bold=True, size=12, color="1E3A5F")
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 22

    headers = ["School", "District", "Division", "Item Name", "Quantity", "Usage Count", "Broken", "Reason"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    ws.row_dimensions[2].height = 30

    row_idx = 3
    for sid in sorted(all_sids, key=lambda x: sch_objs.get(x, type("_",(),({"name":"~"}))()).name):
        sch = sch_objs.get(sid)
        if not sch: continue
        all_keys = set(sch_used[sid].keys()) | set(sch_broken[sid].keys())
        for k in sorted(all_keys):
            u = sch_used[sid].get(k, {})
            b = sch_broken[sid].get(k, {})
            if status == "working" and u.get("quantity", 0) == 0: continue
            if status == "broken"  and b.get("quantity", 0) == 0: continue
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            vals = [
                sch.name, sch.district or "",
                sch.division.name if sch.division else "",
                k,
                u.get("quantity", 0), u.get("usage_count", 0),
                b.get("quantity", 0),
                " | ".join(b.get("reasons", [])) or "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=ci, value=v)
                c.border = thin; c.fill = fill
                c.alignment = left if ci in (1, 4, 8) else ctr
            row_idx += 1

    col_widths = [30, 18, 18, 35, 12, 14, 10, 35]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    fname = f"item_usage{('_'+academic_year.replace('-','_')) if academic_year else ''}.xlsx"
    return _xl_response(wb, fname)


# ═══════════════════════════════════════════════════════════════════════════
#  EQUIPMENT REVIEW EXPORT  (Issue Equipment page)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/equipment-review")
def export_equipment_review(
    division_id: Optional[int] = Query(None),
    _token:      Optional[str] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(get_download_user),
):
    """Excel export of equipment review overview (per-school working / broken totals)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_font, hdr_fill, alt_fill, ctr, left, thin = _make_wb_styles()

    sq = db.query(School)
    if division_id: sq = sq.filter(School.division_id == division_id)
    schools = sq.order_by(School.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Equipment Review"

    ws.merge_cells("A1:J1")
    ws["A1"] = "GTTC — Equipment Review Overview"
    ws["A1"].font = Font(bold=True, size=12, color="1E3A5F"); ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 22

    headers = ["School", "District", "Division", "Item Name",
               "Issued", "Working", "Not Working", "Review Status", "Last Checked", "Notes"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    ws.row_dimensions[2].height = 28

    row_idx = 3
    for sch in schools:
        items = db.query(EquipmentInventory).filter(
            EquipmentInventory.school_id == sch.id
        ).order_by(EquipmentInventory.item_name).all()
        for eq in items:
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            reviewed = "Reviewed" if eq.reviewed_at else "Pending"
            vals = [
                sch.name, sch.district or "",
                sch.division.name if sch.division else "",
                eq.item_name,
                eq.quantity or 0,
                eq.working_qty if eq.working_qty is not None else "—",
                eq.not_working_qty if eq.not_working_qty is not None else "—",
                reviewed,
                str(eq.last_checked) if eq.last_checked else "—",
                eq.review_notes or "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=ci, value=v)
                c.border = thin; c.fill = fill
                c.alignment = left if ci in (1, 4, 10) else ctr
            row_idx += 1

    col_widths = [30, 18, 18, 35, 10, 12, 14, 14, 14, 35]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    fname = f"equipment_review.xlsx"
    return _xl_response(wb, fname)


# ═══════════════════════════════════════════════════════════════════════════
#  PROJECTS EXPORT  (Projects tab)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/projects")
def export_projects(
    academic_year: Optional[str] = Query(None),
    division_id:   Optional[int] = Query(None),
    district:      Optional[str] = Query(None),
    school_id:     Optional[int] = Query(None),
    _token:        Optional[str] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(get_download_user),
):
    """Excel export of projects with their bill-of-materials and school usage counts."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlalchemy import func as sqfunc

    hdr_font, hdr_fill, alt_fill, ctr, left, thin = _make_wb_styles()
    grp_fill = PatternFill("solid", fgColor="D6E4FF")

    pq = db.query(Project)
    if academic_year: pq = pq.filter(Project.academic_year == academic_year)

    usage_q = (db.query(ReportProjectUsage.project_id,
                        sqfunc.count(sqfunc.distinct(ReportProjectUsage.school_id)))
               .join(School, School.id == ReportProjectUsage.school_id))
    if division_id: usage_q = usage_q.filter(School.division_id == division_id)
    if district:    usage_q = usage_q.filter(School.district == district)
    if school_id:   usage_q = usage_q.filter(ReportProjectUsage.school_id == school_id)
    count_map = dict(usage_q.group_by(ReportProjectUsage.project_id).all())

    projects = pq.order_by(Project.exp_no.asc().nullslast(), Project.id.asc()).all()

    wb = openpyxl.Workbook()

    # Sheet 1: Projects summary
    ws = wb.active; ws.title = "Projects"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"GTTC — Projects{(' · '+academic_year) if academic_year else ''}"
    ws["A1"].font = Font(bold=True, size=12, color="1E3A5F"); ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(["Exp #","Project Name","Description","Status","Items","Schools Using"], 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    ws.row_dimensions[2].height = 28

    for ri, p in enumerate(projects, 3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, v in enumerate([
            p.exp_no or "", p.name, p.description or "",
            "Active" if p.is_active else "Inactive",
            len(p.items), count_map.get(p.id, 0),
        ], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin; c.fill = fill
            c.alignment = left if ci in (2, 3) else ctr
    for ci, w in enumerate([8, 50, 40, 12, 10, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2: Bill of Materials (flat)
    ws2 = wb.create_sheet("Bill of Materials")
    for ci, h in enumerate(["Exp #","Project Name","Item Name","Quantity"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    ri = 2
    for p in projects:
        for it in p.items:
            fill = alt_fill if ri % 2 == 0 else PatternFill()
            for ci, v in enumerate([p.exp_no or "", p.name, it.item_name, it.quantity or ""], 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.border = thin; c.fill = fill
                c.alignment = left if ci in (2, 3) else ctr
            ri += 1
    for ci, w in enumerate([8, 40, 35, 15], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    fname = f"projects{('_'+academic_year.replace('-','_')) if academic_year else ''}.xlsx"
    return _xl_response(wb, fname)
