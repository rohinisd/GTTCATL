from datetime import date, datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO

from app.database import get_db
from app.models.users import User, RoleEnum
from app.models.hierarchy import Division, SubDivision, School, SchoolTrainer, SchoolPrincipal
from app.models.reports import MonthlyReport, ActivityLog, BulkUpload, ReportStatusEnum, ATLCompetition, Notification, CalendarEvent, DivisionTarget, EquipmentInventory
from app.schemas.user_schema import UserCreate, UserUpdate, SchoolCreate, SchoolUpdate
from app.schemas.report_schema import ReportCreate, ReportUpdate
from app.core.security import (
    hash_password, get_current_user,
    require_state_admin, require_sub_and_above, require_any,
)

router = APIRouter()

MONTH_NAMES = ["","January","February","March","April","May","June",
               "July","August","September","October","November","December"]


def get_available_school_ids(school_ids, assigned_school_ids):
    assigned = set(assigned_school_ids or [])
    return [school_id for school_id in school_ids if school_id not in assigned]


def _apply_school_scope(q, db: Session, current_user) -> object:
    if current_user.role == RoleEnum.master_trainer:
        districts = current_user.districts or []
        if districts:
            return q.filter(School.district.in_(districts))
        return q.filter(False)
    if current_user.role == RoleEnum.atl_trainer:
        ids = [st.school_id for st in db.query(SchoolTrainer).filter(
            SchoolTrainer.user_id == current_user.id, SchoolTrainer.is_current == True).all()]
        return q.filter(School.id.in_(ids))
    if current_user.role == RoleEnum.principal:
        ids = [sp.school_id for sp in db.query(SchoolPrincipal).filter(
            SchoolPrincipal.user_id == current_user.id, SchoolPrincipal.is_current == True).all()]
        return q.filter(School.id.in_(ids))
    return q


# ════════════════════════════════════════════════════════
# DIVISIONS
# ════════════════════════════════════════════════════════

@router.get("/divisions")
def list_divisions(db: Session = Depends(get_db), _=Depends(require_any)):
    return db.query(Division).filter(Division.is_active == True).all()


@router.post("/divisions")
def create_division(
    name: str, code: str, state: str = "Karnataka",
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    if db.query(Division).filter(Division.code == code).first():
        raise HTTPException(400, "Division code already exists")
    d = Division(name=name, code=code, state=state)
    db.add(d); db.commit(); db.refresh(d)
    return d


@router.delete("/divisions/{division_id}")
def delete_division(
    division_id: int, db: Session = Depends(get_db),
    current_user=Depends(require_state_admin),
):
    d = db.query(Division).filter(Division.id == division_id).first()
    if not d: raise HTTPException(404, "Division not found")
    ActivityLog.log(db, current_user.id, "division_deleted",
        model_type="Division", model_id=division_id,
        description=f"Deleted division: {d.name}")
    db.delete(d); db.commit()
    return {"message": f"Division '{d.name}' deleted"}


# ════════════════════════════════════════════════════════
# SUB-DIVISIONS
# ════════════════════════════════════════════════════════

@router.get("/sub-divisions")
def list_sub_divisions(
    division_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), _=Depends(require_any),
):
    q = db.query(SubDivision).filter(SubDivision.is_active == True)
    if division_id: q = q.filter(SubDivision.division_id == division_id)
    return q.all()


@router.post("/sub-divisions")
def create_sub_division(
    division_id: int, name: str, code: str,
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    if db.query(SubDivision).filter(SubDivision.code == code).first():
        raise HTTPException(400, "Code already exists")
    s = SubDivision(division_id=division_id, name=name, code=code)
    db.add(s); db.commit(); db.refresh(s)
    return s


@router.delete("/sub-divisions/{sub_id}")
def delete_sub_division(
    sub_id: int, db: Session = Depends(get_db),
    current_user=Depends(require_state_admin),
):
    s = db.query(SubDivision).filter(SubDivision.id == sub_id).first()
    if not s: raise HTTPException(404, "Sub-division not found")
    ActivityLog.log(db, current_user.id, "subdivision_deleted",
        model_type="SubDivision", model_id=sub_id,
        description=f"Deleted sub-division: {s.name}")
    db.delete(s); db.commit()
    return {"message": f"Sub-division '{s.name}' deleted"}


# ════════════════════════════════════════════════════════
# USERS
# ════════════════════════════════════════════════════════

@router.get("/users")
def list_users(
    role: Optional[str] = Query(None),
    division_id: Optional[int] = Query(None),
    sub_division_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0), limit: int = Query(50, le=200),
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    q = db.query(User)
    if role:            q = q.filter(User.role == role)
    if division_id:     q = q.filter(User.division_id == division_id)
    if sub_division_id: q = q.filter(User.sub_division_id == sub_division_id)
    if is_active is not None: q = q.filter(User.is_active == is_active)
    if search: q = q.filter(User.name.ilike(f"%{search}%") | User.email.ilike(f"%{search}%"))
    users = q.offset(skip).limit(limit).all()
    return [
        {
            "id": u.id, "name": u.name, "email": u.email, "phone": u.phone,
            "role": u.role.value, "is_active": u.is_active,
            "salary": u.salary or 0,
            "division_id": u.division_id, "sub_division_id": u.sub_division_id,
            "last_login_at": u.last_login_at, "created_at": u.created_at,
        }
        for u in users
    ]


@router.post("/users")
def create_user(
    body: UserCreate,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    # Role-based creation permissions
    allowed_by_role = {
        RoleEnum.state_admin:     [RoleEnum.division_master, RoleEnum.master_trainer, RoleEnum.atl_trainer, RoleEnum.principal],
        RoleEnum.division_master: [RoleEnum.division_master, RoleEnum.master_trainer, RoleEnum.atl_trainer, RoleEnum.principal],
        RoleEnum.master_trainer:  [RoleEnum.atl_trainer, RoleEnum.principal],
        RoleEnum.atl_trainer:     [RoleEnum.principal],
    }
    allowed = allowed_by_role.get(current_user.role, [])
    if body.role not in allowed:
        raise HTTPException(403, f"You cannot create users with role '{body.role.value}'")

    # Scope division/sub-division to the creator's own hierarchy
    div_id  = body.division_id
    sub_id  = body.sub_division_id
    if current_user.role == RoleEnum.atl_trainer:
        div_id = current_user.division_id
        sub_id = current_user.sub_division_id

    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(400, "Email already registered")
    u = User(
        name=body.name, email=body.email, phone=body.phone,
        hashed_password=hash_password(body.password),
        role=body.role, gender=body.gender, caste=body.caste,
        qualification=body.qualification, salary=body.salary or 0, dob=body.dob,
        division_id=div_id, sub_division_id=sub_id,
        districts=body.districts if body.role == RoleEnum.master_trainer else None,
        is_active=True, must_change_password=True,
    )
    db.add(u); db.flush()
    # Assign to school if school_id provided
    if body.school_id:
        from datetime import date
        if body.role == RoleEnum.atl_trainer:
            db.add(SchoolTrainer(school_id=body.school_id, user_id=u.id,
                                 assigned_from=date.today(), is_current=True))
        elif body.role == RoleEnum.principal:
            db.add(SchoolPrincipal(school_id=body.school_id, user_id=u.id,
                                   assigned_from=date.today(), is_current=True))
    ActivityLog.log(db, current_user.id, "user_created",
        model_type="User", model_id=u.id,
        description=f"Created user {u.name} ({u.role.value})")
    db.commit(); db.refresh(u)
    return {"id": u.id, "name": u.name, "email": u.email, "role": u.role.value}


@router.put("/users/{user_id}")
def update_user(
    user_id: int, body: UserUpdate,
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(u, field, val)
    ActivityLog.log(db, current_user.id, "user_updated",
        model_type="User", model_id=u.id, description=f"Updated user {u.name}")
    db.commit(); db.refresh(u)
    return {"id": u.id, "name": u.name, "email": u.email}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    if u.role == RoleEnum.state_admin:
        raise HTTPException(400, "Cannot delete state admin")
    ActivityLog.log(db, current_user.id, "user_deleted",
        model_type="User", model_id=user_id,
        description=f"Deleted user: {u.name} ({u.role.value})")
    db.delete(u); db.commit()
    return {"message": f"User '{u.name}' deleted"}


@router.patch("/users/{user_id}/toggle-active")
def toggle_user_active(
    user_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    if u.role == RoleEnum.state_admin: raise HTTPException(400, "Cannot deactivate admin")
    u.is_active = not u.is_active
    ActivityLog.log(db, current_user.id,
        "user_activated" if u.is_active else "user_deactivated",
        model_type="User", model_id=u.id,
        description=f"{'Activated' if u.is_active else 'Deactivated'} {u.name}")
    db.commit()
    return {"is_active": u.is_active}


@router.post("/users/{user_id}/reset-password")
def reset_password(
    user_id: int, db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    u.hashed_password      = hash_password("Temp@1234")
    u.must_change_password = True
    ActivityLog.log(db, current_user.id, "password_reset",
        model_type="User", model_id=u.id, description=f"Reset password for {u.name}")
    db.commit()
    return {"message": "Password reset to Temp@1234"}


# ════════════════════════════════════════════════════════
# TRAINER SCHOOL ASSIGNMENTS
# ════════════════════════════════════════════════════════

@router.get("/users/{user_id}/schools")
def get_trainer_schools(
    user_id: int,
    db: Session = Depends(get_db), _=Depends(require_sub_and_above),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u or u.role != RoleEnum.atl_trainer:
        raise HTTPException(404, "Trainer not found")
    links = db.query(SchoolTrainer).filter(
        SchoolTrainer.user_id == user_id, SchoolTrainer.is_current == True
    ).all()
    result = []
    for lnk in links:
        s = db.query(School).filter(School.id == lnk.school_id).first()
        if s:
            result.append({
                "school_id": s.id, "school_name": s.name,
                "udise_code": s.udise_code, "district": s.district,
                "assigned_from": str(lnk.assigned_from),
            })
    return result


@router.get("/users/{user_id}/assignable-schools")
def get_assignable_schools(
    user_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u or u.role != RoleEnum.atl_trainer:
        raise HTTPException(404, "Trainer not found")

    q = db.query(School).filter(School.is_active == True)
    q = _apply_school_scope(q, db, current_user)

    assigned_ids = {row.school_id for row in db.query(SchoolTrainer.school_id).filter(SchoolTrainer.is_current == True).all()}
    if assigned_ids:
        q = q.filter(School.id.notin_(list(assigned_ids)))

    schools = q.order_by(School.name).all()
    return [{
        "id": s.id,
        "name": s.name,
        "udise_code": s.udise_code,
        "district": s.district,
        "division_id": s.division_id,
    } for s in schools]


@router.post("/users/{user_id}/assign-school")
def assign_school_to_trainer(
    user_id: int,
    school_id: int = Body(..., embed=True),
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u or u.role != RoleEnum.atl_trainer:
        raise HTTPException(404, "Trainer not found")
    current_count = db.query(SchoolTrainer).filter(
        SchoolTrainer.user_id == user_id, SchoolTrainer.is_current == True
    ).count()
    if current_count >= 2:
        raise HTTPException(400, "Trainer already has 2 schools assigned (maximum allowed)")
    s = db.query(School).filter(School.id == school_id, School.is_active == True).first()
    if not s:
        raise HTTPException(404, "School not found")

    existing_assignment = db.query(SchoolTrainer).filter(
        SchoolTrainer.school_id == school_id, SchoolTrainer.is_current == True,
    ).first()
    if existing_assignment:
        if existing_assignment.user_id == user_id:
            raise HTTPException(400, "This school is already assigned to the trainer")
        raise HTTPException(400, "This school already has an active trainer assignment")

    db.add(SchoolTrainer(school_id=school_id, user_id=user_id,
                         assigned_from=date.today(), is_current=True))
    ActivityLog.log(db, current_user.id, "school_assigned",
        model_type="SchoolTrainer", model_id=school_id,
        description=f"Assigned school '{s.name}' to trainer {u.name}")
    db.commit()
    return {"message": f"School '{s.name}' assigned to {u.name}"}


@router.delete("/users/{user_id}/schools/{school_id}")
def remove_school_from_trainer(
    user_id: int, school_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    link = db.query(SchoolTrainer).filter(
        SchoolTrainer.user_id == user_id, SchoolTrainer.school_id == school_id,
        SchoolTrainer.is_current == True,
    ).first()
    if not link:
        raise HTTPException(404, "Assignment not found")
    link.is_current = False
    link.assigned_to = date.today()
    ActivityLog.log(db, current_user.id, "school_unassigned",
        model_type="SchoolTrainer", model_id=school_id,
        description=f"Removed school {school_id} from trainer {user_id}")
    db.commit()
    return {"message": "School assignment removed"}


# ════════════════════════════════════════════════════════
# SCHOOLS
# ════════════════════════════════════════════════════════

@router.get("/schools")
def list_schools(
    division_id: Optional[int] = Query(None),
    sub_division_id: Optional[int] = Query(None),
    district: Optional[str] = Query(None),
    lab_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0), limit: int = Query(100, le=500),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    q = db.query(School).filter(School.is_active == True)
    q = _apply_school_scope(q, db, current_user)
    if division_id:     q = q.filter(School.division_id == division_id)
    if sub_division_id: q = q.filter(School.sub_division_id == sub_division_id)
    if district: q = q.filter(School.district.ilike(f"%{district}%"))
    if lab_type: q = q.filter(School.lab_type == lab_type)
    if search: q = q.filter(School.name.ilike(f"%{search}%") | School.udise_code.ilike(f"%{search}%"))
    schools = q.order_by(School.name).offset(skip).limit(limit).all()

    # Fetch current trainer for each school in one query
    school_ids = [s.id for s in schools]
    trainer_map = {}
    if school_ids:
        trainer_rows = (
            db.query(SchoolTrainer.school_id, User.name, User.email, User.phone)
            .join(User, SchoolTrainer.user_id == User.id)
            .filter(SchoolTrainer.school_id.in_(school_ids), SchoolTrainer.is_current == True)
            .all()
        )
        for row in trainer_rows:
            trainer_map[row.school_id] = {"name": row.name, "email": row.email, "phone": row.phone}

    principal_ids = set()
    if school_ids:
        from app.models.hierarchy import SchoolPrincipal as _SP
        for row in db.query(_SP.school_id).filter(
            _SP.school_id.in_(school_ids), _SP.is_current == True,
        ).all():
            principal_ids.add(row.school_id)

    return [
        {
            "id": s.id, "udise_code": s.udise_code, "atl_lab_code": s.atl_lab_code,
            "name": s.name, "district": s.district, "pin_code": s.pin_code,
            "division_id": s.division_id, "sub_division_id": s.sub_division_id,
            "division_name": s.division.name if s.division else None,
            "school_type": s.school_type.value if s.school_type else None,
            "lab_type": s.lab_type.value if s.lab_type else None,
            "education_type": s.education_type.value if s.education_type else None,
            "max_grade": s.max_grade,
            "principal_name": s.principal_name, "principal_email": s.principal_email,
            "principal_phone": s.principal_phone, "lab_area_sqft": s.lab_area_sqft,
            "lab_launch_date": str(s.lab_launch_date) if s.lab_launch_date else None,
            "is_active": s.is_active, "created_at": s.created_at,
            "trainer_name":  trainer_map.get(s.id, {}).get("name"),
            "trainer_email": trainer_map.get(s.id, {}).get("email"),
            "trainer_phone": trainer_map.get(s.id, {}).get("phone"),
            "geo_latitude":       s.geo_latitude,
            "geo_longitude":      s.geo_longitude,
            "geo_radius":         s.geo_radius or 200,
            "school_start_time":  s.school_start_time,
            "school_end_time":    s.school_end_time,
            "completeness": _school_completeness(
                s,
                has_trainer   = s.id in trainer_map,
                has_principal = s.id in principal_ids,
            ),
        }
        for s in schools
    ]


@router.post("/schools")
def create_school(
    body: SchoolCreate,
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    if db.query(School).filter(School.udise_code == body.udise_code).first():
        raise HTTPException(400, "UDISE code already exists")
    if body.atl_lab_code and db.query(School).filter(School.atl_lab_code == body.atl_lab_code).first():
        raise HTTPException(400, "ATL Lab code already exists")
    if db.query(User).filter(User.email == body.trainer_email).first():
        raise HTTPException(400, "Trainer email already registered")

    school_data = {k: v for k, v in body.model_dump().items()
                   if not k.startswith("trainer_")}
    school = School(**school_data)
    db.add(school); db.flush()

    trainer = User(
        name=body.trainer_name, email=body.trainer_email,
        phone=body.trainer_phone, gender=body.trainer_gender, caste=body.trainer_caste,
        dob=body.trainer_dob, salary=body.trainer_salary or 0,
        hashed_password=hash_password("Pass@1234"),
        role=RoleEnum.atl_trainer,
        division_id=body.division_id, sub_division_id=body.sub_division_id,
        is_active=True, must_change_password=True,
    )
    db.add(trainer); db.flush()
    db.add(SchoolTrainer(school_id=school.id, user_id=trainer.id,
                          assigned_from=date.today(), is_current=True))

    if body.principal_email:
        if not db.query(User).filter(User.email == body.principal_email).first():
            p = User(
                name=body.principal_name or "Principal",
                email=body.principal_email, phone=body.principal_phone,
                hashed_password=hash_password("Pass@1234"),
                role=RoleEnum.principal,
                division_id=body.division_id, sub_division_id=body.sub_division_id,
                is_active=True, must_change_password=True,
            )
            db.add(p); db.flush()
            db.add(SchoolPrincipal(school_id=school.id, user_id=p.id,
                                    assigned_from=date.today(), is_current=True))

    ActivityLog.log(db, current_user.id, "school_created",
        model_type="School", model_id=school.id,
        description=f"Created school {school.name} (UDISE: {school.udise_code})")
    db.commit(); db.refresh(school)
    return {"id": school.id, "name": school.name, "udise_code": school.udise_code}


@router.put("/schools/{school_id}")
def update_school(
    school_id: int, body: SchoolUpdate,
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    school = db.query(School).filter(School.id == school_id).first()
    if not school: raise HTTPException(404, "School not found")
    if current_user.role == RoleEnum.principal:
        raise HTTPException(403, "Principals cannot edit school data")
    if current_user.role == RoleEnum.atl_trainer:
        link = db.query(SchoolTrainer).filter(
            SchoolTrainer.school_id == school_id,
            SchoolTrainer.user_id == current_user.id,
            SchoolTrainer.is_current == True,
        ).first()
        if not link: raise HTTPException(403, "Not your school")
    old = {k: str(getattr(school, k)) for k in body.model_dump() if body.model_dump()[k] is not None}
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(school, field, val)
    ActivityLog.log(db, current_user.id, "school_updated",
        model_type="School", model_id=school_id,
        old_values=old, description=f"Updated school {school.name}")
    db.commit(); db.refresh(school)
    return {"id": school.id, "name": school.name}


@router.delete("/schools/{school_id}")
def delete_school(
    school_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    school = db.query(School).filter(School.id == school_id).first()
    if not school: raise HTTPException(404, "School not found")
    ActivityLog.log(db, current_user.id, "school_deleted",
        model_type="School", model_id=school_id,
        description=f"Deleted school: {school.name} (UDISE: {school.udise_code})")
    db.delete(school); db.commit()
    return {"message": f"School '{school.name}' deleted"}


# ════════════════════════════════════════════════════════
# MONTHLY REPORTS
# ════════════════════════════════════════════════════════

@router.get("/reports")
def list_reports(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    academic_year: Optional[str] = Query(None),
    school_id: Optional[int] = Query(None),
    division_id: Optional[int] = Query(None),
    sub_division_id: Optional[int] = Query(None),
    district: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    skip: int = Query(0, ge=0), limit: int = Query(100, le=500),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.services.aggregator import get_scoped_reports
    q = get_scoped_reports(db, current_user)  # already joins School
    if year:            q = q.filter(MonthlyReport.report_year == year)
    if month:           q = q.filter(MonthlyReport.report_month == month)
    if academic_year:   q = q.filter(MonthlyReport.academic_year == academic_year)
    if school_id:       q = q.filter(MonthlyReport.school_id == school_id)
    if status:          q = q.filter(MonthlyReport.status == status)
    if division_id:     q = q.filter(School.division_id == division_id)
    if sub_division_id: q = q.filter(School.sub_division_id == sub_division_id)
    if district:        q = q.filter(School.district.ilike(f"%{district}%"))
    rows = q.order_by(MonthlyReport.report_year.desc(),
                      MonthlyReport.report_month.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id": r.id, "school_id": r.school_id, "submitted_by": r.submitted_by,
            "school_name": r.school.name if r.school else None,
            "school_district": r.school.district if r.school else None,
            "report_year": r.report_year, "report_month": r.report_month,
            "academic_year": r.academic_year,
            "students_school": r.students_school, "students_community": r.students_community,
            "students_girls": r.students_girls,
            "workshops_school": r.workshops_school, "workshops_community": r.workshops_community,
            "workshops_count": r.workshops_count,
            "mentoring_school": r.mentoring_school, "mentoring_community": r.mentoring_community,
            "mentoring_sessions": r.mentoring_sessions,
            "innovation_school": r.innovation_school, "innovation_community": r.innovation_community,
            "innovation_projects": r.innovation_projects,
            "patents_school": r.patents_school, "patents_community": r.patents_community,
            "patents_filed": r.patents_filed,
            "copyrights_school": r.copyrights_school, "copyrights_community": r.copyrights_community,
            "copyrights_filed": r.copyrights_filed,
            "atl_competitions_participated": r.atl_competitions_participated,
            "atl_competitions_won": r.atl_competitions_won,
            "other_competitions_participated": r.other_competitions_participated,
            "other_competitions_won": r.other_competitions_won,
            "highlight_of_month": r.highlight_of_month,
            "social_post_link_1": r.social_post_link_1,
            "status": r.status.value, "submitted_at": r.submitted_at,
            "total_students": r.total_students, "total_won": r.total_won,
        }
        for r in rows
    ]


@router.post("/reports")
def create_report(
    body: ReportCreate,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    if current_user.role == RoleEnum.principal:
        raise HTTPException(403, "Principals cannot submit reports")
    if current_user.role == RoleEnum.atl_trainer:
        link = db.query(SchoolTrainer).filter(
            SchoolTrainer.school_id == body.school_id,
            SchoolTrainer.user_id == current_user.id,
            SchoolTrainer.is_current == True,
        ).first()
        if not link: raise HTTPException(403, "Not your school")

    existing = db.query(MonthlyReport).filter(
        MonthlyReport.school_id == body.school_id,
        MonthlyReport.report_year == body.report_year,
        MonthlyReport.report_month == body.report_month,
    ).first()

    if existing:
        if existing.status == ReportStatusEnum.submitted and current_user.role == RoleEnum.atl_trainer:
            raise HTTPException(400, "Already submitted. Contact Division Master (SPM GTTC) to edit.")
        for f, v in body.model_dump(exclude={"school_id"}).items():
            if v is not None: setattr(existing, f, v)
        report = existing
    else:
        report = MonthlyReport(**body.model_dump(), submitted_by=current_user.id)
        db.add(report)

    ActivityLog.log(db, current_user.id, "report_saved",
        model_type="MonthlyReport", description=f"Saved report {body.report_month}/{body.report_year}")
    db.commit(); db.refresh(report)
    return {"id": report.id, "status": report.status.value}


@router.post("/reports/{report_id}/submit")
def submit_report(
    report_id: int, db: Session = Depends(get_db), current_user=Depends(require_any),
):
    r = db.query(MonthlyReport).filter(MonthlyReport.id == report_id).first()
    if not r: raise HTTPException(404, "Report not found")
    if r.status == ReportStatusEnum.submitted: raise HTTPException(400, "Already submitted")
    r.status = ReportStatusEnum.submitted
    r.submitted_at = datetime.utcnow()
    r.submitted_by = current_user.id
    ActivityLog.log(db, current_user.id, "report_submitted",
        model_type="MonthlyReport", model_id=report_id,
        description=f"Submitted report {r.report_month}/{r.report_year} for school {r.school_id}")
    db.commit()
    return {"message": "Report submitted successfully"}


@router.put("/reports/{report_id}")
def update_report(
    report_id: int, body: ReportUpdate,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    r = db.query(MonthlyReport).filter(MonthlyReport.id == report_id).first()
    if not r: raise HTTPException(404, "Report not found")
    if current_user.role == RoleEnum.atl_trainer and r.status == ReportStatusEnum.submitted:
        raise HTTPException(403, "Cannot edit submitted report")
    if current_user.role == RoleEnum.principal:
        raise HTTPException(403, "Principals cannot edit reports")
    old = {f: getattr(r, f) for f in body.model_dump() if body.model_dump()[f] is not None}
    for f, v in body.model_dump(exclude_none=True).items():
        setattr(r, f, v)
    ActivityLog.log(db, current_user.id, "report_updated",
        model_type="MonthlyReport", model_id=report_id,
        old_values={k: str(v) for k, v in old.items()},
        description=f"Updated report {r.id}")
    db.commit(); db.refresh(r)
    return {"id": r.id, "status": r.status.value}


@router.delete("/reports/{report_id}")
def delete_report(
    report_id: int, db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    r = db.query(MonthlyReport).filter(MonthlyReport.id == report_id).first()
    if not r: raise HTTPException(404, "Report not found")
    ActivityLog.log(db, current_user.id, "report_deleted",
        model_type="MonthlyReport", model_id=report_id,
        description=f"DELETED report {r.report_month}/{r.report_year} for school {r.school_id}")
    db.delete(r); db.commit()
    return {"message": "Report deleted"}


# ════════════════════════════════════════════════════════
# EXPORT — EXCEL
# ════════════════════════════════════════════════════════

@router.get("/export/excel")
def export_excel(
    academic_year: str = Query(None),
    division_id: Optional[int] = Query(None),
    sub_division_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.services.aggregator import get_scoped_reports

    if not academic_year:
        latest_row = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest_row[0] if latest_row else "2025-26"

    q = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.status == ReportStatusEnum.submitted,
    )
    if division_id:     q = q.filter(School.division_id == division_id)        # School already joined
    if sub_division_id: q = q.filter(School.sub_division_id == sub_division_id)  # in get_scoped_reports
    reports = q.order_by(MonthlyReport.report_year, MonthlyReport.report_month).all()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Monthly Reports"

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))
    alt_fill  = PatternFill("solid", fgColor="EEF4FF")

    ws.merge_cells("A1:T1")
    ws["A1"] = f"GTTC Robotics Portal — Monthly Report ({academic_year})"
    ws["A1"].font      = Font(bold=True, size=13, color="1E3A5F")
    ws["A1"].alignment = ctr_align
    ws.row_dimensions[1].height = 25

    headers = [
        "School Name","UDISE Code","District","Division","Sub-Division","Lab Type",
        "Month","Year","Students (School)","Students (Community)","Girls",
        "Workshops","Mentoring","Projects","Patents","Copyrights",
        "ATL Won","Other Won","Status","Submitted On",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = ctr_align
        cell.border    = thin
    ws.row_dimensions[2].height = 35

    for ri, r in enumerate(reports, 3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        row_data = [
            r.school.name if r.school else "",
            r.school.udise_code if r.school else "",
            r.school.district if r.school else "",
            r.school.division.name if r.school and r.school.division else "",
            r.school.sub_division.name if r.school and r.school.sub_division else "",
            r.school.lab_type.value.upper() if r.school and r.school.lab_type else "",
            MONTH_NAMES[r.report_month] if r.report_month else "",
            r.report_year,
            r.students_school, r.students_community, r.students_girls,
            r.workshops_count, r.mentoring_sessions, r.innovation_projects,
            r.patents_filed, r.copyrights_filed,
            r.atl_competitions_won, r.other_competitions_won,
            r.status.value.capitalize(),
            str(r.submitted_at.date()) if r.submitted_at else "",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin; cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for ci in range(1, len(headers)+1):
        max_len = max((len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, len(reports)+3)), default=10)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

    # Latest report per school for accurate YTD totals
    latest_by_school = {}
    for r in reports:
        ym = r.report_year * 100 + r.report_month
        prev = latest_by_school.get(r.school_id)
        if prev is None or ym > prev[0]:
            latest_by_school[r.school_id] = (ym, r)
    latest = [v[1] for v in latest_by_school.values()]

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"; ws2["B1"] = "Value"
    ws2["A1"].font = Font(bold=True); ws2["B1"].font = Font(bold=True)
    summary = [
        ("Schools Reporting",          len(latest_by_school)),
        ("Total Students (School)",    sum(r.students_school or 0 for r in latest)),
        ("Total Students (Community)", sum(r.students_community or 0 for r in latest)),
        ("Total Girls",                sum(r.students_girls or 0 for r in latest)),
        ("Total Workshops",            sum(r.workshops_count or 0 for r in latest)),
        ("Total Mentoring",            sum(r.mentoring_sessions or 0 for r in latest)),
        ("Total Projects",             sum(r.innovation_projects or 0 for r in latest)),
        ("Total Patents",              sum(r.patents_filed or 0 for r in latest)),
        ("ATL Competitions Won",       sum(r.atl_competitions_won or 0 for r in latest)),
        ("Other Competitions Won",     sum(r.other_competitions_won or 0 for r in latest)),
    ]
    for i, (m, v) in enumerate(summary, 2):
        ws2[f"A{i}"] = m; ws2[f"B{i}"] = v
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"gttc_{academic_year.replace('-','_')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ════════════════════════════════════════════════════════
# EXPORT — PDF
# ════════════════════════════════════════════════════════

@router.get("/export/pdf")
def export_pdf(
    academic_year: str = Query(None),
    division_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from app.services.aggregator import get_scoped_reports

    if not academic_year:
        latest_row = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest_row[0] if latest_row else "2025-26"

    q = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.status == ReportStatusEnum.submitted,
    )
    if division_id: q = q.filter(School.division_id == division_id)  # School already joined in get_scoped_reports
    reports = q.order_by(MonthlyReport.report_year, MonthlyReport.report_month).all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Title"], fontSize=14,
                              textColor=colors.HexColor("#1E3A5F"))
    sub_s   = ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                              textColor=colors.grey, alignment=TA_CENTER)
    cell_s  = ParagraphStyle("c", parent=styles["Normal"], fontSize=7)

    story = [
        Paragraph("GTTC Robotics Training Portal", title_s),
        Paragraph(f"Monthly Report — {academic_year}", sub_s),
        Spacer(1, 0.3*cm),
    ]
    headers = ["School","UDISE","District","Month","Students","Community","Workshops","Projects","Won","Status"]
    data = [headers] + [
        [Paragraph(r.school.name[:28] if r.school else "", cell_s),
         r.school.udise_code if r.school else "",
         r.school.district if r.school else "",
         f"{MONTH_NAMES[r.report_month][:3]} {r.report_year}" if r.report_month else "",
         str(r.students_school), str(r.students_community),
         str(r.workshops_count), str(r.innovation_projects),
         str(r.atl_competitions_won + r.other_competitions_won),
         r.status.value.capitalize()]
        for r in reports
    ]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0),  8),
        ("FONTSIZE",     (0,1), (-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#EEF4FF")]),
        ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]))
    latest_by_school_pdf = {}
    for r in reports:
        ym = r.report_year * 100 + r.report_month
        prev = latest_by_school_pdf.get(r.school_id)
        if prev is None or ym > prev[0]:
            latest_by_school_pdf[r.school_id] = (ym, r)
    latest_pdf = [v[1] for v in latest_by_school_pdf.values()]

    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Total Schools: {len(latest_by_school_pdf)} | "
        f"Total Students: {sum((r.students_school or 0)+(r.students_community or 0) for r in latest_pdf):,} | "
        f"Total Projects: {sum(r.innovation_projects or 0 for r in latest_pdf):,} | "
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_s))
    doc.build(story)
    buf.seek(0)
    fname = f"gttc_{academic_year.replace('-','_')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ════════════════════════════════════════════════════════
# SYSTEM STATS + ACTIVITY LOG
# ════════════════════════════════════════════════════════

@router.get("/system-stats")
def system_stats(db: Session = Depends(get_db), _=Depends(require_state_admin)):
    return {
        "total_users":       db.query(User).count(),
        "active_users":      db.query(User).filter(User.is_active == True).count(),
        "total_schools":     db.query(School).filter(School.is_active == True).count(),
        "total_divisions":   db.query(Division).count(),
        "total_subs":        db.query(SubDivision).count(),
        "total_reports":     db.query(MonthlyReport).count(),
        "submitted_reports": db.query(MonthlyReport).filter(MonthlyReport.status == "submitted").count(),
        "total_logs":        db.query(ActivityLog).count(),
    }


@router.get("/activity-log")
def full_activity_log(
    skip: int = Query(0, ge=0), limit: int = Query(100, le=500),
    action: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
    db: Session = Depends(get_db), _=Depends(require_sub_and_above),
):
    q = db.query(ActivityLog)
    if action:  q = q.filter(ActivityLog.action == action)
    if user_id: q = q.filter(ActivityLog.user_id == user_id)
    logs = q.order_by(ActivityLog.created_at.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id": log.id, "user_name": log.user.name if log.user else "System",
            "user_role": log.user.role.value if log.user else "-",
            "action": log.action, "model_type": log.model_type, "model_id": log.model_id,
            "description": log.description, "ip_address": log.ip_address, "created_at": log.created_at,
        }
        for log in logs
    ]


@router.delete("/activity-log/{log_id}")
def delete_activity_log(
    log_id: int,
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    log = db.query(ActivityLog).filter(ActivityLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    db.delete(log)
    db.commit()
    return {"ok": True}


@router.delete("/activity-log")
def clear_all_activity_logs(
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    db.query(ActivityLog).delete()
    db.commit()
    return {"ok": True, "message": "All notifications cleared"}


# ════════════════════════════════════════════════════════
# BULK UPLOAD
# ════════════════════════════════════════════════════════

@router.get("/bulk-uploads")
def list_bulk_uploads(
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    uploads = db.query(BulkUpload).order_by(BulkUpload.created_at.desc()).limit(limit).all()
    return [
        {
            "id": u.id, "filename": u.file_name, "upload_type": u.record_type,
            "total_rows": u.total_rows, "success_rows": u.success_rows,
            "failed_rows": u.failed_rows, "status": u.status,
            "created_at": u.created_at,
        }
        for u in uploads
    ]


@router.get("/bulk-template/{upload_type}")
def download_template(upload_type: str, _=Depends(require_state_admin)):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    if upload_type == "schools":
        ws.title = "Schools"
        headers = [
            "udise_code", "atl_lab_code", "name", "district", "pin_code",
            "division_id", "sub_division_id", "school_type", "lab_type",
            "max_grade", "education_type", "address", "principal_name",
            "principal_email", "principal_phone", "lab_incharge_name",
            "lab_area_sqft", "lab_launch_date",
            "trainer_name", "trainer_email", "trainer_phone",
            "trainer_gender", "trainer_caste", "trainer_dob",
        ]
        ws.append(headers)
        ws.append([
            "29XXXXXXXXXX", "ATL-KA-001", "Government High School Example",
            "Mysore", "570001", "1", "1", "government", "ATL", "10",
            "secondary", "Main Road, Mysore", "Dr. Principal Name",
            "principal@school.edu", "+91 9876543210", "Lab Incharge Name",
            "1200", "2023-01-15",
            "Trainer Full Name", "trainer@gttc.gov.in", "+91 9876543211",
            "male", "general", "1990-05-20",
        ])
    elif upload_type == "trainers":
        ws.title = "Trainers"
        headers = [
            "school_udise_code", "trainer_name", "trainer_email",
            "trainer_phone", "trainer_gender", "trainer_caste", "trainer_dob",
        ]
        ws.append(headers)
        ws.append([
            "29XXXXXXXXXX", "Trainer Full Name", "trainer@gttc.gov.in",
            "+91 9876543211", "male", "general", "1990-05-20",
        ])
    else:
        from fastapi import HTTPException as _H
        raise _H(400, f"Unknown template type: {upload_type}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="template_{upload_type}.xlsx"'},
    )


@router.post("/bulk-upload/schools")
async def bulk_upload_schools(
    file: UploadFile = File(...),
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    import openpyxl
    content = await file.read()
    buf = BytesIO(content)
    try:
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Cannot parse file: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx]
            return str(v).strip() if v is not None else ""
        except (ValueError, IndexError):
            return ""

    success, failed, errors = 0, 0, []
    for i, row in enumerate(data_rows, start=2):
        udise = col(row, "udise_code")
        name  = col(row, "name")
        if not udise or not name:
            failed += 1; errors.append(f"Row {i}: missing udise_code or name"); continue
        try:
            div_id = int(col(row, "division_id") or 1)
            sub_id = int(col(row, "sub_division_id") or 1)
        except ValueError:
            div_id, sub_id = 1, 1
        if db.query(School).filter(School.udise_code == udise).first():
            failed += 1; errors.append(f"Row {i}: UDISE {udise} already exists"); continue
        trainer_email = col(row, "trainer_email")
        if trainer_email and db.query(User).filter(User.email == trainer_email).first():
            failed += 1; errors.append(f"Row {i}: Trainer email {trainer_email} already registered"); continue
        try:
            from datetime import date as _date
            lab_dt = col(row, "lab_launch_date")
            lab_date = _date.fromisoformat(lab_dt) if lab_dt else None
            area_raw = col(row, "lab_area_sqft")
            area = int(area_raw) if area_raw else None
            grade_raw = col(row, "max_grade")
            grade = int(grade_raw) if grade_raw else 10
            school = School(
                udise_code=udise, name=name,
                district=col(row, "district") or "Unknown",
                pin_code=col(row, "pin_code") or "000000",
                division_id=div_id, sub_division_id=sub_id,
                atl_lab_code=col(row, "atl_lab_code") or None,
                school_type=col(row, "school_type") or "government",
                lab_type=col(row, "lab_type") or "atl",
                education_type=col(row, "education_type") or "secondary",
                max_grade=grade, address=col(row, "address") or None,
                principal_name=col(row, "principal_name") or None,
                principal_email=col(row, "principal_email") or None,
                principal_phone=col(row, "principal_phone") or None,
                lab_incharge_name=col(row, "lab_incharge_name") or None,
                lab_area_sqft=area, lab_launch_date=lab_date,
            )
            db.add(school); db.flush()
            if trainer_email:
                tname = col(row, "trainer_name") or "Trainer"
                trainer = User(
                    name=tname, email=trainer_email,
                    phone=col(row, "trainer_phone") or None,
                    gender=col(row, "trainer_gender") or None,
                    caste=col(row, "trainer_caste") or None,
                    hashed_password=hash_password("Pass@1234"),
                    role=RoleEnum.atl_trainer,
                    division_id=div_id, sub_division_id=sub_id,
                    is_active=True, must_change_password=True,
                )
                db.add(trainer); db.flush()
                db.add(SchoolTrainer(
                    school_id=school.id, user_id=trainer.id,
                    assigned_from=date.today(), is_current=True,
                ))
            db.commit()
            success += 1
        except Exception as e:
            db.rollback()
            failed += 1; errors.append(f"Row {i}: {e}")

    upload_rec = BulkUpload(
        uploaded_by=current_user.id, file_name=file.filename,
        record_type="schools", total_rows=len(data_rows),
        success_rows=success, failed_rows=failed,
        error_log="\n".join(errors) if errors else None,
        status="completed" if failed == 0 else "partial",
    )
    db.add(upload_rec); db.commit()
    ActivityLog.log(db, current_user.id, "bulk_upload",
        description=f"Bulk school upload: {success} created, {failed} failed")
    db.commit()
    return {"success": success, "failed": failed, "errors": errors[:20]}


@router.post("/bulk-upload/trainers")
async def bulk_upload_trainers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    import openpyxl
    content = await file.read()
    buf = BytesIO(content)
    try:
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Cannot parse file: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx]
            return str(v).strip() if v is not None else ""
        except (ValueError, IndexError):
            return ""

    success, failed, errors = 0, 0, []
    for i, row in enumerate(data_rows, start=2):
        udise = col(row, "school_udise_code")
        email = col(row, "trainer_email")
        if not udise or not email:
            failed += 1; errors.append(f"Row {i}: missing school_udise_code or trainer_email"); continue
        school = db.query(School).filter(School.udise_code == udise).first()
        if not school:
            failed += 1; errors.append(f"Row {i}: school UDISE {udise} not found"); continue
        if db.query(User).filter(User.email == email).first():
            failed += 1; errors.append(f"Row {i}: email {email} already registered"); continue
        try:
            trainer = User(
                name=col(row, "trainer_name") or "Trainer",
                email=email,
                phone=col(row, "trainer_phone") or None,
                gender=col(row, "trainer_gender") or None,
                caste=col(row, "trainer_caste") or None,
                hashed_password=hash_password("Pass@1234"),
                role=RoleEnum.atl_trainer,
                division_id=school.division_id,
                sub_division_id=school.sub_division_id,
                is_active=True, must_change_password=True,
            )
            db.add(trainer); db.flush()
            db.add(SchoolTrainer(
                school_id=school.id, user_id=trainer.id,
                assigned_from=date.today(), is_current=True,
            ))
            db.commit()
            success += 1
        except Exception as e:
            db.rollback()
            failed += 1; errors.append(f"Row {i}: {e}")

    upload_rec = BulkUpload(
        uploaded_by=current_user.id, file_name=file.filename,
        record_type="trainers", total_rows=len(data_rows),
        success_rows=success, failed_rows=failed,
        error_log="\n".join(errors) if errors else None,
        status="completed" if failed == 0 else "partial",
    )
    db.add(upload_rec); db.commit()
    ActivityLog.log(db, current_user.id, "bulk_upload",
        description=f"Bulk trainer upload: {success} assigned, {failed} failed")
    db.commit()
    return {"success": success, "failed": failed, "errors": errors[:20]}


# ════════════════════════════════════════════════════════
# ATL COMPETITIONS
# ════════════════════════════════════════════════════════

@router.get("/atl-competitions")
def list_atl_competitions(
    skip: int = Query(0, ge=0), limit: int = Query(50, le=200),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    rows = db.query(ATLCompetition)\
        .order_by(ATLCompetition.competition_date.desc())\
        .offset(skip).limit(limit).all()
    return [
        {
            "id":                         r.id,
            "competition_date":           str(r.competition_date),
            "venue_school":               r.venue_school,
            "sub_district":               r.sub_district,
            "division":                   r.division.name if r.division else None,
            "atl_teams_participated":     r.atl_teams_participated,
            "other_teams_participated":   r.other_teams_participated,
            "atl_teams_not_participated": r.atl_teams_not_participated,
            "atl_teams_won":              r.atl_teams_won,
            "others_won":                 r.others_won,
            "submitted_by":               r.submitter.name if r.submitter else None,
        }
        for r in rows
    ]


@router.post("/atl-competitions")
def create_atl_competition(
    payload: dict,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from datetime import date as date_cls
    comp = ATLCompetition(
        competition_date          = date_cls.fromisoformat(payload["competition_date"]),
        venue_school              = payload.get("venue_school", ""),
        sub_district              = payload.get("sub_district", ""),
        division_id               = payload.get("division_id"),
        atl_teams_participated    = int(payload.get("atl_teams_participated", 0)),
        other_teams_participated  = int(payload.get("other_teams_participated", 0)),
        atl_teams_not_participated= int(payload.get("atl_teams_not_participated", 0)),
        atl_teams_won             = int(payload.get("atl_teams_won", 0)),
        others_won                = int(payload.get("others_won", 0)),
        submitted_by              = current_user.id,
    )
    db.add(comp)
    db.commit()
    db.refresh(comp)
    ActivityLog.log(db, current_user.id, "report_submitted",
        description=f"ATL Competition recorded: {comp.venue_school} on {comp.competition_date}")
    db.commit()
    return {"id": comp.id, "ok": True}


@router.delete("/atl-competitions/{comp_id}")
def delete_atl_competition(
    comp_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    if current_user.role not in (RoleEnum.state_admin, RoleEnum.division_master):
        raise HTTPException(403, "Not authorised")
    comp = db.query(ATLCompetition).filter(ATLCompetition.id == comp_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="Not found")
    if current_user.role == RoleEnum.division_master and comp.division_id != current_user.division_id:
        raise HTTPException(403, "Cannot delete competitions from other divisions")
    db.delete(comp)
    db.commit()
    return {"ok": True}


@router.get("/highlights")
def list_highlights(
    academic_year: str = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.services.aggregator import get_scoped_reports
    if not academic_year:
        latest_row = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest_row[0] if latest_row else "2025-26"
    rows = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.highlight_of_month != None,
        MonthlyReport.highlight_of_month != "",
        MonthlyReport.status == ReportStatusEnum.submitted,
    ).order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).all()
    return [
        {
            "school":    r.school.name if r.school else "",
            "district":  r.school.district if r.school else "",
            "month":     r.report_month,
            "year":      r.report_year,
            "highlight": r.highlight_of_month,
        }
        for r in rows
    ]


# ════════════════════════════════════════════════════════
# PENDING SUBMISSIONS
# ════════════════════════════════════════════════════════

@router.get("/pending-submissions")
def pending_submissions(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.services.aggregator import get_scoped_schools
    today = date.today()
    target_year  = year  or today.year
    target_month = month or today.month

    schools = get_scoped_schools(db, current_user).all()
    submitted_ids = {
        r.school_id
        for r in db.query(MonthlyReport.school_id).filter(
            MonthlyReport.report_year  == target_year,
            MonthlyReport.report_month == target_month,
            MonthlyReport.status       == ReportStatusEnum.submitted,
        ).all()
    }

    result = []
    for s in schools:
        if s.id in submitted_ids:
            continue
        last = (
            db.query(MonthlyReport)
            .filter(MonthlyReport.school_id == s.id, MonthlyReport.status == ReportStatusEnum.submitted)
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc())
            .first()
        )
        last_date = None
        if last:
            last_date = date(last.report_year, last.report_month, 1)

        deadline   = date(target_year, target_month, 10)
        days_overdue = max(0, (today - deadline).days) if today > deadline else 0

        result.append({
            "school_id":    s.id,
            "school_name":  s.name,
            "district":     s.district or "",
            "division_id":  s.division_id,
            "last_submitted": last_date.isoformat() if last_date else None,
            "days_overdue": days_overdue,
        })

    result.sort(key=lambda x: x["days_overdue"], reverse=True)
    return result


@router.post("/pending-submissions/{school_id}/remind")
def send_reminder(
    school_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(404, "School not found")

    trainers = (
        db.query(SchoolTrainer)
        .filter(SchoolTrainer.school_id == school_id, SchoolTrainer.is_current == True)
        .all()
    )
    today = date.today()
    for t in trainers:
        notif = Notification(
            user_id   = t.user_id,
            title     = "Reminder: Monthly Report Pending",
            body      = f"Please submit the report for {MONTH_NAMES[today.month]} {today.year}.",
            notif_type= "pending_report",
            link_page = "reports",
        )
        db.add(notif)

    db.add(ActivityLog(
        user_id = current_user.id,
        action  = f"Sent reminder to school {school.name} (id={school_id})",
    ))
    db.commit()
    return {"ok": True, "reminders_sent": len(trainers)}


# ════════════════════════════════════════════════════════
# NOTIFICATIONS
# ════════════════════════════════════════════════════════

@router.get("/notifications")
def list_notifications(
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    notifs = (
        db.query(Notification)
        .filter(Notification.user_id == current_user.id)
        .order_by(Notification.created_at.desc())
        .limit(50)
        .all()
    )
    unread = sum(1 for n in notifs if not n.is_read)
    return {
        "unread": unread,
        "items": [
            {
                "id":         n.id,
                "title":      n.title,
                "body":       n.body,
                "notif_type": n.notif_type,
                "link_page":  n.link_page,
                "is_read":    n.is_read,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            }
            for n in notifs
        ],
    }


@router.patch("/notifications/{notif_id}/read")
def mark_notification_read(
    notif_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    n = db.query(Notification).filter(
        Notification.id == notif_id,
        Notification.user_id == current_user.id,
    ).first()
    if not n:
        raise HTTPException(404, "Not found")
    n.is_read = True
    db.commit()
    return {"ok": True}


@router.patch("/notifications/read-all")
def mark_all_notifications_read(
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.is_read == False,
    ).update({"is_read": True})
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════
# CALENDAR EVENTS
# ════════════════════════════════════════════════════════

@router.get("/calendar-events")
def list_calendar_events(
    year: Optional[int]  = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    q = db.query(CalendarEvent)
    if year:  q = q.filter(CalendarEvent.event_date >= date(year, 1, 1),
                            CalendarEvent.event_date <= date(year, 12, 31))
    if month and year:
        import calendar as _cal
        last_day = _cal.monthrange(year, month)[1]
        q = q.filter(CalendarEvent.event_date >= date(year, month, 1),
                     CalendarEvent.event_date <= date(year, month, last_day))
    events = q.order_by(CalendarEvent.event_date).all()
    return [
        {
            "id":               e.id,
            "title":            e.title,
            "description":      e.description,
            "event_date":       e.event_date.isoformat(),
            "end_date":         e.end_date.isoformat() if e.end_date else None,
            "event_type":       e.event_type,
            "division_id":      e.division_id,
            "is_all_divisions": e.is_all_divisions,
        }
        for e in events
    ]


@router.post("/calendar-events")
def create_calendar_event(
    body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    ev = CalendarEvent(
        title            = body.get("title"),
        description      = body.get("description"),
        event_date       = date.fromisoformat(body["event_date"]),
        end_date         = date.fromisoformat(body["end_date"]) if body.get("end_date") else None,
        event_type       = body.get("event_type", "event"),
        division_id      = body.get("division_id"),
        is_all_divisions = body.get("is_all_divisions", True),
        created_by       = current_user.id,
    )
    db.add(ev)
    db.commit()
    db.refresh(ev)
    return {"id": ev.id, "ok": True}


@router.put("/calendar-events/{event_id}")
def update_calendar_event(
    event_id: int, body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    ev = db.query(CalendarEvent).filter(CalendarEvent.id == event_id).first()
    if not ev:
        raise HTTPException(404, "Event not found")
    for field in ("title", "description", "event_type", "division_id", "is_all_divisions"):
        if field in body:
            setattr(ev, field, body[field])
    if "event_date" in body:
        ev.event_date = date.fromisoformat(body["event_date"])
    if "end_date" in body:
        ev.end_date = date.fromisoformat(body["end_date"]) if body["end_date"] else None
    db.commit()
    return {"ok": True}


@router.delete("/calendar-events/{event_id}")
def delete_calendar_event(
    event_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    ev = db.query(CalendarEvent).filter(CalendarEvent.id == event_id).first()
    if not ev:
        raise HTTPException(404, "Event not found")
    db.delete(ev)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════
# EQUIPMENT INVENTORY
# ════════════════════════════════════════════════════════

@router.get("/equipment")
def list_equipment(
    school_id: int = Query(...),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    items = db.query(EquipmentInventory).filter(
        EquipmentInventory.school_id == school_id
    ).order_by(EquipmentInventory.item_name).all()
    return [
        {
            "id": i.id, "school_id": i.school_id, "item_name": i.item_name,
            "quantity": i.quantity, "condition": i.condition,
            "last_checked": i.last_checked.isoformat() if i.last_checked else None,
            "notes": i.notes,
            "added_by": i.added_by,
            "created_at": i.created_at.isoformat() if i.created_at else None,
        }
        for i in items
    ]


@router.post("/equipment")
def add_equipment(
    body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    school_id = body.get("school_id")
    if not school_id:
        raise HTTPException(400, "school_id required")
    last_checked = None
    if body.get("last_checked"):
        try:
            last_checked = date.fromisoformat(body["last_checked"])
        except ValueError:
            pass
    item = EquipmentInventory(
        school_id=school_id,
        item_name=body.get("item_name", "").strip(),
        quantity=int(body.get("quantity", 1)),
        condition=body.get("condition", "good"),
        last_checked=last_checked,
        notes=body.get("notes", "").strip() or None,
        added_by=current_user.id,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id, "ok": True}


@router.put("/equipment/{item_id}")
def update_equipment(
    item_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    item = db.query(EquipmentInventory).filter(EquipmentInventory.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    for field in ("item_name", "quantity", "condition", "notes"):
        if field in body and body[field] is not None:
            setattr(item, field, body[field])
    if body.get("last_checked"):
        try:
            item.last_checked = date.fromisoformat(body["last_checked"])
        except ValueError:
            pass
    db.commit()
    return {"id": item.id, "ok": True}


@router.delete("/equipment/{item_id}")
def delete_equipment(
    item_id: int,
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    item = db.query(EquipmentInventory).filter(EquipmentInventory.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════
# DIVISION COMPARISON
# ════════════════════════════════════════════════════════

@router.get("/division-comparison")
def division_comparison(
    academic_year: Optional[str] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.services.aggregator import get_scoped_reports
    if not academic_year:
        latest = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest[0] if latest else "2025-26"

    reports = get_scoped_reports(db, current_user).filter(
        MonthlyReport.academic_year == academic_year,
        MonthlyReport.status == ReportStatusEnum.submitted,
    ).all()

    # Reports are running YTD totals — keep only the latest report per school
    # so totals reconcile with the dashboard KPI cards (no month-over-month double counting).
    latest_by_school = {}
    for r in reports:
        ym = r.report_year * 100 + r.report_month
        prev = latest_by_school.get(r.school_id)
        if prev is None or ym > prev[0]:
            latest_by_school[r.school_id] = (ym, r)
    reports = [v[1] for v in latest_by_school.values()]

    divisions = db.query(Division).all()
    div_map = {d.id: d.name for d in divisions}

    school_div = {}
    for s in db.query(School).all():
        school_div[s.id] = s.division_id

    buckets: dict = {}
    for r in reports:
        div_id = school_div.get(r.school_id)
        if div_id not in buckets:
            buckets[div_id] = {
                "division_id":   div_id,
                "division_name": div_map.get(div_id, "Unknown"),
                "schools_submitted": set(),
                "students":    0,
                "workshops":   0,
                "mentoring":   0,
                "projects":    0,
                "competitions_won": 0,
                "industrial_visits": 0,
            }
        b = buckets[div_id]
        b["schools_submitted"].add(r.school_id)
        b["students"]    += r.total_students
        b["workshops"]   += r.workshops_count
        b["mentoring"]   += r.mentoring_sessions
        b["projects"]    += r.innovation_projects
        b["competitions_won"] += r.total_won
        b["industrial_visits"] += (r.industrial_visits or 0)

    total_schools_by_div = {}
    for s in db.query(School).all():
        d = s.division_id
        total_schools_by_div[d] = total_schools_by_div.get(d, 0) + 1

    result = []
    for div_id, b in buckets.items():
        submitted = len(b["schools_submitted"])
        total     = total_schools_by_div.get(div_id, 0)
        result.append({
            "division_id":       b["division_id"],
            "division_name":     b["division_name"],
            "total_schools":     total,
            "schools_submitted": submitted,
            "submission_rate":   round(submitted / total * 100) if total else 0,
            "students":          b["students"],
            "workshops":         b["workshops"],
            "mentoring":         b["mentoring"],
            "projects":          b["projects"],
            "competitions_won":  b["competitions_won"],
            "industrial_visits": b["industrial_visits"],
        })

    result.sort(key=lambda x: x["students"], reverse=True)
    return {"academic_year": academic_year, "divisions": result}


# ════════════════════════════════════════════════════════
# TRAINER PERFORMANCE STATS
# ════════════════════════════════════════════════════════

@router.get("/trainer-stats")
def trainer_stats(
    academic_year: Optional[str] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.models.hierarchy import SchoolTrainer
    from app.models.users import RoleEnum as RE

    if not academic_year:
        latest = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest[0] if latest else "2025-26"

    trainers = db.query(User).filter(User.role == RE.atl_trainer, User.is_active == True).all()

    result = []
    for t in trainers:
        links = db.query(SchoolTrainer).filter(
            SchoolTrainer.user_id == t.id, SchoolTrainer.is_current == True,
        ).all()
        school_ids = [l.school_id for l in links]
        school_names = [db.query(School).filter(School.id == sid).first() for sid in school_ids]
        school_name = school_names[0].name if school_names and school_names[0] else "—"

        submitted = db.query(MonthlyReport).filter(
            MonthlyReport.school_id.in_(school_ids),
            MonthlyReport.academic_year == academic_year,
            MonthlyReport.status == ReportStatusEnum.submitted,
        ).count() if school_ids else 0

        last_report = db.query(MonthlyReport).filter(
            MonthlyReport.school_id.in_(school_ids),
            MonthlyReport.status == ReportStatusEnum.submitted,
        ).order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first() if school_ids else None

        # Streak: consecutive months submitted (current year, going back from latest)
        from datetime import date as dt
        now = dt.today()
        streak = 0
        if school_ids:
            for m in range(now.month, 0, -1):
                found = db.query(MonthlyReport).filter(
                    MonthlyReport.school_id.in_(school_ids),
                    MonthlyReport.report_year == now.year,
                    MonthlyReport.report_month == m,
                    MonthlyReport.status == ReportStatusEnum.submitted,
                ).first()
                if found: streak += 1
                else: break

        result.append({
            "trainer_id":   t.id,
            "trainer_name": t.name,
            "email":        t.email,
            "school_name":  school_name,
            "schools_count":len(school_ids),
            "months_submitted": submitted,
            "streak":       streak,
            "last_submitted": f"{last_report.report_year}-{last_report.report_month:02d}" if last_report else None,
            "rating": "excellent" if submitted >= 9 else "good" if submitted >= 6 else "fair" if submitted >= 3 else "poor",
        })

    result.sort(key=lambda x: x["months_submitted"], reverse=True)
    return {"academic_year": academic_year, "trainers": result}


# ════════════════════════════════════════════════════════
# SCHOOL PROFILE COMPLETENESS
# ════════════════════════════════════════════════════════

def _school_completeness(s, has_trainer: bool, has_principal: bool) -> int:
    fields = [
        bool(s.udise_code),
        bool(s.atl_lab_code),
        bool(s.lab_type),
        bool(s.lab_area_sqft),
        bool(s.district),
        bool(s.division_id),
        has_trainer,
        has_principal,
    ]
    return round(sum(fields) / len(fields) * 100)


# ════════════════════════════════════════════════════════
# GLOBAL SEARCH
# ════════════════════════════════════════════════════════

@router.get("/search")
def global_search(
    q: str = Query(..., min_length=2),
    limit: int = Query(8, le=20),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.services.aggregator import get_scoped_reports, get_scoped_schools
    term = f"%{q}%"

    schools = get_scoped_schools(db, current_user).filter(
        (School.name.ilike(term)) | (School.udise_code.ilike(term)) | (School.district.ilike(term))
    ).limit(limit).all()

    users = db.query(User).filter(
        User.is_active == True,
        (User.name.ilike(term)) | (User.email.ilike(term)),
    ).limit(limit).all()

    highlights = get_scoped_reports(db, current_user).filter(
        MonthlyReport.highlight_of_month.ilike(term),
        MonthlyReport.status == ReportStatusEnum.submitted,
    ).order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).limit(5).all()

    return {
        "query": q,
        "schools": [{"id": s.id, "name": s.name, "district": s.district or "",
                     "lab_type": s.lab_type.value if s.lab_type else ""} for s in schools],
        "users":   [{"id": u.id, "name": u.name, "email": u.email, "role": u.role.value} for u in users],
        "highlights": [{"report_id": r.id, "school": r.school.name if r.school else "",
                        "month": r.report_month, "year": r.report_year,
                        "text": r.highlight_of_month} for r in highlights],
    }


# ════════════════════════════════════════════════════════
# DIVISION KPI TARGETS
# ════════════════════════════════════════════════════════

@router.get("/division-targets")
def list_division_targets(
    academic_year: Optional[str] = Query(None),
    db: Session = Depends(get_db), current_user=Depends(require_any),
):
    from app.models.reports import DivisionTarget
    if not academic_year:
        latest = db.query(MonthlyReport.academic_year)\
            .order_by(MonthlyReport.report_year.desc(), MonthlyReport.report_month.desc()).first()
        academic_year = latest[0] if latest else "2025-26"
    targets = db.query(DivisionTarget).filter(DivisionTarget.academic_year == academic_year).all()
    return [
        {
            "id":               t.id,
            "division_id":      t.division_id,
            "division_name":    t.division.name if t.division else "",
            "academic_year":    t.academic_year,
            "target_students":  t.target_students,
            "target_workshops": t.target_workshops,
            "target_projects":  t.target_projects,
            "target_wins":      t.target_wins,
            "target_industrial":t.target_industrial,
            "target_atl_comp":   t.target_atl_comp,
            "target_other_comp": t.target_other_comp,
            "target_mentoring":  t.target_mentoring,
            "target_exhibitions":t.target_exhibitions,
        }
        for t in targets
    ]


@router.put("/division-targets/{division_id}")
def set_division_target(
    division_id: int,
    body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_sub_and_above),
):
    from app.models.reports import DivisionTarget
    academic_year = body.get("academic_year", "2025-26")
    t = db.query(DivisionTarget).filter(
        DivisionTarget.division_id == division_id,
        DivisionTarget.academic_year == academic_year,
    ).first()
    fields = ("target_students","target_workshops","target_projects","target_wins","target_industrial",
              "target_atl_comp","target_other_comp","target_mentoring","target_exhibitions")
    if t:
        for f in fields:
            if f in body: setattr(t, f, body[f])
    else:
        t = DivisionTarget(
            division_id   = division_id,
            academic_year = academic_year,
            created_by    = current_user.id,
            **{f: body.get(f, 0) for f in fields},
        )
        db.add(t)

    # ── Notify the master trainers of this division about the new targets ──
    div = db.query(Division).filter(Division.id == division_id).first()
    div_name = div.name if div else f"Division #{division_id}"
    summary = (
        f"{t.target_students} students, {t.target_workshops} workshops, "
        f"{t.target_projects} projects, {t.target_wins} wins, "
        f"{t.target_industrial} industrial visits"
    )
    recipients = (
        db.query(User)
        .filter(
            User.role == RoleEnum.master_trainer,
            User.division_id == division_id,
            User.is_active == True,
        )
        .all()
    )
    for u in recipients:
        db.add(Notification(
            user_id    = u.id,
            title      = f"New KPI Targets for {div_name} ({academic_year})",
            body       = f"{current_user.name} set your division's targets: {summary}.",
            notif_type = "target_set",
            link_page  = "dashboard",
        ))

    db.add(ActivityLog(
        user_id     = current_user.id,
        action      = "Set division KPI targets",
        model_type  = "DivisionTarget",
        model_id    = division_id,
        description = f"Targets for {div_name} ({academic_year}) → {summary}; notified {len(recipients)} master trainer(s)",
    ))
    db.commit()
    return {"ok": True, "notified": len(recipients)}


# ════════════════════════════════════════════════════════
# GLOBAL SETTINGS (academic year, etc.)
# ════════════════════════════════════════════════════════

@router.get("/settings")
def get_settings(
    db: Session = Depends(get_db), _=Depends(require_sub_and_above),
):
    from app.core.settings_store import (
        get_setting, get_current_academic_year,
        ACADEMIC_YEAR_KEY, REPORT_DEADLINE_KEY,
    )
    return {
        "academic_year": get_current_academic_year(db),
        "report_deadline_day": get_setting(db, REPORT_DEADLINE_KEY, "10"),
    }


@router.put("/settings")
def update_settings(
    body: dict = Body(...),
    db: Session = Depends(get_db), current_user=Depends(require_state_admin),
):
    from app.core.settings_store import set_setting, ACADEMIC_YEAR_KEY, REPORT_DEADLINE_KEY
    ay = (body.get("academic_year") or "").strip()
    if ay:
        set_setting(db, ACADEMIC_YEAR_KEY, ay)
    if body.get("report_deadline_day") is not None:
        set_setting(db, REPORT_DEADLINE_KEY, str(body.get("report_deadline_day")))
    db.add(ActivityLog(
        user_id     = current_user.id,
        action      = "Updated global settings",
        description = f"academic_year={ay or '(unchanged)'}",
    ))
    db.commit()
    from app.core.settings_store import get_current_academic_year
    return {"ok": True, "academic_year": get_current_academic_year(db)}


# ════════════════════════════════════════════════════════
# DATABASE BACKUP
# ════════════════════════════════════════════════════════

@router.post("/backup")
def create_backup(_=Depends(require_state_admin)):
    """Make a consistent copy of the SQLite database into ./backups."""
    import os, sqlite3, datetime
    src = "gttc_portal.db"
    if not os.path.exists(src):
        raise HTTPException(404, "Database file not found")
    os.makedirs("backups", exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join("backups", f"gttc_portal_{ts}.db")
    s = sqlite3.connect(src)
    d = sqlite3.connect(dest)
    try:
        with d:
            s.backup(d)          # consistent online backup
    finally:
        s.close(); d.close()
    return {
        "ok": True,
        "file": os.path.basename(dest),
        "size_kb": round(os.path.getsize(dest) / 1024, 1),
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }


@router.get("/backups")
def list_backups(_=Depends(require_state_admin)):
    """List existing database backups (most recent first)."""
    import os, datetime
    folder = "backups"
    if not os.path.isdir(folder):
        return []
    out = []
    for f in sorted(os.listdir(folder), reverse=True):
        if f.endswith(".db"):
            path = os.path.join(folder, f)
            out.append({
                "file": f,
                "size_kb": round(os.path.getsize(path) / 1024, 1),
                "created_at": datetime.datetime.fromtimestamp(os.path.getmtime(path)).isoformat(timespec="seconds"),
            })
    return out


# ════════════════════════════════════════════════════════
# REMINDER STATUS
# ════════════════════════════════════════════════════════

@router.post("/trigger-reminders")
def trigger_reminders(
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    from app.main import send_pending_report_reminders
    try:
        send_pending_report_reminders()
        return {"ok": True, "message": "Reminders sent successfully"}
    except Exception as e:
        raise HTTPException(500, f"Reminder job failed: {e}")


@router.get("/reminder-status")
def reminder_status(
    db: Session = Depends(get_db), _=Depends(require_state_admin),
):
    # Last auto-reminder notification created by the system (no user linked = system-generated)
    last = db.query(Notification).filter(
        Notification.notif_type == "pending_report",
        Notification.title == "Monthly Report Reminder",
    ).order_by(Notification.created_at.desc()).first()

    from app import main as _main
    jobs = []
    try:
        for j in _main._scheduler.get_jobs():
            next_run = j.next_run_time
            jobs.append({
                "id": j.id,
                "next_run": next_run.isoformat() if next_run else None,
            })
    except Exception:
        pass

    return {
        "last_reminder_at": last.created_at.isoformat() if last else None,
        "scheduler_jobs": jobs,
    }