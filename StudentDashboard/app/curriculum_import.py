from pathlib import Path

import openpyxl

from app.models import Course, Lesson


CURRICULUM_COURSE_TITLE = "ATL Curriculum and Innovation Calendar 2026-27"

REFERENCE_FILES = [
    {
        "title": "ATL Handbook Volume I - Learning by Doing",
        "content_type": "pdf",
        "resource_url": "/static/uploads/curriculum/atl-handbook-vol-i-karnataka.pdf",
        "body": "Reference handbook Volume I for ATL activities and learning-by-doing guidance.",
    },
    {
        "title": "ATL Handbook Volume II - Grades 6-8",
        "content_type": "pdf",
        "resource_url": "/static/uploads/curriculum/atl-handbook-vol-ii-karnataka.pdf",
        "body": "Reference handbook Volume II for integrating school curricula with ATL for Grades 6-8.",
    },
    {
        "title": "ATL Handbook Volume III - Grades 9-10",
        "content_type": "pdf",
        "resource_url": "/static/uploads/curriculum/atl-handbook-vol-iii-karnataka.pdf",
        "body": "Reference handbook Volume III for integrating school curricula with ATL for Grades 9-10.",
    },
    {
        "title": "ATL Innovation Odyssey Calendar 2026",
        "content_type": "pdf",
        "resource_url": "/static/uploads/curriculum/atl-calendar-2026.pdf",
        "body": "ATL Innovation Odyssey calendar for annual planning and innovation milestones.",
    },
    {
        "title": "ATL Curriculum Activity Plan Workbook 2026-27",
        "content_type": "xlsx",
        "resource_url": "/static/uploads/curriculum/atl-curriculum-2026-27.xlsx",
        "body": "Original Excel workbook used to import the monthly ATL curriculum activity plan.",
    },
]


def _curriculum_workbook_path():
    return Path(__file__).resolve().parent / "static" / "uploads" / "curriculum" / "atl-curriculum-2026-27.xlsx"


def _read_activity_plan():
    workbook_path = _curriculum_workbook_path()
    if not workbook_path.exists():
        return []

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook["Activity_Plan_2026_27"] if "Activity_Plan_2026_27" in workbook.sheetnames else workbook.active

    activities = []
    current_month = ""
    current_theme = ""
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        month = row[2] or current_month
        theme = row[3] or current_theme
        week = row[4]
        category = row[5]
        activity = row[6]
        outcome = row[7]

        if month:
            current_month = str(month).strip()
        if theme:
            current_theme = str(theme).strip()
        if not week and not category and not activity and not outcome:
            continue

        activities.append(
            {
                "month": current_month,
                "theme": current_theme,
                "week": str(week or "").strip(),
                "category": str(category or "").strip(),
                "activity": str(activity or "").strip(),
                "outcome": str(outcome or "").strip(),
            }
        )
    return activities


def seed_curriculum_content(db):
    course = db.query(Course).filter(Course.title == CURRICULUM_COURSE_TITLE).first()
    if not course:
        course = Course(
            title=CURRICULUM_COURSE_TITLE,
            description=(
                "Imported annual ATL curriculum plan for 2026-27 with month-wise activities, "
                "learning outcomes, handbooks, and calendar resources."
            ),
            level="ATL Curriculum 2026-27",
        )
        db.add(course)
        db.flush()

    if db.query(Lesson).filter(Lesson.course_id == course.id, Lesson.title == "Curriculum Overview").first():
        db.commit()
        return

    sort_order = 1
    db.add(
        Lesson(
            course_id=course.id,
            title="Curriculum Overview",
            content_type="article",
            content_body=(
                "This course organizes the ATL Curriculum and Innovation Calendar Integrated Plan "
                "for 2026-27. Use the month-wise lessons for weekly delivery planning and the reference "
                "resources for detailed activity instructions."
            ),
            sort_order=sort_order,
        )
    )
    sort_order += 1

    for reference in REFERENCE_FILES:
        db.add(
            Lesson(
                course_id=course.id,
                title=reference["title"],
                content_type=reference["content_type"],
                content_body=reference["body"],
                resource_url=reference["resource_url"],
                sort_order=sort_order,
            )
        )
        sort_order += 1

    for activity in _read_activity_plan():
        title_parts = [activity["month"], activity["week"], activity["category"]]
        title = " - ".join(part for part in title_parts if part)
        body = (
            f"Theme / DT Stage: {activity['theme']}\n"
            f"Activity Plan: {activity['activity']}\n\n"
            f"Learning Outcomes: {activity['outcome']}"
        )
        db.add(
            Lesson(
                course_id=course.id,
                title=title[:220] or "Curriculum Activity",
                content_type="curriculum_activity",
                content_body=body,
                resource_url="/static/uploads/curriculum/atl-curriculum-2026-27.xlsx",
                sort_order=sort_order,
            )
        )
        sort_order += 1

    db.commit()
