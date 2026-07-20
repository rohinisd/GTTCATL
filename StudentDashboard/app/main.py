import json
import mimetypes
import os
import re
import secrets
import urllib.error
import urllib.request
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote, urlencode, urlparse

import bcrypt
import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, inspect, or_, text

from app.database import Base, SessionLocal, engine
from app import models  # noqa: F401 - registers SQLAlchemy models
from app.content_seed import seed_lms_content
from app.curriculum_import import seed_curriculum_content
from app.forms_catalog import ATL_MER_FORMS
from app.handbook_import import VOLUME_COURSES, seed_handbook_volume_courses
from app.seed import DEFAULT_TRAINER_PASSWORD, seed_initial_data, seed_trainer_accounts
from app.student_options import (
    CATEGORY_OPTIONS,
    RELIGION_OPTIONS,
    category_display,
    normalize_category,
    normalize_religion,
)


APP_DIR = Path(__file__).resolve().parent

Base.metadata.create_all(bind=engine)


def _ensure_columns():
    inspector = inspect(engine)

    def add_missing(table_name, columns):
        try:
            existing = {column["name"] for column in inspector.get_columns(table_name)}
        except Exception:
            return

        with engine.begin() as connection:
            for name, definition in columns.items():
                if name not in existing:
                    connection.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {name} {definition}"))

    add_missing(
        "trainers",
        {
            "role": "VARCHAR(80) DEFAULT 'atl_trainer'",
            "gender": "VARCHAR(40)",
            "caste": "VARCHAR(40)",
            "division": "VARCHAR(120)",
            "districts": "TEXT",
            "assigned_school": "VARCHAR(220)",
            "is_active": "BOOLEAN DEFAULT TRUE",
        },
    )
    add_missing(
        "schools",
        {
            "atl_lab_code": "VARCHAR(60)",
            "state": "VARCHAR(120) DEFAULT 'Karnataka'",
            "pin_code": "VARCHAR(20)",
            "school_type": "VARCHAR(80) DEFAULT 'government'",
            "lab_type": "VARCHAR(80) DEFAULT 'atl'",
            "education_type": "VARCHAR(80) DEFAULT 'secondary'",
            "max_grade": "INTEGER DEFAULT 10",
            "principal_name": "VARCHAR(160)",
            "principal_email": "VARCHAR(180)",
            "principal_phone": "VARCHAR(40)",
            "lab_area_sqft": "INTEGER",
            "lab_launch_date": "DATE",
            "assigned_trainer": "VARCHAR(160)",
            "current_students": "INTEGER DEFAULT 0",
            "girls_count": "INTEGER DEFAULT 0",
        },
    )
    add_missing(
        "students",
        {
            "father_name": "VARCHAR(160)",
            "mother_name": "VARCHAR(160)",
            "age": "INTEGER",
            "gender": "VARCHAR(40)",
            "caste": "VARCHAR(80)",
            "category": "VARCHAR(80)",
            "phone": "VARCHAR(40)",
            "urban_rural": "VARCHAR(40)",
            "income_status": "VARCHAR(20)",
            "physically_challenged": "VARCHAR(10)",
            "medium": "VARCHAR(80)",
            "state": "VARCHAR(120)",
            "district": "VARCHAR(120)",
            "taluk": "VARCHAR(120)",
            "village": "VARCHAR(160)",
            "pincode": "VARCHAR(20)",
            "address": "TEXT",
        },
    )
    add_missing(
        "courses",
        {
            "sector": "VARCHAR(160)",
            "sub_sector": "VARCHAR(160)",
            "occupation": "VARCHAR(160)",
            "reference_id": "VARCHAR(120)",
            "resource_url": "VARCHAR(500)",
        },
    )
    add_missing(
        "batches",
        {
            "start_date": "DATE",
            "end_date": "DATE",
        },
    )
    add_missing(
        "student_teamwork_badges",
        {
            "course_id": "INTEGER",
        },
    )
    add_missing(
        "accounts",
        {
            "plain_password": "VARCHAR(240)",
        },
    )
    add_missing(
        "enrollments",
        {
            "batch_id": "INTEGER",
        },
    )
    add_missing(
        "atl_forms",
        {
            "description": "TEXT",
            "is_hidden": "BOOLEAN DEFAULT FALSE",
        },
    )


_ensure_columns()
Base.metadata.create_all(bind=engine)

with SessionLocal() as db:
    seed_initial_data(db)
    seed_trainer_accounts(db)
    seed_lms_content(db)
    seed_curriculum_content(db)
    seed_handbook_volume_courses(db)

app = FastAPI(
    title="GTTC Student Dashboard & LMS",
    description="Student, trainer, school, and LMS extension for the GTTC portal.",
    version="0.1.0",
)

templates = Jinja2Templates(directory=APP_DIR / "templates")

default_upload_dir = "/tmp/student-dashboard/uploads/lms" if os.getenv("VERCEL") else str(APP_DIR / "static" / "uploads" / "lms")
UPLOAD_DIR = Path(os.getenv("STUDENT_DASHBOARD_UPLOAD_DIR", default_upload_dir))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

default_forms_upload_dir = "/tmp/student-dashboard/uploads/forms/custom" if os.getenv("VERCEL") else str(APP_DIR / "static" / "uploads" / "forms" / "custom")
FORMS_UPLOAD_DIR = Path(os.getenv("STUDENT_DASHBOARD_FORMS_UPLOAD_DIR", default_forms_upload_dir))
FORMS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ATL_FORM_EXTENSIONS = {".doc", ".docx", ".pdf", ".xlsx"}
BLOB_READ_WRITE_TOKEN = os.getenv("BLOB_READ_WRITE_TOKEN", "").strip()


def _media_type_for_name(filename: str) -> str:
    guessed, _encoding = mimetypes.guess_type(filename)
    return guessed or "application/octet-stream"


def _resolve_lms_upload_path(filename: str) -> Path | None:
    safe_name = Path(str(filename or "")).name
    if not safe_name or safe_name in {".", ".."}:
        return None
    candidates = [
        UPLOAD_DIR / safe_name,
        APP_DIR / "static" / "uploads" / "lms" / safe_name,
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _store_lms_upload_in_db(safe_name: str, content: bytes, content_type: str):
    with SessionLocal() as db:
        existing = db.query(models.MediaFile).filter(models.MediaFile.stored_name == safe_name).first()
        if existing:
            existing.data = content
            existing.content_type = content_type
            existing.original_name = safe_name
        else:
            db.add(
                models.MediaFile(
                    stored_name=safe_name,
                    original_name=safe_name,
                    content_type=content_type,
                    data=content,
                )
            )
        db.commit()


def _load_lms_upload_from_db(safe_name: str) -> tuple[bytes, str] | None:
    with SessionLocal() as db:
        row = db.query(models.MediaFile).filter(models.MediaFile.stored_name == safe_name).first()
        if not row or not row.data:
            return None
        return bytes(row.data), (row.content_type or _media_type_for_name(safe_name))


def _upload_to_vercel_blob(safe_name: str, content: bytes, content_type: str) -> str | None:
    if not BLOB_READ_WRITE_TOKEN:
        return None
    pathname = quote(f"lms/{safe_name}", safe="/")
    request = urllib.request.Request(
        f"https://blob.vercel-storage.com/{pathname}",
        data=content,
        method="PUT",
        headers={
            "Authorization": f"Bearer {BLOB_READ_WRITE_TOKEN}",
            "x-api-version": "7",
            "x-content-type": content_type,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError, UnicodeDecodeError):
        return None
    url = str(payload.get("url") or "").strip()
    return url or None


def _persist_lms_upload(safe_name: str, content: bytes) -> str:
    """Store an uploaded LMS file and return a URL that can be opened in the app."""
    content_type = _media_type_for_name(safe_name)
    blob_url = _upload_to_vercel_blob(safe_name, content, content_type)
    if blob_url:
        return blob_url

    target = UPLOAD_DIR / safe_name
    target.write_bytes(content)
    # Persist in DB so Vercel /tmp cold starts do not lose uploaded PDFs.
    try:
        _store_lms_upload_in_db(safe_name, content, content_type)
    except Exception:
        pass
    return f"/media/lms/{safe_name}"


def _serve_lms_file(filename: str):
    safe_name = Path(str(filename or "")).name
    if not safe_name or safe_name in {".", ".."}:
        raise HTTPException(status_code=400, detail="Invalid file name.")

    target = _resolve_lms_upload_path(safe_name)
    if target:
        return FileResponse(
            target,
            filename=target.name,
            media_type=_media_type_for_name(target.name),
            headers={"Content-Disposition": f'inline; filename="{target.name}"'},
        )

    stored = _load_lms_upload_from_db(safe_name)
    if stored:
        data, content_type = stored
        # Warm local cache for subsequent requests on this instance.
        try:
            (UPLOAD_DIR / safe_name).write_bytes(data)
        except Exception:
            pass
        return StreamingResponse(
            BytesIO(data),
            media_type=content_type,
            headers={"Content-Disposition": f'inline; filename="{safe_name}"'},
        )

    raise HTTPException(
        status_code=404,
        detail="Uploaded file not found. Please re-upload the PDF/resource from Content Library.",
    )


@app.get("/media/lms/{filename}")
async def serve_lms_media(filename: str, request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    return _serve_lms_file(filename)


@app.get("/static/uploads/lms/{filename}")
async def serve_legacy_lms_static(filename: str, request: Request):
    """Serve previously saved /static/uploads/lms URLs from disk or database."""
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    return _serve_lms_file(filename)


app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")

SESSION_COOKIE = "student_dashboard_session"
SESSION_VALUE = os.getenv("STUDENT_DASHBOARD_SESSION", "student-dashboard-local-session")
LOGIN_USERNAME = os.getenv("STUDENT_DASHBOARD_USERNAME", "admin@gttc.gov.in")
LOGIN_PASSWORD = os.getenv("STUDENT_DASHBOARD_PASSWORD", "Admin@123")


def _is_authenticated(request: Request):
    session = request.cookies.get(SESSION_COOKIE)
    return bool(session and (session == SESSION_VALUE or session.startswith("account:")))


def _current_account(request: Request):
    session = request.cookies.get(SESSION_COOKIE, "")
    if session == SESSION_VALUE:
        return {"name": "Admin User", "email": LOGIN_USERNAME, "role": "admin"}
    if not session.startswith("account:"):
        return None
    account_id = _parse_int(session.split(":")[1] if len(session.split(":")) > 1 else None)
    if not account_id:
        return None
    with SessionLocal() as db:
        account = db.query(models.Account).filter(models.Account.id == account_id).first()
        if not account:
            return None
        return {"name": account.name, "email": account.email, "role": account.role}


def _is_admin(request: Request):
    account = _current_account(request)
    return bool(account and account["role"] == "admin")


def _is_master_trainer(request: Request):
    account = _current_account(request)
    if not account:
        return False
    if account.get("role") == "admin":
        return True
    if account.get("role") != "trainer":
        return False
    with SessionLocal() as db:
        trainer = (
            db.query(models.Trainer)
            .filter(func.lower(models.Trainer.email) == account.get("email", "").lower())
            .first()
        )
        return bool(trainer and trainer.role == "master_trainer")


def _can_manage_forms(request: Request):
    return _is_admin(request) or _is_master_trainer(request)


def _official_form_codes():
    return {form["code"].upper() for form in ATL_MER_FORMS}


def _validate_form_upload(file: UploadFile | None):
    if not file or not file.filename:
        raise ValueError("Please choose a form file to upload.")
    extension = Path(file.filename).suffix.lower()
    if extension not in ATL_FORM_EXTENSIONS:
        allowed = ", ".join(sorted(ATL_FORM_EXTENSIONS))
        raise ValueError(f"Invalid file type. Allowed: {allowed}.")


async def _read_form_upload(file: UploadFile) -> bytes:
    _validate_form_upload(file)
    content = await file.read()
    if not content:
        raise ValueError("Uploaded form file is empty.")
    return content


def _write_form_file(stored_name: str, content: bytes):
    target = FORMS_UPLOAD_DIR / stored_name
    target.write_bytes(content)
    return target


def _remove_form_file(stored_name: str | None):
    if not stored_name:
        return
    target = FORMS_UPLOAD_DIR / stored_name
    if target.exists():
        target.unlink()


def _atl_form_record(db, form_id: int):
    form = db.query(models.AtlForm).filter(models.AtlForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found.")
    return form


def _form_card_description(source: str, description: str | None = None, uploaded_by: str | None = None) -> str:
    cleaned = str(description or "").strip()
    if cleaned:
        return cleaned
    if source == "official_override":
        return f"Updated by {uploaded_by or 'Admin'}. This replaces the default official file until restored."
    if source == "uploaded":
        return f"Added by {uploaded_by or 'Admin'}. Download, print, and complete manually."
    return "Official ATL form ready for printing and manual completion."


def _clean_form_description(description: str) -> str | None:
    cleaned = str(description or "").strip()
    if not cleaned:
        return None
    return cleaned[:500]


def _can_manage_courses(request: Request):
    return _is_master_trainer(request)


def _can_manage_trainers(request: Request):
    return _is_admin(request) or _is_master_trainer(request)


def _build_profile_details(db, current_account):
    if not current_account:
        return {}

    role_label = current_account.get("role", "").replace("_", " ").title()
    details = {
        "name": current_account.get("name"),
        "email": current_account.get("email"),
        "role": current_account.get("role"),
        "role_label": role_label,
        "phone": None,
        "division": None,
        "districts": None,
        "assigned_school": None,
        "specialization": None,
        "trainer_role": None,
        "trainer_role_label": None,
    }

    if current_account.get("role") == "trainer":
        trainer = (
            db.query(models.Trainer)
            .filter(func.lower(models.Trainer.email) == current_account.get("email", "").lower())
            .first()
        )
        if trainer:
            trainer_role_label = trainer.role.replace("_", " ").title() if trainer.role else None
            details.update(
                {
                    "phone": trainer.phone,
                    "division": trainer.division,
                    "districts": trainer.districts,
                    "assigned_school": _format_assigned_schools(db, trainer.assigned_school) or trainer.assigned_school,
                    "specialization": trainer.specialization,
                    "trainer_role": trainer.role,
                    "trainer_role_label": trainer_role_label,
                    "role_label": trainer_role_label or role_label,
                }
            )

    return details


def _hash_password(password: str):
    return bcrypt.hashpw(password[:72].encode(), bcrypt.gensalt()).decode()


def _verify_password(password: str, hashed_password: str):
    return bcrypt.checkpw(password[:72].encode(), hashed_password.encode())


def _set_session_cookie(response: RedirectResponse, value: str):
    response.set_cookie(
        SESSION_COOKIE,
        value,
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 8,
    )


def _dashboard_redirect(message: str, kind: str = "success"):
    return RedirectResponse(f"/dashboard?{urlencode({'notice': message, 'notice_kind': kind})}", status_code=303)


def _parse_int(value, default=None):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _split_assigned_schools(value: str):
    if not value:
        return []
    parts = re.split(r"\s*(?:,|;|\|)\s*", str(value).strip())
    return [part for part in (item.strip() for item in parts) if part]


def _join_assigned_schools(*values):
    schools = []
    for value in values:
        for school in _split_assigned_schools(value):
            if school and school.lower() not in {item.lower() for item in schools}:
                schools.append(school)
            if len(schools) == 3:
                return ", ".join(schools)
    return ", ".join(schools) or None


def _school_ids_from_assigned_school_value(db, value: str | None):
    if not value:
        return []

    school_lookup = {school.id: school for school in db.query(models.School).all() if school.id is not None}
    matched_ids = []
    seen_ids = set()
    for token in _split_assigned_schools(value):
        cleaned = token.strip()
        if not cleaned:
            continue
        if cleaned.isdigit():
            school_id = int(cleaned)
            if school_id in school_lookup and school_id not in seen_ids:
                matched_ids.append(school_id)
                seen_ids.add(school_id)
            continue

        token_l = cleaned.lower()
        for school_id, school in school_lookup.items():
            school_name = (school.name or "").strip().lower()
            school_label = f"{school_name} - {(school.district or '').strip().lower()}".strip(" -")
            if school_name and (
                school_name == token_l
                or school_label == token_l
                or school_name.startswith(token_l)
                or token_l.startswith(school_name)
            ):
                if school_id not in seen_ids:
                    matched_ids.append(school_id)
                    seen_ids.add(school_id)
                break
    return matched_ids


def _validate_trainer_assignment_selection(db, selected_school_ids, current_trainer_id=None):
    selected_ids = {int(school_id) for school_id in selected_school_ids if str(school_id).strip()}
    if not selected_ids:
        return None

    active_trainers = (
        db.query(models.Trainer)
        .filter(models.Trainer.is_active.is_(True))
        .all()
    )
    for trainer in active_trainers:
        if current_trainer_id is not None and trainer.id == current_trainer_id:
            continue
        assigned_ids = set(_school_ids_from_assigned_school_value(db, trainer.assigned_school))
        if assigned_ids & selected_ids:
            return f"{trainer.name} already has one of the selected schools assigned."
    return None


def _active_school_assignment_map(db):
    assignment_map = {}
    active_trainers = db.query(models.Trainer).filter(models.Trainer.is_active.is_(True)).all()
    for trainer in active_trainers:
        for school_id in _school_ids_from_assigned_school_value(db, trainer.assigned_school):
            assignment_map.setdefault(school_id, trainer.id)
    return assignment_map


def _trainer_school_scope_ids(db, account):
    if not account or account.get("role") != "trainer":
        return None
    trainer = (
        db.query(models.Trainer)
        .filter(func.lower(models.Trainer.email) == account.get("email", "").lower())
        .first()
    )
    if not trainer:
        return set()
    # Master trainers can work across all schools.
    if trainer.role == "master_trainer":
        return None

    tokens = _split_assigned_schools(trainer.assigned_school)
    schools = db.query(models.School).all()
    matched_ids = set()

    for token in tokens:
        cleaned = token.strip()
        if not cleaned:
            continue
        if cleaned.isdigit():
            matched_ids.add(int(cleaned))
            continue
        token_l = cleaned.lower()
        for school in schools:
            school_name = (school.name or "").strip().lower()
            school_label = f"{school_name} - {(school.district or '').strip().lower()}".strip(" -")
            if not school_name:
                continue
            if (
                school_name == token_l
                or school_label == token_l
                or school_name.startswith(token_l)
                or token_l.startswith(school_name)
            ):
                matched_ids.add(school.id)

    trainer_name_l = (trainer.name or "").strip().lower()
    if trainer_name_l:
        for school in schools:
            assigned_trainer = (school.assigned_trainer or "").strip().lower()
            if assigned_trainer and (
                assigned_trainer == trainer_name_l
                or trainer_name_l in assigned_trainer
                or assigned_trainer in trainer_name_l
            ):
                matched_ids.add(school.id)

    return matched_ids


def _school_in_scope(scope_school_ids, school_id):
    return scope_school_ids is None or (school_id is not None and school_id in scope_school_ids)


def _request_school_scope_ids(request: Request, db):
    return _trainer_school_scope_ids(db, _current_account(request))


def _format_assigned_schools(db, value: str | None):
    tokens = _split_assigned_schools(value)
    if not tokens:
        return None
    school_by_id = {
        str(school.id): school.name
        for school in db.query(models.School).all()
        if school.id is not None
    }
    labels = []
    for token in tokens:
        if token.isdigit() and token in school_by_id:
            labels.append(school_by_id[token])
        else:
            labels.append(token)
    return ", ".join(labels) or None


def _is_atl_trainer_account(request: Request):
    account = _current_account(request)
    if not account or account.get("role") != "trainer":
        return False
    with SessionLocal() as db:
        trainer = (
            db.query(models.Trainer)
            .filter(func.lower(models.Trainer.email) == account.get("email", "").lower())
            .first()
        )
        return bool(trainer and trainer.role != "master_trainer")


def _compose_student_address(state="", district="", taluk="", village="", pincode="", fallback=""):
    parts = [village, taluk, district, state, pincode]
    cleaned = [str(part).strip() for part in parts if str(part or "").strip()]
    return ", ".join(cleaned) or (fallback.strip() if fallback else None)


PERFORMANCE_LEVELS = {"not_assessed", "needs_support", "developing", "proficient", "excellent"}
TEAMWORK_BADGES = {
    "ideator": "Ideator",
    "goal_getter": "Goal Getter",
    "co_creator": "Co-Creator",
    "questioner": "Questioner",
}


def _performance_level(value: str):
    return value if value in PERFORMANCE_LEVELS else "not_assessed"


def _worksheet_rows(upload: bytes):
    try:
        workbook = openpyxl.load_workbook(BytesIO(upload), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Cannot parse Excel file: {exc}") from exc

    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty Excel file")

    headers = [str(header).strip().lower() if header else "" for header in rows[0]]

    def col(row, name):
        try:
            value = row[headers.index(name)]
        except (ValueError, IndexError):
            return ""
        return str(value).strip() if value is not None else ""

    return rows[1:], col


def _matches_search(row: dict, search: str):
    if not search:
        return True
    needle = search.lower()
    return any(needle in str(value or "").lower() for value in row.values())


def _export_workbook(sheet_name: str, headers: list[str], rows: list[dict], filename: str):
    return _export_workbook_sheets([{"name": sheet_name, "headers": headers, "rows": rows}], filename)


def _autosize_worksheet(worksheet):
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 45)


def _append_labeled_table(worksheet, title: str, columns: list[tuple[str, str]], rows: list[dict]):
    worksheet.append([title])
    keys = [key for key, _label in columns]
    worksheet.append([label for _key, label in columns])
    for row in rows:
        worksheet.append([row.get(key, "") for key in keys])
    worksheet.append([])


def _export_workbook_sheets(sheets: list[dict], filename: str):
    workbook = openpyxl.Workbook()
    first = True
    for sheet in sheets:
        if first:
            worksheet = workbook.active
            first = False
        else:
            worksheet = workbook.create_sheet()
        worksheet.title = str(sheet["name"])[:31]
        columns = sheet.get("columns")
        if columns:
            _append_labeled_table(worksheet, sheet.get("title") or sheet["name"], columns, sheet.get("rows") or [])
            # Remove the trailing blank row left by helper when it's a dedicated sheet
            if worksheet.max_row > 1 and all(cell.value in (None, "") for cell in worksheet[worksheet.max_row]):
                worksheet.delete_rows(worksheet.max_row)
        else:
            headers = sheet["headers"]
            worksheet.append(headers)
            for row in sheet.get("rows") or []:
                worksheet.append([row.get(header, "") for header in headers])
        _autosize_worksheet(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_student_combination_workbook(summary_rows: list[dict], detail_rows: list[dict], filename: str):
    """One sheet with Matching Students first, then Summary — so details are always visible."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Student Report"
    _append_labeled_table(
        worksheet,
        f"Matching Students ({len(detail_rows)})",
        REPORT_TABLE_COLUMNS["student_detail_list"],
        detail_rows,
    )
    _append_labeled_table(
        worksheet,
        f"Summary ({len(summary_rows)})",
        REPORT_TABLE_COLUMNS["student_combination_summary"],
        summary_rows,
    )
    if worksheet.max_row > 1 and all(cell.value in (None, "") for cell in worksheet[worksheet.max_row]):
        worksheet.delete_rows(worksheet.max_row)
    _autosize_worksheet(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _report_label(value, fallback: str = "Unknown"):
    text = str(value or "").strip()
    return text if text else fallback


def _report_key(value):
    return _report_label(value).lower()


def _report_percent(part: int, total: int):
    return round((part / total) * 100) if total else 0


def _report_increment(groups: dict, key, base: dict):
    bucket = groups.setdefault(key, dict(base))
    bucket["total"] = bucket.get("total", 0) + 1
    return bucket


def _report_filter_options(existing_values, fixed_values=()):
    options = {}
    for value in [*fixed_values, *existing_values]:
        label = _report_label(value)
        if label == "Unknown":
            continue
        options.setdefault(label.lower(), label)
    return sorted(options.values(), key=str.lower)


def _school_scope_filter(query, column, scope_school_ids):
    if scope_school_ids is None:
        return query
    return query.filter(column.in_(scope_school_ids))


def _matches_report_filters(row: dict, filters: dict):
    for key, value in filters.items():
        if not value:
            continue
        if key == "category":
            row_code = normalize_category(row.get("category"))
            filter_code = normalize_category(value)
            if row_code and filter_code:
                if row_code != filter_code:
                    return False
                continue
        if key == "religion":
            row_religion = normalize_religion(row.get("religion"))
            filter_religion = normalize_religion(value)
            if row_religion and filter_religion:
                if row_religion.lower() != filter_religion.lower():
                    return False
                continue
        if key not in row:
            continue
        row_value = str(row.get(key) or "").strip()
        if row_value.lower() != str(value).strip().lower():
            return False
    return True


def _student_matches_combination(detail: dict, summary: dict) -> bool:
    """Match a student to a summary combo row on the dimensions that define that combo."""
    fields = (
        "school_id",
        "district",
        "gender",
        "income_status",
        "physically_challenged",
        "medium",
        "urban_rural",
    )
    for field in fields:
        left = str(detail.get(field) or "").strip().lower()
        right = str(summary.get(field) or "").strip().lower()
        if left != right:
            return False
    return True


def _date_in_report_range(value, start_date, end_date):
    parsed = _parse_date(value)
    if not parsed:
        return True
    if start_date and parsed < start_date:
        return False
    if end_date and parsed > end_date:
        return False
    return True


REPORT_TABLE_COLUMNS = {
    "school_summary": [
        ("school", "School"), ("district", "District"), ("division", "Division"),
        ("assigned_trainer", "Trainer"), ("registered_students", "Registered Students"),
        ("reported_students", "Reported Students"), ("girls_reported", "Girls Reported"),
    ],
    "schools_by_district": [
        ("state", "State"), ("division", "Division"), ("district", "District"), ("total", "Schools"),
    ],
    "student_school_summary": [
        ("school", "School"), ("district", "District"), ("total", "Students"),
        ("girls", "Girls"), ("boys", "Boys"), ("other_gender", "Other"),
        ("apl", "APL"), ("bpl", "BPL"), ("physically_challenged_yes", "Physically Challenged"),
    ],
    "student_location_summary": [
        ("state", "State"), ("district", "District"), ("taluk", "Taluk"),
        ("village", "Village"), ("pincode", "Pincode"), ("total", "Students"),
        ("girls", "Girls"), ("boys", "Boys"), ("apl", "APL"), ("bpl", "BPL"),
    ],
    "student_demographics": [
        ("dimension", "Dimension"), ("value", "Value"), ("district", "District"),
        ("total", "Students"),
    ],
    "student_gender_summary": [
        ("school", "School"), ("district", "District"), ("gender", "Gender"), ("total", "Students"),
    ],
    "student_income_summary": [
        ("school", "School"), ("district", "District"), ("income_status", "Income Status"), ("total", "Students"),
    ],
    "student_physical_summary": [
        ("school", "School"), ("district", "District"), ("physically_challenged", "Physically Challenged"), ("total", "Students"),
    ],
    "student_medium_summary": [
        ("school", "School"), ("district", "District"), ("medium", "Medium"), ("total", "Students"),
    ],
    "student_religion_category_summary": [
        ("school", "School"), ("district", "District"), ("religion", "Religion"), ("category", "Category"), ("total", "Students"),
    ],
    "student_combination_summary": [
        ("school", "School"), ("district", "District"), ("gender", "Gender"), ("income_status", "Income"),
        ("physically_challenged", "Physically Challenged"), ("medium", "Medium"), ("urban_rural", "Urban/Rural"), ("total", "Students"),
    ],
    "student_detail_list": [
        ("name", "Name"), ("email", "Email"), ("grade", "Grade"), ("father_name", "Father"), ("mother_name", "Mother"),
        ("age", "Age"), ("gender", "Gender"), ("religion", "Religion"), ("category", "Category"), ("phone", "Phone"),
        ("urban_rural", "Urban/Rural"), ("income_status", "Income"), ("physically_challenged", "Physically Challenged"),
        ("medium", "Medium"), ("state", "State"), ("district", "District"), ("taluk", "Taluk"), ("village", "Village"),
        ("pincode", "Pincode"), ("school", "School"), ("trainer", "Trainer"),
    ],
    "trainer_summary": [
        ("trainer", "Trainer"), ("email", "Email"), ("role", "Role"), ("division", "Division"),
        ("districts", "Districts"), ("assigned_school_count", "Schools Assigned"), ("assigned_schools", "Assigned Schools"),
    ],
    "batch_summary": [
        ("school", "School"), ("district", "District"), ("batch", "Batch"), ("trainer", "Trainer"),
        ("start_date", "Start Date"), ("end_date", "End Date"),
    ],
    "enrollment_summary": [
        ("school", "School"), ("district", "District"), ("course", "Course"), ("status", "Status"),
        ("total", "Enrollments"), ("average_progress", "Avg Progress"),
    ],
    "attendance_summary": [
        ("school", "School"), ("district", "District"), ("batch", "Batch"),
        ("total", "Records"), ("present", "Present"), ("absent", "Absent"), ("present_percent", "Present %"),
    ],
    "performance_summary": [
        ("school", "School"), ("district", "District"), ("batch", "Batch"), ("trainer", "Trainer"),
        ("concept_understanding", "Concept"), ("project_understanding", "Project"),
        ("design_thinking", "Design Thinking"), ("assessed_by", "Assessed By"),
    ],
}


REPORT_TABLE_TITLES = {
    "school_summary": "Registered Schools",
    "schools_by_district": "Schools by District",
    "student_school_summary": "Students by School",
    "student_location_summary": "Students by Location",
    "student_demographics": "Student Demographic Combinations",
    "student_gender_summary": "Students by Gender",
    "student_income_summary": "Students by APL/BPL",
    "student_physical_summary": "Students by Physically Challenged Status",
    "student_medium_summary": "Students by Medium",
    "student_religion_category_summary": "Students by Religion and Category",
    "student_combination_summary": "Student Detail Combination Report",
    "student_detail_list": "Matching Students",
    "trainer_summary": "Trainer Assignment Summary",
    "batch_summary": "Batches by School and Trainer",
    "enrollment_summary": "Enrollments by Course and Status",
    "attendance_summary": "Attendance by School and Batch",
    "performance_summary": "Performance by Batch",
}


def _build_reports_payload(db, current_account):
    scope_school_ids = _trainer_school_scope_ids(db, current_account)
    schools = _school_scope_filter(db.query(models.School), models.School.id, scope_school_ids).order_by(models.School.name).all()
    school_by_id = {school.id: school for school in schools}
    scoped_school_ids = set(school_by_id)

    students_query = db.query(models.Student, models.School).outerjoin(models.School, models.School.id == models.Student.school_id)
    if scope_school_ids is not None:
        students_query = students_query.filter(models.Student.school_id.in_(scope_school_ids))
    students = students_query.order_by(models.Student.name).all()

    trainers_query = db.query(models.Trainer)
    if current_account.get("role") == "trainer":
        trainers_query = trainers_query.filter(func.lower(models.Trainer.email) == current_account.get("email", "").lower())
    trainers = trainers_query.order_by(models.Trainer.name).all()

    batches_query = db.query(models.Batch, models.School).join(models.School, models.School.id == models.Batch.school_id)
    batches_query = _school_scope_filter(batches_query, models.Batch.school_id, scope_school_ids)
    batches = batches_query.order_by(models.School.name, models.Batch.name).all()

    enrollments_query = (
        db.query(models.Enrollment, models.Student, models.Course, models.School)
        .join(models.Student, models.Student.id == models.Enrollment.student_id)
        .join(models.Course, models.Course.id == models.Enrollment.course_id)
        .outerjoin(models.School, models.School.id == models.Student.school_id)
    )
    if scope_school_ids is not None:
        enrollments_query = enrollments_query.filter(models.Student.school_id.in_(scope_school_ids))
    enrollments = enrollments_query.all()

    attendance_query = (
        db.query(models.AttendanceRecord, models.Batch, models.Student, models.School)
        .join(models.Batch, models.Batch.id == models.AttendanceRecord.batch_id)
        .join(models.Student, models.Student.id == models.AttendanceRecord.student_id)
        .join(models.School, models.School.id == models.Batch.school_id)
    )
    attendance_query = _school_scope_filter(attendance_query, models.Batch.school_id, scope_school_ids)
    attendance_records = attendance_query.all()

    assessments_query = (
        db.query(models.BatchPerformanceAssessment, models.Batch, models.School)
        .join(models.Batch, models.Batch.id == models.BatchPerformanceAssessment.batch_id)
        .join(models.School, models.School.id == models.Batch.school_id)
    )
    assessments_query = _school_scope_filter(assessments_query, models.Batch.school_id, scope_school_ids)
    assessments = assessments_query.all()

    school_summary = [
        {
            "school_id": school.id,
            "school": school.name,
            "state": _report_label(school.state),
            "district": _report_label(school.district),
            "division": _report_label(school.division),
            "assigned_trainer": _report_label(school.assigned_trainer),
            "registered_students": sum(1 for student, _school in students if student.school_id == school.id),
            "reported_students": school.current_students or 0,
            "girls_reported": school.girls_count or 0,
            "trainer": _report_label(school.assigned_trainer),
        }
        for school in schools
    ]

    schools_by_district_groups = {}
    for school in schools:
        key = (_report_key(school.state), _report_key(school.division), _report_key(school.district))
        _report_increment(
            schools_by_district_groups,
            key,
            {
                "state": _report_label(school.state),
                "division": _report_label(school.division),
                "district": _report_label(school.district),
            },
        )
    schools_by_district = list(schools_by_district_groups.values())

    student_school_groups = {}
    student_location_groups = {}
    student_demographic_groups = {}
    student_gender_groups = {}
    student_income_groups = {}
    student_physical_groups = {}
    student_medium_groups = {}
    student_religion_category_groups = {}
    student_combination_groups = {}
    for student, school in students:
        school_name = school.name if school else "Unknown"
        district = _report_label(student.district or (school.district if school else ""))
        state = _report_label(student.state or (school.state if school else ""))
        taluk = _report_label(student.taluk)
        village = _report_label(student.village)
        pincode = _report_label(student.pincode)
        gender_label = _report_label(student.gender)
        income_label = _report_label(student.income_status)
        challenged_label = _report_label(student.physically_challenged)
        medium_label = _report_label(student.medium)
        religion_label = _report_label(normalize_religion(student.caste) or student.caste)
        category_code = normalize_category(student.category) or student.category
        category_label_value = category_display(category_code) if category_code else "Unknown"
        urban_rural_label = _report_label(student.urban_rural)
        gender = _report_key(student.gender)
        income = _report_key(student.income_status)
        challenged = _report_key(student.physically_challenged)

        school_bucket = _report_increment(
            student_school_groups,
            student.school_id or "unknown",
            {"school_id": student.school_id or "", "school": school_name, "district": district},
        )
        location_bucket = _report_increment(
            student_location_groups,
            (state.lower(), district.lower(), taluk.lower(), village.lower(), pincode.lower()),
            {"state": state, "district": district, "taluk": taluk, "village": village, "pincode": pincode},
        )
        for bucket in (school_bucket, location_bucket):
            if gender == "female":
                bucket["girls"] = bucket.get("girls", 0) + 1
            elif gender == "male":
                bucket["boys"] = bucket.get("boys", 0) + 1
            else:
                bucket["other_gender"] = bucket.get("other_gender", 0) + 1
            if income == "apl":
                bucket["apl"] = bucket.get("apl", 0) + 1
            elif income == "bpl":
                bucket["bpl"] = bucket.get("bpl", 0) + 1
            if challenged == "yes":
                bucket["physically_challenged_yes"] = bucket.get("physically_challenged_yes", 0) + 1

        for dimension, value in [
            ("Gender", student.gender),
            ("Income Status", student.income_status),
            ("Physically Challenged", student.physically_challenged),
            ("Urban/Rural", student.urban_rural),
            ("Medium", student.medium),
            ("Religion", student.caste),
            ("Category", student.category),
        ]:
            _report_increment(
                student_demographic_groups,
                (dimension, _report_key(value), district.lower()),
                {"dimension": dimension, "value": _report_label(value), "district": district},
            )

        _report_increment(
            student_gender_groups,
            (student.school_id or "unknown", district.lower(), _report_key(gender_label)),
            {"school_id": student.school_id or "", "school": school_name, "district": district, "gender": gender_label},
        )
        _report_increment(
            student_income_groups,
            (student.school_id or "unknown", district.lower(), _report_key(income_label)),
            {"school_id": student.school_id or "", "school": school_name, "district": district, "income_status": income_label},
        )
        _report_increment(
            student_physical_groups,
            (student.school_id or "unknown", district.lower(), _report_key(challenged_label)),
            {"school_id": student.school_id or "", "school": school_name, "district": district, "physically_challenged": challenged_label},
        )
        _report_increment(
            student_medium_groups,
            (student.school_id or "unknown", district.lower(), _report_key(medium_label)),
            {"school_id": student.school_id or "", "school": school_name, "district": district, "medium": medium_label},
        )
        _report_increment(
            student_religion_category_groups,
            (student.school_id or "unknown", district.lower(), _report_key(religion_label), _report_key(category_code or category_label_value)),
            {
                "school_id": student.school_id or "",
                "school": school_name,
                "district": district,
                "religion": religion_label,
                "category": category_code or "",
            },
        )
        _report_increment(
            student_combination_groups,
            (
                student.school_id or "unknown",
                district.lower(),
                _report_key(gender_label),
                _report_key(income_label),
                _report_key(challenged_label),
                _report_key(medium_label),
                _report_key(urban_rural_label),
            ),
            {
                "school_id": student.school_id or "",
                "school": school_name,
                "district": district,
                "taluk": taluk,
                "village": village,
                "trainer": _report_label(school.assigned_trainer if school else ""),
                "gender": gender_label,
                "income_status": income_label,
                "physically_challenged": challenged_label,
                "medium": medium_label,
                "urban_rural": urban_rural_label,
                "religion": religion_label,
                "category": category_code or "",
            },
        )

    student_school_summary = list(student_school_groups.values())
    student_location_summary = list(student_location_groups.values())
    student_demographics = list(student_demographic_groups.values())
    student_gender_summary = list(student_gender_groups.values())
    student_income_summary = list(student_income_groups.values())
    student_physical_summary = list(student_physical_groups.values())
    student_medium_summary = list(student_medium_groups.values())
    student_religion_category_summary = list(student_religion_category_groups.values())
    student_combination_summary = list(student_combination_groups.values())
    student_detail_list = [
        {
            "name": student.name,
            "email": student.email or "",
            "grade": student.grade or "",
            "father_name": student.father_name or "",
            "mother_name": student.mother_name or "",
            "age": student.age or "",
            "gender": _report_label(student.gender),
            "religion": _report_label(normalize_religion(student.caste) or student.caste),
            "category": normalize_category(student.category) or "",
            "phone": student.phone or "",
            "urban_rural": _report_label(student.urban_rural),
            "income_status": _report_label(student.income_status),
            "physically_challenged": _report_label(student.physically_challenged),
            "medium": _report_label(student.medium),
            "state": _report_label(student.state or (school.state if school else "")),
            "district": _report_label(student.district or (school.district if school else "")),
            "taluk": _report_label(student.taluk),
            "village": _report_label(student.village),
            "pincode": _report_label(student.pincode),
            "school_id": student.school_id or "",
            "school": school.name if school else "Unknown",
            "trainer": _report_label(school.assigned_trainer if school else ""),
        }
        for student, school in students
    ]

    trainer_summary = [
        {
            "trainer": trainer.name,
            "email": trainer.email,
            "role": trainer.role,
            "division": _report_label(trainer.division),
            "districts": _report_label(trainer.districts),
            "assigned_school_count": len(_split_assigned_schools(trainer.assigned_school)),
            "assigned_schools": trainer.assigned_school or "",
        }
        for trainer in trainers
    ]

    batch_summary = [
        {
            "school_id": school.id,
            "school": school.name,
            "district": _report_label(school.district),
            "batch": batch.name,
            "trainer": _report_label(batch.trainer_name),
            "start_date": batch.start_date.isoformat() if batch.start_date else "",
            "end_date": batch.end_date.isoformat() if batch.end_date else "",
        }
        for batch, school in batches
    ]

    enrollment_groups = {}
    for enrollment, _student, course, school in enrollments:
        key = (school.id if school else "", course.id, enrollment.status)
        bucket = _report_increment(
            enrollment_groups,
            key,
            {
                "school_id": school.id if school else "",
                "school": school.name if school else "Unknown",
                "district": _report_label(school.district if school else ""),
                "course": course.title,
                "status": enrollment.status,
                "progress_total": 0,
            },
        )
        bucket["progress_total"] += enrollment.progress or 0
        bucket["average_progress"] = _report_percent(bucket["progress_total"], bucket["total"])
    enrollment_summary = [
        {key: value for key, value in row.items() if key != "progress_total"}
        for row in enrollment_groups.values()
    ]

    attendance_groups = {}
    for record, batch, _student, school in attendance_records:
        key = (school.id, batch.id)
        bucket = _report_increment(
            attendance_groups,
            key,
            {
                "school_id": school.id,
                "school": school.name,
                "district": _report_label(school.district),
                "batch": batch.name,
                "present": 0,
                "absent": 0,
            },
        )
        if record.status == "present":
            bucket["present"] += 1
        elif record.status == "absent":
            bucket["absent"] += 1
        bucket["present_percent"] = _report_percent(bucket["present"], bucket["total"])
    attendance_summary = list(attendance_groups.values())

    performance_summary = [
        {
            "school_id": school.id,
            "school": school.name,
            "district": _report_label(school.district),
            "batch": batch.name,
            "trainer": _report_label(batch.trainer_name),
            "concept_understanding": assessment.concept_understanding,
            "project_understanding": assessment.project_understanding,
            "design_thinking": assessment.design_thinking,
            "assessed_by": _report_label(assessment.assessed_by),
        }
        for assessment, batch, school in assessments
    ]

    tables = {
        "school_summary": school_summary,
        "schools_by_district": schools_by_district,
        "student_school_summary": student_school_summary,
        "student_gender_summary": student_gender_summary,
        "student_income_summary": student_income_summary,
        "student_physical_summary": student_physical_summary,
        "student_medium_summary": student_medium_summary,
        "student_religion_category_summary": student_religion_category_summary,
        "student_combination_summary": student_combination_summary,
        "student_detail_list": student_detail_list,
        "student_location_summary": student_location_summary,
        "student_demographics": student_demographics,
        "trainer_summary": trainer_summary,
        "batch_summary": batch_summary,
        "enrollment_summary": enrollment_summary,
        "attendance_summary": attendance_summary,
        "performance_summary": performance_summary,
    }

    sections = []
    for key, rows in tables.items():
        if key == "student_detail_list":
            continue
        section = {
            "key": key,
            "title": REPORT_TABLE_TITLES[key],
            "columns": [{"key": column_key, "label": label} for column_key, label in REPORT_TABLE_COLUMNS[key]],
            "rows": rows,
        }
        if key == "student_combination_summary":
            section["detail_columns"] = [
                {"key": column_key, "label": label}
                for column_key, label in REPORT_TABLE_COLUMNS["student_detail_list"]
            ]
            section["detail_rows"] = tables.get("student_detail_list", [])
        sections.append(section)

    girls = sum(1 for student, _school in students if _report_key(student.gender) == "female")
    boys = sum(1 for student, _school in students if _report_key(student.gender) == "male")
    challenged_yes = sum(1 for student, _school in students if _report_key(student.physically_challenged) == "yes")
    apl = sum(1 for student, _school in students if _report_key(student.income_status) == "apl")
    bpl = sum(1 for student, _school in students if _report_key(student.income_status) == "bpl")

    return {
        "kpis": {
            "schools": len(schools),
            "students": len(students),
            "girls": girls,
            "boys": boys,
            "apl": apl,
            "bpl": bpl,
            "physically_challenged": challenged_yes,
            "trainers": len(trainers),
            "batches": len(batches),
            "enrollments": len(enrollments),
        },
        "filters": {
            "schools": [{"id": school.id, "name": school.name} for school in schools],
            "districts": _report_filter_options((school.district for school in schools)),
            "taluks": _report_filter_options((student.taluk for student, _school in students)),
            "villages": _report_filter_options((student.village for student, _school in students)),
            "trainers": _report_filter_options((school.assigned_trainer for school in schools)),
            "mediums": _report_filter_options((student.medium for student, _school in students), ("Kannada", "English")),
            "genders": _report_filter_options((student.gender for student, _school in students), ("Male", "Female", "Other")),
            "income_statuses": _report_filter_options((student.income_status for student, _school in students), ("APL", "BPL")),
            "physically_challenged": _report_filter_options((student.physically_challenged for student, _school in students), ("Yes", "No")),
            "urban_rural": _report_filter_options((student.urban_rural for student, _school in students), ("Urban", "Rural")),
            "religions": _report_filter_options(
                (normalize_religion(student.caste) or student.caste for student, _school in students),
                RELIGION_OPTIONS,
            ),
            "categories": _report_filter_options(
                (normalize_category(student.category) or student.category for student, _school in students),
                [code for code, _label in CATEGORY_OPTIONS],
            ),
            "category_labels": {code: label for code, label in CATEGORY_OPTIONS},
        },
        "sections": sections,
        "tables": tables,
    }


def _build_dashboard_analytics(students: list[dict], batches: list[dict], enrollments: list[dict]):
    """Board-ready dashboard metrics and chart series from live LMS records."""
    today = date.today()
    girls = boys = other_gender = 0
    apl = bpl = income_other = 0
    urban = rural = locale_other = 0
    district_counts: dict[str, int] = {}

    for student in students:
        gender = _report_key(student.get("gender"))
        if gender == "female":
            girls += 1
        elif gender == "male":
            boys += 1
        else:
            other_gender += 1

        income = _report_key(student.get("income_status"))
        if income == "apl":
            apl += 1
        elif income == "bpl":
            bpl += 1
        else:
            income_other += 1

        locale = _report_key(student.get("urban_rural"))
        if locale == "urban":
            urban += 1
        elif locale == "rural":
            rural += 1
        else:
            locale_other += 1

        district = _report_label(student.get("district"), "Unspecified")
        district_counts[district] = district_counts.get(district, 0) + 1

    batch_completed = batch_in_progress = batch_upcoming = batch_unscheduled = 0
    for batch in batches:
        start = batch.get("start_date")
        end = batch.get("end_date")
        if isinstance(start, str):
            start = _parse_date(start)
        if isinstance(end, str):
            end = _parse_date(end)

        if end and end < today:
            batch_completed += 1
        elif start and start > today:
            batch_upcoming += 1
        elif start or end:
            batch_in_progress += 1
        else:
            batch_unscheduled += 1

    enrollment_status_counts = {
        "assigned": 0,
        "in_progress": 0,
        "completed": 0,
        "dropped": 0,
    }
    progress_total = 0
    for enrollment in enrollments:
        status = _report_key(enrollment.get("status")) or "assigned"
        if status not in enrollment_status_counts:
            status = "assigned"
        enrollment_status_counts[status] += 1
        progress_total += int(enrollment.get("progress") or 0)

    enrollment_total = len(enrollments)
    students_total = len(students)
    batches_total = len(batches)
    top_districts = sorted(district_counts.items(), key=lambda item: (-item[1], item[0]))[:8]

    return {
        "students_total": students_total,
        "batches_total": batches_total,
        "enrollments_total": enrollment_total,
        "girls": girls,
        "boys": boys,
        "other_gender": other_gender,
        "girl_percent": _report_percent(girls, students_total),
        "boy_percent": _report_percent(boys, students_total),
        "batch_completed": batch_completed,
        "batch_in_progress": batch_in_progress,
        "batch_upcoming": batch_upcoming,
        "batch_unscheduled": batch_unscheduled,
        "batch_completion_percent": _report_percent(batch_completed, batches_total),
        "enrollment_completed": enrollment_status_counts["completed"],
        "enrollment_in_progress": enrollment_status_counts["in_progress"],
        "enrollment_assigned": enrollment_status_counts["assigned"],
        "enrollment_dropped": enrollment_status_counts["dropped"],
        "avg_progress": round(progress_total / enrollment_total) if enrollment_total else 0,
        "apl": apl,
        "bpl": bpl,
        "charts": {
            "gender": {
                "labels": ["Girls", "Boys", "Other"],
                "values": [girls, boys, other_gender],
                "colors": ["#db2777", "#2563eb", "#94a3b8"],
            },
            "batches": {
                "labels": ["Completed", "In Progress", "Upcoming", "Not Scheduled"],
                "values": [batch_completed, batch_in_progress, batch_upcoming, batch_unscheduled],
                "colors": ["#059669", "#d97706", "#0891b2", "#94a3b8"],
            },
            "enrollments": {
                "labels": ["Assigned", "In Progress", "Completed", "Dropped"],
                "values": [
                    enrollment_status_counts["assigned"],
                    enrollment_status_counts["in_progress"],
                    enrollment_status_counts["completed"],
                    enrollment_status_counts["dropped"],
                ],
                "colors": ["#64748b", "#2563eb", "#059669", "#dc2626"],
            },
            "districts": {
                "labels": [name for name, _count in top_districts],
                "values": [count for _name, count in top_districts],
                "colors": ["#1e3a8a", "#2563eb", "#0891b2", "#0f766e", "#d97706", "#db2777", "#475569", "#94a3b8"],
            },
            "income": {
                "labels": ["APL", "BPL", "Other"],
                "values": [apl, bpl, income_other],
                "colors": ["#0f766e", "#d97706", "#94a3b8"],
            },
        },
    }


def _safe_upload_name(filename: str):
    stem = Path(filename).stem or "content"
    suffix = Path(filename).suffix.lower()
    safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "-", stem).strip("-")[:80] or "content"
    return f"{safe_stem}-{secrets.token_hex(6)}{suffix}"


def _build_atl_forms(db):
    official_codes = _official_form_codes()
    uploaded = db.query(models.AtlForm).order_by(models.AtlForm.created_at.desc(), models.AtlForm.title).all()
    overrides = {}
    hidden_official = set()
    custom_forms = []
    for form in uploaded:
        code_key = (form.code or "").strip().upper()
        if form.is_hidden:
            if code_key and code_key in official_codes:
                hidden_official.add(code_key)
            continue
        if code_key and code_key in official_codes:
            overrides[code_key] = form
        else:
            custom_forms.append(form)

    forms = []
    for catalog in ATL_MER_FORMS:
        code = catalog["code"]
        if code.upper() in hidden_official:
            continue
        override = overrides.get(code.upper())
        if override and override.stored_name:
            source = "official_override"
            forms.append(
                {
                    "id": override.id,
                    "code": code,
                    "title": override.title or catalog["title"],
                    "description": _form_card_description(source, override.description, override.uploaded_by),
                    "description_raw": override.description or "",
                    "filename": override.filename,
                    "url": f"/forms/download/{override.id}",
                    "source": source,
                    "uploaded_by": override.uploaded_by,
                }
            )
        else:
            metadata = override if override and not override.stored_name else None
            source = "official"
            forms.append(
                {
                    "id": metadata.id if metadata else None,
                    "code": code,
                    "title": catalog["title"],
                    "description": _form_card_description(
                        source,
                        metadata.description if metadata else None,
                        metadata.uploaded_by if metadata else None,
                    ),
                    "description_raw": metadata.description if metadata else "",
                    "filename": catalog["filename"],
                    "url": catalog["url"],
                    "source": source,
                    "uploaded_by": metadata.uploaded_by if metadata else None,
                }
            )

    for form in custom_forms:
        if not form.stored_name:
            continue
        forms.append(
            {
                "id": form.id,
                "code": form.code or f"FORM-{form.id}",
                "title": form.title,
                "description": _form_card_description("uploaded", form.description, form.uploaded_by),
                "description_raw": form.description or "",
                "filename": form.filename,
                "url": f"/forms/download/{form.id}",
                "source": "uploaded",
                "uploaded_by": form.uploaded_by,
            }
        )
    return forms


def _validate_content_upload(content_type: str, upload: UploadFile | None, resource_url: str):
    allowed_extensions = {
        "pdf": {".pdf"},
        "ppt": {".ppt", ".pptx"},
        "video_file": {".mp4", ".mov", ".webm", ".mkv"},
    }
    if content_type == "article":
        return
    if content_type == "video_link":
        _validate_video_link(resource_url)
        return

    if content_type not in allowed_extensions:
        raise ValueError("Unsupported content type.")
    if not upload or not upload.filename:
        raise ValueError("Please choose a file to upload.")

    extension = Path(upload.filename).suffix.lower()
    if extension not in allowed_extensions[content_type]:
        allowed = ", ".join(sorted(allowed_extensions[content_type]))
        raise ValueError(f"Invalid file type. Allowed: {allowed}.")


def _normalize_resource_url(resource_url: str):
    cleaned_url = resource_url.strip()
    if cleaned_url.startswith("/"):
        return cleaned_url
    if cleaned_url and not cleaned_url.startswith(("http://", "https://")):
        cleaned_url = f"https://{cleaned_url}"
    return cleaned_url


def _validate_video_link(resource_url: str):
    if not resource_url:
        raise ValueError("Video link is required.")
    parsed = urlparse(resource_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("Video link must be a valid http:// or https:// URL.")


def _validate_optional_content_file(content_type: str, upload: UploadFile | None):
    allowed_extensions = {
        "pdf": {".pdf"},
        "ppt": {".ppt", ".pptx"},
        "video_file": {".mp4", ".mov", ".webm", ".mkv"},
    }
    if content_type in {"article", "video_link"}:
        return
    if content_type not in allowed_extensions:
        raise ValueError("Unsupported content type.")
    if not upload or not upload.filename:
        return
    extension = Path(upload.filename).suffix.lower()
    if extension not in allowed_extensions[content_type]:
        allowed = ", ".join(sorted(allowed_extensions[content_type]))
        raise ValueError(f"Invalid file type. Allowed: {allowed}.")


async def _save_content_upload(content_type: str, upload: UploadFile | None, resource_url: str):
    if content_type == "video_link":
        return resource_url.strip()
    if upload and upload.filename:
        safe_name = _safe_upload_name(upload.filename)
        content = await upload.read()
        if not content:
            raise ValueError("Uploaded file is empty.")
        return _persist_lms_upload(safe_name, content)
    return ""


def _content_group(course_title: str, content_type: str):
    if course_title == "Volume I - Learning by Doing":
        return "volume-1"
    if course_title == "Volume II - Grades 6-8 ATL Activities":
        return "volume-2"
    if course_title == "Volume III - Grades 9-10 ATL Activities":
        return "volume-3"
    if content_type in {"pdf", "ppt", "video_link", "video_file"} and not course_title.startswith("Experiment "):
        return "uploaded"
    if course_title.startswith("Experiment "):
        return "experiments"
    if course_title == "ATL Curriculum and Innovation Calendar 2026-27":
        return "curriculum"
    return "other"


@app.get("/")
async def root(request: Request):
    if _is_authenticated(request):
        return RedirectResponse("/dashboard")
    return RedirectResponse("/login")


@app.get("/login")
async def login_page(request: Request):
    if _is_authenticated(request):
        return RedirectResponse("/dashboard")
    return templates.TemplateResponse(
        request,
        "login.html",
        {
            "app_name": "GTTC Student Dashboard",
            "error": None,
            "username": LOGIN_USERNAME,
        },
    )


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    normalized_username = username.strip().lower()

    with SessionLocal() as db:
        account = db.query(models.Account).filter(func.lower(models.Account.email) == normalized_username).first()

    if account and _verify_password(password, account.hashed_password):
        if account.role == "trainer":
            with SessionLocal() as db:
                trainer = (
                    db.query(models.Trainer)
                    .filter(func.lower(models.Trainer.email) == normalized_username)
                    .first()
                )
                if trainer and trainer.is_active is False:
                    return templates.TemplateResponse(
                        request,
                        "login.html",
                        {
                            "app_name": "GTTC Student Dashboard",
                            "error": "This trainer account is deactivated. Contact admin or master trainer.",
                            "username": username,
                        },
                        status_code=401,
                    )
        response = RedirectResponse("/dashboard", status_code=303)
        _set_session_cookie(response, f"account:{account.id}:{secrets.token_urlsafe(16)}")
        return response

    valid_username = secrets.compare_digest(normalized_username, LOGIN_USERNAME.lower())
    valid_password = secrets.compare_digest(password, LOGIN_PASSWORD)

    if not (valid_username and valid_password):
        return templates.TemplateResponse(
            request,
            "login.html",
            {
                "app_name": "GTTC Student Dashboard",
                "error": "Invalid username or password.",
                "username": username,
            },
            status_code=401,
        )

    response = RedirectResponse("/dashboard", status_code=303)
    _set_session_cookie(response, SESSION_VALUE)
    return response


@app.get("/signup")
async def signup_page(request: Request):
    return RedirectResponse("/login", status_code=303)


@app.post("/signup")
async def signup(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form("student"),
    password: str = Form(...),
    confirm_password: str = Form(...),
):
    return RedirectResponse("/login", status_code=303)


@app.get("/logout")
async def logout():
    response = RedirectResponse("/login")
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/bulk-template/{record_type}")
async def bulk_template(record_type: str, request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if record_type == "trainers" and not _can_manage_trainers(request):
        return _dashboard_redirect("Only admins and master trainers can download trainer bulk templates.", "error")
    if record_type == "schools" and not _is_admin(request):
        return _dashboard_redirect("Only admin can download school bulk templates.", "error")

    templates = {
        "trainers": {
            "sheet": "Trainers",
            "headers": [
                "name", "email", "phone", "role", "division", "districts",
                "assigned_school_1", "assigned_school_2", "assigned_school_3", "password",
            ],
            "sample": [
                "Trainer Full Name", "trainer@example.com", "+91 9876543210", "atl_trainer",
                "Bengaluru", "Bengaluru South",
                "GHS Jayanagar", "GHS Basavanagudi", "GHS Malleshwaram", "Trainer@123",
            ],
        },
        "schools": {
            "sheet": "Schools",
            "headers": [
                "udise_code", "atl_lab_code", "name", "district", "division",
                "pin_code", "principal_name", "assigned_trainer",
            ],
            "sample": [
                "29XXXXXXXXX", "ATL-KA-999", "Government High School Example", "Mysore",
                "Mysuru", "570001", "Dr. Principal Name", "Trainer Full Name",
            ],
        },
        "students": {
            "sheet": "Students",
            "headers": [
                "name", "email", "grade", "school_udise_code", "father_name", "mother_name",
                "age", "gender", "religion", "category", "phone", "urban_rural", "income_status",
                "physically_challenged", "medium", "state", "district", "taluk", "village", "pincode",
            ],
            "sample": [
                "Student Full Name", "student@example.com", "8", "29XXXXXXXXX", "Father Name",
                "Mother Name", "13", "female", "Hindu", "GM", "+91 9876543210", "rural",
                "BPL", "no", "Kannada", "Karnataka", "Bengaluru Urban", "Bengaluru South",
                "Jayanagar", "560011",
            ],
        },
    }
    config = templates.get(record_type)
    if not config:
        raise HTTPException(404, "Unknown template type")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = config["sheet"]
    worksheet.append(config["headers"])
    worksheet.append(config["sample"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="student_dashboard_{record_type}_template.xlsx"'},
    )


@app.get("/export/{record_type}")
async def export_records(record_type: str, request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    search = request.query_params.get("q", "").strip()
    sort_key = request.query_params.get("sort", "").strip()
    direction = request.query_params.get("direction", "asc").strip().lower()
    reverse = direction == "desc"

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        if record_type == "trainers":
            if not _is_admin(request):
                return _dashboard_redirect("Only admin can export trainer records.", "error")
            headers = [
                "name", "email", "phone", "role", "division", "districts",
                "assigned_school_1", "assigned_school_2", "assigned_school_3",
            ]
            rows = []
            for trainer in db.query(models.Trainer).order_by(models.Trainer.name).all():
                schools = (_split_assigned_schools(trainer.assigned_school) + ["", "", ""])[:3]
                rows.append(
                    {
                        "name": trainer.name,
                        "email": trainer.email,
                        "phone": trainer.phone,
                        "role": trainer.role,
                        "division": trainer.division,
                        "districts": trainer.districts,
                        "assigned_school_1": schools[0],
                        "assigned_school_2": schools[1],
                        "assigned_school_3": schools[2],
                    }
                )
            filename = "student_dashboard_trainers.xlsx"
            sheet_name = "Trainers"
        elif record_type == "schools":
            headers = [
                "udise_code", "atl_lab_code", "name", "division", "district", "principal_name",
                "assigned_trainer",
            ]
            school_query = db.query(models.School)
            if scope_school_ids is not None:
                school_query = school_query.filter(models.School.id.in_(scope_school_ids))
            rows = [
                {
                    "udise_code": school.udise_code,
                    "atl_lab_code": school.atl_lab_code,
                    "name": school.name,
                    "division": school.division,
                    "district": school.district,
                    "principal_name": school.principal_name,
                    "assigned_trainer": school.assigned_trainer,
                }
                for school in school_query.all()
            ]
            filename = "student_dashboard_schools.xlsx"
            sheet_name = "Schools"
        elif record_type == "students":
            headers = [
                "name", "email", "grade", "school_name", "school_udise_code", "father_name",
                "mother_name", "age", "gender", "religion", "category", "phone", "urban_rural",
                "income_status", "physically_challenged", "medium", "state", "district",
                "taluk", "village", "pincode", "address",
            ]
            student_query = (
                db.query(models.Student, models.School.name.label("school_name"), models.School.udise_code.label("school_udise_code"))
                .outerjoin(models.School, models.School.id == models.Student.school_id)
            )
            if scope_school_ids is not None:
                student_query = student_query.filter(models.Student.school_id.in_(scope_school_ids))
            rows = [
                {
                    "name": student.name,
                    "email": student.email,
                    "grade": student.grade,
                    "school_name": school_name,
                    "school_udise_code": school_udise_code,
                    "father_name": student.father_name,
                    "mother_name": student.mother_name,
                    "age": student.age,
                    "gender": student.gender,
                    "religion": normalize_religion(student.caste) or student.caste,
                    "category": normalize_category(student.category) or student.category,
                    "phone": student.phone,
                    "urban_rural": student.urban_rural,
                    "income_status": student.income_status,
                    "physically_challenged": student.physically_challenged,
                    "medium": student.medium,
                    "state": student.state,
                    "district": student.district,
                    "taluk": student.taluk,
                    "village": student.village,
                    "pincode": student.pincode,
                    "address": student.address,
                }
                for student, school_name, school_udise_code in student_query.all()
            ]
            filename = "student_dashboard_students.xlsx"
            sheet_name = "Students"
        else:
            raise HTTPException(404, "Unknown export type")

    filtered_rows = [row for row in rows if _matches_search(row, search)]
    if sort_key in headers:
        filtered_rows.sort(key=lambda row: str(row.get(sort_key) or "").lower(), reverse=reverse)

    return _export_workbook(sheet_name, headers, filtered_rows, filename)


@app.post("/manual/trainer")
async def add_trainer(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    role: str = Form("atl_trainer"),
    division: str = Form(""),
    districts: str = Form(""),
    assigned_school: str = Form(""),
    assigned_school_1: str = Form(""),
    assigned_school_2: str = Form(""),
    assigned_school_3: str = Form(""),
    password: str = Form(""),
    confirm_password: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_trainers(request):
        return _dashboard_redirect("Only admins and master trainers can add trainers.", "error")

    normalized_email = email.strip().lower()
    cleaned_name = name.strip()
    allowed_trainer_roles = {"atl_trainer", "master_trainer"}
    if _is_admin(request):
        selected_role = role if role in allowed_trainer_roles else "atl_trainer"
        chosen_password = password
        if len(chosen_password) < 6:
            return _dashboard_redirect("Trainer login password must be at least 6 characters.", "error")
        if chosen_password != confirm_password:
            return _dashboard_redirect("Trainer login passwords do not match.", "error")
    else:
        selected_role = "atl_trainer"
        chosen_password = DEFAULT_TRAINER_PASSWORD

    if len(cleaned_name) < 2:
        return _dashboard_redirect("Trainer name is required.", "error")
    if "@" not in normalized_email:
        return _dashboard_redirect("Trainer email must be valid.", "error")

    with SessionLocal() as db:
        if db.query(models.Trainer).filter(func.lower(models.Trainer.email) == normalized_email).first():
            return _dashboard_redirect(f"Trainer email already exists: {email}", "error")
        if db.query(models.Account).filter(func.lower(models.Account.email) == normalized_email).first():
            return _dashboard_redirect(f"A login account already exists for: {email}", "error")

        selected_school_ids = [
            value for value in [assigned_school_1, assigned_school_2, assigned_school_3] if value
        ]
        conflict_message = _validate_trainer_assignment_selection(db, selected_school_ids)
        if conflict_message:
            return _dashboard_redirect(conflict_message, "error")

        assigned_schools = _join_assigned_schools(
            assigned_school_1,
            assigned_school_2,
            assigned_school_3,
            assigned_school,
        )
        db.add(
            models.Trainer(
                name=cleaned_name,
                email=normalized_email,
                phone=phone.strip() or None,
                role=selected_role,
                division=division.strip() or None,
                districts=districts.strip() or None,
                assigned_school=assigned_schools,
                specialization="ATL trainer",
                is_active=True,
            )
        )
        db.add(
            models.Account(
                name=cleaned_name,
                email=normalized_email,
                role="trainer",
                hashed_password=_hash_password(chosen_password),
                plain_password=chosen_password,
            )
        )
        db.commit()
    if _is_admin(request):
        return _dashboard_redirect("Trainer added successfully. Login credentials are ready.")
    return _dashboard_redirect(f"Trainer added successfully. Default login password is {DEFAULT_TRAINER_PASSWORD}.")


@app.post("/trainers/{trainer_id}/toggle-active")
async def toggle_trainer_active(request: Request, trainer_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_trainers(request):
        return _dashboard_redirect("Only admins and master trainers can activate or deactivate trainers.", "error")

    current_account = _current_account(request) or {}
    with SessionLocal() as db:
        trainer = db.query(models.Trainer).filter(models.Trainer.id == trainer_id).first()
        if not trainer:
            return _dashboard_redirect("Trainer was not found.", "error")
        if current_account.get("email") and trainer.email.lower() == current_account.get("email", "").lower():
            return _dashboard_redirect("You cannot deactivate your own account.", "error")
        if not _is_admin(request) and trainer.role == "master_trainer":
            return _dashboard_redirect("Only admin can activate or deactivate master trainers.", "error")

        current_active = True if trainer.is_active is None else bool(trainer.is_active)
        trainer.is_active = not current_active
        trainer_name = trainer.name
        db.commit()
        state = "activated" if trainer.is_active else "deactivated"

    return _dashboard_redirect(f"Trainer {trainer_name} has been {state}.")


@app.post("/trainers/{trainer_id}/update")
async def update_trainer(
    request: Request,
    trainer_id: int,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    role: str = Form("atl_trainer"),
    division: str = Form(""),
    districts: str = Form(""),
    assigned_school: str = Form(""),
    assigned_school_1: str = Form(""),
    assigned_school_2: str = Form(""),
    assigned_school_3: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_trainers(request):
        return _dashboard_redirect("Only admins and master trainers can edit trainer details.", "error")

    cleaned_name = name.strip()
    normalized_email = email.strip().lower()
    allowed_trainer_roles = {"atl_trainer", "master_trainer"}

    if len(cleaned_name) < 2:
        return _dashboard_redirect("Trainer name is required.", "error")
    if "@" not in normalized_email:
        return _dashboard_redirect("Trainer email must be valid.", "error")

    with SessionLocal() as db:
        trainer = db.query(models.Trainer).filter(models.Trainer.id == trainer_id).first()
        if not trainer:
            return _dashboard_redirect("Trainer was not found.", "error")
        if not _is_admin(request) and trainer.role == "master_trainer":
            return _dashboard_redirect("Only admin can edit master trainer roles.", "error")

        existing_trainer = (
            db.query(models.Trainer)
            .filter(models.Trainer.id != trainer_id, func.lower(models.Trainer.email) == normalized_email)
            .first()
        )
        if existing_trainer:
            return _dashboard_redirect(f"Trainer email already exists: {email}", "error")

        existing_account = db.query(models.Account).filter(func.lower(models.Account.email) == normalized_email).first()
        if existing_account and existing_account.email.lower() != trainer.email.lower():
            return _dashboard_redirect(f"A login account already exists for: {email}", "error")

        selected_school_ids = [
            value for value in [assigned_school_1, assigned_school_2, assigned_school_3] if value
        ]
        conflict_message = _validate_trainer_assignment_selection(db, selected_school_ids, trainer.id)
        if conflict_message:
            return _dashboard_redirect(conflict_message, "error")

        selected_role = trainer.role
        if _is_admin(request) and role in allowed_trainer_roles:
            selected_role = role

        trainer.name = cleaned_name
        trainer.email = normalized_email
        trainer.phone = phone.strip() or None
        trainer.role = selected_role
        trainer.division = division.strip() or None
        trainer.districts = districts.strip() or None
        trainer.assigned_school = _join_assigned_schools(*selected_school_ids)
        trainer.specialization = trainer.specialization or "ATL trainer"
        db.commit()

        account = db.query(models.Account).filter(func.lower(models.Account.email) == trainer.email.lower()).first()
        if account:
            account.name = cleaned_name
            account.email = normalized_email
            db.commit()

    return _dashboard_redirect(f"Trainer {cleaned_name} was updated successfully.")


@app.post("/manual/school")
async def add_school(
    request: Request,
    udise_code: str = Form(...),
    atl_lab_code: str = Form(""),
    name: str = Form(...),
    district: str = Form(""),
    division: str = Form(""),
    pin_code: str = Form(""),
    principal_name: str = Form(""),
    assigned_trainer: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _is_admin(request):
        return _dashboard_redirect("Only admin can add schools.", "error")

    with SessionLocal() as db:
        if db.query(models.School).filter(models.School.udise_code == udise_code.strip()).first():
            return _dashboard_redirect(f"School UDISE already exists: {udise_code}", "error")
        db.add(
            models.School(
                udise_code=udise_code.strip(),
                atl_lab_code=atl_lab_code.strip() or None,
                name=name.strip(),
                district=district.strip() or None,
                division=division.strip() or None,
                pin_code=pin_code.strip() or None,
                principal_name=principal_name.strip() or None,
                assigned_trainer=assigned_trainer.strip() or None,
            )
        )
        db.commit()
    return _dashboard_redirect("School added successfully.")


@app.post("/manual/student")
async def add_student(
    request: Request,
    name: str = Form(...),
    email: str = Form(""),
    grade: str = Form(""),
    school_udise_code: str = Form(""),
    father_name: str = Form(""),
    mother_name: str = Form(""),
    age: str = Form(""),
    gender: str = Form(""),
    religion: str = Form(""),
    category: str = Form(""),
    phone: str = Form(""),
    urban_rural: str = Form(""),
    income_status: str = Form(""),
    physically_challenged: str = Form(""),
    medium: str = Form(""),
    state: str = Form(""),
    district: str = Form(""),
    taluk: str = Form(""),
    village: str = Form(""),
    pincode: str = Form(""),
    address: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _is_atl_trainer_account(request):
        return _dashboard_redirect("Only trainers can add students. Admins and master trainers can view student details.", "error")

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        school = None
        if school_udise_code.strip():
            school = db.query(models.School).filter(models.School.udise_code == school_udise_code.strip()).first()
            if not school:
                return _dashboard_redirect(f"School UDISE not found: {school_udise_code}", "error")
            if not _school_in_scope(scope_school_ids, school.id):
                return _dashboard_redirect("You can add students only for your assigned schools.", "error")
        elif scope_school_ids is not None:
            return _dashboard_redirect("Trainer student records must be assigned to one of your schools.", "error")
        if email.strip() and db.query(models.Student).filter(func.lower(models.Student.email) == email.strip().lower()).first():
            return _dashboard_redirect(f"Student email already exists: {email}", "error")
        db.add(
            models.Student(
                name=name.strip(),
                email=email.strip().lower() or None,
                grade=grade.strip() or None,
                father_name=father_name.strip() or None,
                mother_name=mother_name.strip() or None,
                age=_parse_int(age),
                gender=gender.strip() or None,
                caste=normalize_religion(religion),
                category=normalize_category(category),
                phone=phone.strip() or None,
                urban_rural=urban_rural.strip() or None,
                income_status=income_status.strip() or None,
                physically_challenged=physically_challenged.strip() or None,
                medium=medium.strip() or None,
                state=state.strip() or None,
                district=district.strip() or None,
                taluk=taluk.strip() or None,
                village=village.strip() or None,
                pincode=pincode.strip() or None,
                address=_compose_student_address(state, district, taluk, village, pincode, address),
                school_id=school.id if school else None,
            )
        )
        db.commit()
    return _dashboard_redirect("Student added successfully.")


@app.post("/bulk-upload/{record_type}")
async def bulk_upload(record_type: str, request: Request, file: UploadFile = File(...)):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    if record_type not in {"trainers", "schools", "students"}:
        raise HTTPException(404, "Unknown bulk upload type")
    if record_type == "trainers" and not _can_manage_trainers(request):
        return _dashboard_redirect("Only admins and master trainers can bulk upload trainers.", "error")
    if record_type == "schools" and not _is_admin(request):
        return _dashboard_redirect("Only admin can bulk upload school records.", "error")
    if record_type == "students" and not _is_atl_trainer_account(request):
        return _dashboard_redirect("Only trainers can bulk upload students. Admins and master trainers can view student details.", "error")

    rows, col = _worksheet_rows(await file.read())
    success, failed, errors = 0, 0, []
    is_admin_user = _is_admin(request)

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        for index, row in enumerate(rows, start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            try:
                if record_type == "trainers":
                    name = col(row, "name") or col(row, "trainer_name")
                    email = (col(row, "email") or col(row, "trainer_email") or "").lower()
                    if not name or not email:
                        raise ValueError("missing name/email")
                    if db.query(models.Trainer).filter(func.lower(models.Trainer.email) == email).first():
                        raise ValueError(f"trainer email already exists: {email}")
                    if db.query(models.Account).filter(func.lower(models.Account.email) == email).first():
                        raise ValueError(f"login account already exists: {email}")
                    role_value = (col(row, "role") or "atl_trainer").strip().lower()
                    if not is_admin_user:
                        role_value = "atl_trainer"
                    elif role_value not in {"atl_trainer", "master_trainer"}:
                        role_value = "atl_trainer"
                    chosen_password = (col(row, "password") or col(row, "login_password") or "").strip()
                    if not chosen_password:
                        chosen_password = DEFAULT_TRAINER_PASSWORD
                    if len(chosen_password) < 6:
                        raise ValueError("password must be at least 6 characters")
                    db.add(
                        models.Trainer(
                            name=name,
                            email=email,
                            phone=col(row, "phone") or col(row, "trainer_phone") or None,
                            role=role_value,
                            division=col(row, "division") or None,
                            districts=col(row, "districts") or col(row, "district") or None,
                            assigned_school=_join_assigned_schools(
                                col(row, "assigned_school_1"),
                                col(row, "assigned_school_2"),
                                col(row, "assigned_school_3"),
                                col(row, "assigned_school"),
                            ),
                            specialization="ATL trainer",
                            is_active=True,
                        )
                    )
                    db.add(
                        models.Account(
                            name=name,
                            email=email,
                            role="trainer",
                            hashed_password=_hash_password(chosen_password),
                            plain_password=chosen_password,
                        )
                    )

                elif record_type == "schools":
                    udise = col(row, "udise_code")
                    name = col(row, "name")
                    if not udise or not name:
                        raise ValueError("missing udise_code/name")
                    if db.query(models.School).filter(models.School.udise_code == udise).first():
                        raise ValueError(f"school UDISE already exists: {udise}")
                    db.add(
                        models.School(
                            udise_code=udise,
                            atl_lab_code=col(row, "atl_lab_code") or None,
                            name=name,
                            district=col(row, "district") or None,
                            division=col(row, "division") or None,
                            pin_code=col(row, "pin_code") or None,
                            principal_name=col(row, "principal_name") or None,
                            assigned_trainer=col(row, "assigned_trainer") or col(row, "trainer_name") or None,
                        )
                    )

                else:
                    name = col(row, "name") or col(row, "student_name")
                    email = (col(row, "email") or col(row, "student_email")).lower()
                    school_udise = col(row, "school_udise_code") or col(row, "udise_code")
                    if not name:
                        raise ValueError("missing student name")
                    school = None
                    if school_udise:
                        school = db.query(models.School).filter(models.School.udise_code == school_udise).first()
                        if not school:
                            raise ValueError(f"school UDISE not found: {school_udise}")
                        if not _school_in_scope(scope_school_ids, school.id):
                            raise ValueError("student school is not assigned to this trainer")
                    elif scope_school_ids is not None:
                        raise ValueError("trainer student rows must include an assigned school UDISE")
                    if email and db.query(models.Student).filter(func.lower(models.Student.email) == email).first():
                        raise ValueError(f"student email already exists: {email}")
                    db.add(
                        models.Student(
                            name=name,
                            email=email or None,
                            grade=col(row, "grade") or None,
                            father_name=col(row, "father_name") or col(row, "fathers_name") or None,
                            mother_name=col(row, "mother_name") or col(row, "mothers_name") or None,
                            age=_parse_int(col(row, "age")),
                            gender=col(row, "gender") or None,
                            caste=normalize_religion(col(row, "religion") or col(row, "caste")),
                            category=normalize_category(col(row, "category")),
                            phone=col(row, "phone") or col(row, "phone_number") or None,
                            urban_rural=col(row, "urban_rural") or None,
                            income_status=col(row, "income_status") or None,
                            physically_challenged=col(row, "physically_challenged") or None,
                            medium=col(row, "medium") or col(row, "language_medium") or None,
                            state=col(row, "state") or None,
                            district=col(row, "district") or None,
                            taluk=col(row, "taluk") or None,
                            village=col(row, "village") or col(row, "ward") or None,
                            pincode=col(row, "pincode") or col(row, "pin_code") or None,
                            address=_compose_student_address(
                                col(row, "state"),
                                col(row, "district"),
                                col(row, "taluk"),
                                col(row, "village") or col(row, "ward"),
                                col(row, "pincode") or col(row, "pin_code"),
                                col(row, "address"),
                            ),
                            school_id=school.id if school else None,
                        )
                    )

                db.commit()
                success += 1
            except Exception as exc:
                db.rollback()
                failed += 1
                errors.append(f"Row {index}: {exc}")

    message = f"Bulk {record_type} upload complete: {success} success, {failed} failed."
    if errors:
        message += " " + " | ".join(errors[:3])
    return _dashboard_redirect(message, "error" if failed else "success")


@app.post("/courses/create")
async def create_course(
    request: Request,
    title: str = Form(...),
    level: str = Form(""),
    sector: str = Form(""),
    sub_sector: str = Form(""),
    occupation: str = Form(""),
    reference_id: str = Form(""),
    description: str = Form(""),
    pdf_file: UploadFile | None = File(None),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can add new courses.", "error")

    cleaned_title = title.strip()
    if not cleaned_title:
        return _dashboard_redirect("Course title is required.", "error")
    course_pdf_url = ""
    if pdf_file and pdf_file.filename:
        if Path(pdf_file.filename).suffix.lower() != ".pdf":
            return _dashboard_redirect("Course PDF must be a .pdf file.", "error")
        safe_name = _safe_upload_name(pdf_file.filename)
        content = await pdf_file.read()
        if not content:
            return _dashboard_redirect("Course PDF file is empty.", "error")
        course_pdf_url = _persist_lms_upload(safe_name, content)
    form = await request.form()
    item_titles = [str(value).strip() for value in form.getlist("item_title")]
    item_descriptions = [str(value).strip() for value in form.getlist("item_description")]
    lesson_items = [
        (item_title, item_descriptions[index] if index < len(item_descriptions) else "")
        for index, item_title in enumerate(item_titles)
        if item_title
    ]
    if not lesson_items:
        lesson_items = [("Course Overview", "Add PDFs, PPTs, videos, or notes for this course from the Content Library.")]

    with SessionLocal() as db:
        exists = db.query(models.Course).filter(func.lower(models.Course.title) == cleaned_title.lower()).first()
        if exists:
            return _dashboard_redirect(f"Course already exists: {cleaned_title}", "error")
        course = models.Course(
            title=cleaned_title,
            level=level.strip() or "Custom Course",
            sector=sector.strip() or None,
            sub_sector=sub_sector.strip() or None,
            occupation=occupation.strip() or None,
            reference_id=reference_id.strip() or None,
            resource_url=course_pdf_url or None,
            description=description.strip() or None,
        )
        db.add(course)
        db.flush()
        for index, (item_title, item_description) in enumerate(lesson_items, start=1):
            db.add(
                models.Lesson(
                    course_id=course.id,
                    title=item_title,
                    content_type="article",
                    content_body=item_description or "Add PDFs, PPTs, videos, or notes for this course from the Content Library.",
                    sort_order=index,
                )
            )
        db.commit()

    return _dashboard_redirect("New course added successfully.")


@app.post("/courses/{course_id}/pdf")
async def upload_course_pdf(request: Request, course_id: int, pdf_file: UploadFile = File(...)):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can add course PDFs.", "error")
    if not pdf_file or not pdf_file.filename:
        return _dashboard_redirect("Choose a PDF file to upload.", "error")
    if Path(pdf_file.filename).suffix.lower() != ".pdf":
        return _dashboard_redirect("Course PDF must be a .pdf file.", "error")

    content = await pdf_file.read()
    if not content:
        return _dashboard_redirect("Course PDF file is empty.", "error")

    safe_name = _safe_upload_name(pdf_file.filename)
    saved_url = _persist_lms_upload(safe_name, content)

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        course.resource_url = saved_url
        db.commit()

    return _dashboard_redirect("Course PDF added successfully.")


@app.post("/courses/{course_id}/delete")
async def delete_course(request: Request, course_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can delete courses.", "error")

    protected_titles = {"ATL Curriculum and Innovation Calendar 2026-27"}

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        if course.title in protected_titles or course.title.startswith("Experiment "):
            return _dashboard_redirect("Built-in handbook and seed courses cannot be deleted.", "error")

        lessons = db.query(models.Lesson).filter(models.Lesson.course_id == course.id).all()
        for lesson in lessons:
            db.query(models.LessonResource).filter(models.LessonResource.lesson_id == lesson.id).delete()
            db.delete(lesson)
        db.query(models.Enrollment).filter(models.Enrollment.course_id == course.id).delete()
        db.delete(course)
        db.commit()

    return _dashboard_redirect("Course deleted successfully.")


@app.post("/courses/{course_id}/update")
async def update_course(
    request: Request,
    course_id: int,
    title: str = Form(...),
    level: str = Form(""),
    sector: str = Form(""),
    sub_sector: str = Form(""),
    occupation: str = Form(""),
    reference_id: str = Form(""),
    description: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can modify courses.", "error")

    cleaned_title = title.strip()
    if not cleaned_title:
        return _dashboard_redirect("Course title is required.", "error")

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        duplicate = (
            db.query(models.Course)
            .filter(models.Course.id != course.id, func.lower(models.Course.title) == cleaned_title.lower())
            .first()
        )
        if duplicate:
            return _dashboard_redirect(f"Course already exists: {cleaned_title}", "error")
        course.title = cleaned_title
        course.level = level.strip() or "Custom Course"
        course.sector = sector.strip() or None
        course.sub_sector = sub_sector.strip() or None
        course.occupation = occupation.strip() or None
        course.reference_id = reference_id.strip() or None
        course.description = description.strip() or None
        db.commit()

    return _dashboard_redirect("Course updated successfully.")


@app.post("/courses/{course_id}/manage")
async def manage_course(request: Request, course_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can modify courses.", "error")

    form = await request.form()
    cleaned_title = str(form.get("title") or "").strip()
    if not cleaned_title:
        return _dashboard_redirect("Course title is required.", "error")

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        duplicate = (
            db.query(models.Course)
            .filter(models.Course.id != course.id, func.lower(models.Course.title) == cleaned_title.lower())
            .first()
        )
        if duplicate:
            return _dashboard_redirect(f"Course already exists: {cleaned_title}", "error")

        course.title = cleaned_title
        course.level = str(form.get("level") or "").strip() or "Custom Course"
        course.sector = str(form.get("sector") or "").strip() or None
        course.sub_sector = str(form.get("sub_sector") or "").strip() or None
        course.occupation = str(form.get("occupation") or "").strip() or None
        course.reference_id = str(form.get("reference_id") or "").strip() or None
        course.description = str(form.get("description") or "").strip() or None

        delete_ids = {_parse_int(value) for value in form.getlist("delete_lesson_id")}
        delete_ids.discard(None)
        existing_ids = [_parse_int(value) for value in form.getlist("existing_lesson_id")]
        existing_titles = [str(value).strip() for value in form.getlist("existing_item_title")]
        existing_descriptions = [str(value).strip() for value in form.getlist("existing_item_description")]

        for index, lesson_id in enumerate(existing_ids):
            if not lesson_id:
                continue
            lesson = (
                db.query(models.Lesson)
                .filter(models.Lesson.id == lesson_id, models.Lesson.course_id == course.id)
                .first()
            )
            if not lesson:
                continue
            if lesson_id in delete_ids:
                db.query(models.LessonResource).filter(models.LessonResource.lesson_id == lesson.id).delete()
                db.delete(lesson)
                continue
            title_value = existing_titles[index] if index < len(existing_titles) else ""
            description_value = existing_descriptions[index] if index < len(existing_descriptions) else ""
            if title_value:
                lesson.title = title_value
                lesson.content_body = description_value or None

        max_sort = db.query(func.max(models.Lesson.sort_order)).filter(models.Lesson.course_id == course.id).scalar() or 0
        new_titles = [str(value).strip() for value in form.getlist("new_item_title")]
        new_descriptions = [str(value).strip() for value in form.getlist("new_item_description")]
        for index, item_title in enumerate(new_titles):
            if not item_title:
                continue
            max_sort += 1
            item_description = new_descriptions[index] if index < len(new_descriptions) else ""
            db.add(
                models.Lesson(
                    course_id=course.id,
                    title=item_title,
                    content_type="article",
                    content_body=item_description or "Add PDFs, PPTs, videos, or notes for this item from the Content Library.",
                    sort_order=max_sort,
                )
            )

        db.commit()

    return _dashboard_redirect("Course changes saved successfully.")


@app.post("/courses/{course_id}/lessons/add")
async def add_course_lesson(
    request: Request,
    course_id: int,
    item_title: str = Form(...),
    item_description: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can add course items.", "error")

    cleaned_title = item_title.strip()
    if not cleaned_title:
        return _dashboard_redirect("Item title is required.", "error")

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        max_sort = db.query(func.max(models.Lesson.sort_order)).filter(models.Lesson.course_id == course.id).scalar() or 0
        db.add(
            models.Lesson(
                course_id=course.id,
                title=cleaned_title,
                content_type="article",
                content_body=item_description.strip() or "Add PDFs, PPTs, videos, or notes for this item from the Content Library.",
                sort_order=max_sort + 1,
            )
        )
        db.commit()

    return _dashboard_redirect("Course item added successfully.")


@app.post("/lessons/{lesson_id}/update")
async def update_lesson(
    request: Request,
    lesson_id: int,
    item_title: str = Form(...),
    item_description: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can modify course items.", "error")

    cleaned_title = item_title.strip()
    if not cleaned_title:
        return _dashboard_redirect("Item title is required.", "error")

    with SessionLocal() as db:
        lesson = db.query(models.Lesson).filter(models.Lesson.id == lesson_id).first()
        if not lesson:
            return _dashboard_redirect("Selected course item was not found.", "error")
        lesson.title = cleaned_title
        lesson.content_body = item_description.strip() or None
        db.commit()

    return _dashboard_redirect("Course item updated successfully.")


@app.post("/lessons/{lesson_id}/delete")
async def delete_lesson(request: Request, lesson_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_courses(request):
        return _dashboard_redirect("Only admin or master trainer can delete course items.", "error")

    with SessionLocal() as db:
        lesson = db.query(models.Lesson).filter(models.Lesson.id == lesson_id).first()
        if not lesson:
            return _dashboard_redirect("Selected course item was not found.", "error")
        db.query(models.LessonResource).filter(models.LessonResource.lesson_id == lesson.id).delete()
        db.delete(lesson)
        db.commit()

    return _dashboard_redirect("Course item deleted successfully.")


@app.post("/content/upload")
async def upload_course_content(
    request: Request,
    course_id: int = Form(...),
    lesson_id: int = Form(...),
    title: str = Form(""),
    content_type: str = Form(...),
    resource_url: str = Form(""),
    content_body: str = Form(""),
    file: UploadFile | None = File(None),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _is_admin(request):
        return _dashboard_redirect("Only admin can manage Content Library materials.", "error")

    cleaned_url = _normalize_resource_url(resource_url)

    try:
        _validate_content_upload(content_type, file, cleaned_url)
    except ValueError as exc:
        return _dashboard_redirect(str(exc), "error")

    saved_url = cleaned_url
    if content_type != "video_link" and file and file.filename:
        safe_name = _safe_upload_name(file.filename)
        content = await file.read()
        if not content:
            return _dashboard_redirect("Uploaded file is empty.", "error")
        saved_url = _persist_lms_upload(safe_name, content)

    with SessionLocal() as db:
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        lesson = (
            db.query(models.Lesson)
            .filter(models.Lesson.id == lesson_id, models.Lesson.course_id == course_id)
            .first()
        )
        if not lesson:
            return _dashboard_redirect("Selected lesson was not found for this course.", "error")
        db.add(
            models.LessonResource(
                lesson_id=lesson.id,
                title=title.strip() or f"{lesson.title} - {content_type.replace('_', ' ').title()}",
                content_type=content_type,
                content_body=content_body.strip() or f"Trainer uploaded {content_type.replace('_', ' ')} content.",
                resource_url=saved_url,
            )
        )
        db.commit()

    return _dashboard_redirect("Lesson resource uploaded successfully.")


@app.post("/content/update/{content_kind}/{content_id}")
async def update_course_content(
    request: Request,
    content_kind: str,
    content_id: int,
    title: str = Form(...),
    content_type: str = Form(...),
    resource_url: str = Form(""),
    content_body: str = Form(""),
    file: UploadFile | None = File(None),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _is_admin(request):
        return _dashboard_redirect("Only admin can modify Content Library materials.", "error")

    cleaned_title = title.strip()
    cleaned_url = _normalize_resource_url(resource_url)
    if not cleaned_title:
        return _dashboard_redirect("Lesson title is required.", "error")
    if content_type == "video_link":
        try:
            _validate_video_link(cleaned_url)
        except ValueError as exc:
            return _dashboard_redirect(str(exc), "error")

    try:
        _validate_optional_content_file(content_type, file)
        saved_url = await _save_content_upload(content_type, file, cleaned_url)
    except ValueError as exc:
        return _dashboard_redirect(str(exc), "error")

    with SessionLocal() as db:
        if content_kind == "resource":
            content_item = db.query(models.LessonResource).filter(models.LessonResource.id == content_id).first()
            lesson_id = content_item.lesson_id if content_item else None
        else:
            content_item = db.query(models.Lesson).filter(models.Lesson.id == content_id).first()
            lesson_id = content_item.id if content_item else None
        if not content_item:
            return _dashboard_redirect("Selected content was not found.", "error")
        if content_type != content_item.content_type:
            if not saved_url:
                return _dashboard_redirect("Choose a file or paste a valid link to add this content type.", "error")
            db.add(
                models.LessonResource(
                    lesson_id=lesson_id,
                    title=cleaned_title,
                    content_type=content_type,
                    content_body=content_body.strip() or f"Trainer uploaded {content_type.replace('_', ' ')} content.",
                    resource_url=saved_url,
                )
            )
            db.commit()
            return _dashboard_redirect("New content added without replacing existing content.")
        if content_type != "video_link" and not saved_url and not content_item.resource_url:
            return _dashboard_redirect("Please choose a file or keep an existing resource.", "error")
        content_item.title = cleaned_title
        content_item.content_type = content_type
        content_item.content_body = content_body.strip() or ""
        if saved_url:
            content_item.resource_url = saved_url
        db.commit()

    return _dashboard_redirect("Content updated successfully.")


@app.post("/content/delete/{content_kind}/{content_id}")
async def delete_course_content(request: Request, content_kind: str, content_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _is_admin(request):
        return _dashboard_redirect("Only admin can delete Content Library materials.", "error")

    with SessionLocal() as db:
        if content_kind == "resource":
            content_item = db.query(models.LessonResource).filter(models.LessonResource.id == content_id).first()
            if not content_item:
                return _dashboard_redirect("Selected content was not found.", "error")
            db.delete(content_item)
        else:
            lesson = db.query(models.Lesson).filter(models.Lesson.id == content_id).first()
            if not lesson:
                return _dashboard_redirect("Selected lesson was not found.", "error")
            lesson.content_type = "article"
            lesson.content_body = ""
            lesson.resource_url = None
        db.commit()

    return _dashboard_redirect("Content deleted successfully.")


@app.post("/enrollments/create")
async def create_enrollment(
    request: Request,
    student_id: int = Form(...),
    course_id: int = Form(...),
    batch_id: int = Form(...),
    assigned_by: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    current_account = _current_account(request) or {"name": "Trainer"}
    assigned_by_name = assigned_by.strip() or current_account.get("name") or "Trainer"

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        if not student:
            return _dashboard_redirect("Selected student was not found.", "error")
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        if not batch:
            return _dashboard_redirect("Selected batch was not found.", "error")
        if not _school_in_scope(scope_school_ids, student.school_id):
            return _dashboard_redirect("You can enroll students only from your assigned schools.", "error")
        if not _school_in_scope(scope_school_ids, batch.school_id):
            return _dashboard_redirect("You can enroll only into batches from your assigned schools.", "error")
        if student.school_id and student.school_id != batch.school_id:
            return _dashboard_redirect("Selected student does not belong to the selected batch school.", "error")
        existing = (
            db.query(models.Enrollment)
            .filter(
                models.Enrollment.student_id == student_id,
                models.Enrollment.course_id == course_id,
                models.Enrollment.batch_id == batch.id,
            )
            .first()
        )
        if existing:
            return _dashboard_redirect("This student is already enrolled in the selected course for this batch.", "error")
        db.add(
            models.Enrollment(
                student_id=student_id,
                course_id=course_id,
                batch_id=batch.id,
                status="assigned",
                progress=0,
                assigned_by=assigned_by_name,
            )
        )
        db.commit()

    return _dashboard_redirect("Student enrolled into course successfully.")


@app.post("/enrollments/update")
async def update_enrollment(
    request: Request,
    enrollment_id: int = Form(...),
    status: str = Form(...),
    progress: int = Form(...),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    allowed_statuses = {"assigned", "in_progress", "completed", "dropped"}
    selected_status = status if status in allowed_statuses else "assigned"
    cleaned_progress = max(0, min(100, progress))

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        enrollment = db.query(models.Enrollment).filter(models.Enrollment.id == enrollment_id).first()
        if not enrollment:
            return _dashboard_redirect("Enrollment was not found.", "error")
        student = db.query(models.Student).filter(models.Student.id == enrollment.student_id).first()
        if student and not _school_in_scope(scope_school_ids, student.school_id):
            return _dashboard_redirect("You can update enrollments only for your assigned schools.", "error")
        enrollment.status = selected_status
        enrollment.progress = cleaned_progress
        if selected_status == "completed" and cleaned_progress < 100:
            enrollment.progress = 100
        if selected_status == "completed":
            enrollment.completed_at = datetime.utcnow()
        db.commit()

    return _dashboard_redirect("Enrollment updated successfully.")


@app.post("/attendance/batches/create")
async def create_attendance_batch(
    request: Request,
    school_id: int = Form(...),
    name: str = Form(...),
    trainer_name: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    cleaned_name = name.strip()
    if not cleaned_name:
        return _dashboard_redirect("Batch name is required.", "error")
    parsed_start_date = _parse_date(start_date)
    parsed_end_date = _parse_date(end_date)
    if parsed_start_date and parsed_end_date and parsed_end_date < parsed_start_date:
        return _dashboard_redirect("Batch end date cannot be before start date.", "error")
    current_account = _current_account(request) or {"name": "Trainer", "role": "trainer"}
    if _is_atl_trainer_account(request):
        saved_trainer_name = current_account["name"]
    else:
        saved_trainer_name = trainer_name.strip() or current_account["name"]

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        school = db.query(models.School).filter(models.School.id == school_id).first()
        if not school:
            return _dashboard_redirect("Selected school was not found.", "error")
        if not _school_in_scope(scope_school_ids, school.id):
            return _dashboard_redirect("You can create batches only for your assigned schools.", "error")
        if not _is_atl_trainer_account(request) and not trainer_name.strip():
            return _dashboard_redirect("Select the trainer this batch belongs to.", "error")
        exists = (
            db.query(models.Batch)
            .filter(models.Batch.school_id == school_id, func.lower(models.Batch.name) == cleaned_name.lower())
            .first()
        )
        if exists:
            return _dashboard_redirect("This batch already exists for the selected school.", "error")
        db.add(
            models.Batch(
                school_id=school_id,
                name=cleaned_name,
                trainer_name=saved_trainer_name,
                start_date=parsed_start_date,
                end_date=parsed_end_date,
            )
        )
        db.commit()

    return _dashboard_redirect("Attendance batch created successfully.")


@app.post("/attendance/mark")
async def mark_attendance(
    request: Request,
    batch_id: int = Form(...),
    student_id: int = Form(...),
    attendance_date: str = Form(...),
    status: str = Form(...),
    marked_by: str = Form("Trainer"),
    remarks: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    selected_date = _parse_date(attendance_date)
    if not selected_date:
        return _dashboard_redirect("Attendance date is required.", "error")
    selected_status = status if status in {"present", "absent", "late", "excused"} else "present"

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        if not batch or not student:
            return _dashboard_redirect("Selected batch or student was not found.", "error")
        if not _school_in_scope(scope_school_ids, batch.school_id):
            return _dashboard_redirect("You can mark attendance only for your assigned schools.", "error")
        if student.school_id and student.school_id != batch.school_id:
            return _dashboard_redirect("Student does not belong to the selected batch school.", "error")
        record = (
            db.query(models.AttendanceRecord)
            .filter(
                models.AttendanceRecord.batch_id == batch_id,
                models.AttendanceRecord.student_id == student_id,
                models.AttendanceRecord.attendance_date == selected_date,
            )
            .first()
        )
        if not record:
            record = models.AttendanceRecord(batch_id=batch_id, student_id=student_id, attendance_date=selected_date)
            db.add(record)
        record.status = selected_status
        record.marked_by = marked_by.strip() or "Trainer"
        record.remarks = remarks.strip() or None
        db.commit()

    return _dashboard_redirect("Attendance saved successfully.")


@app.post("/attendance/mark-bulk")
async def mark_bulk_attendance(request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    form = await request.form()
    batch_id = _parse_int(form.get("batch_id"))
    selected_date = _parse_date(form.get("attendance_date"))
    marked_by = str(form.get("marked_by") or "Trainer").strip() or "Trainer"
    remarks = str(form.get("remarks") or "").strip() or None
    if not batch_id or not selected_date:
        return _dashboard_redirect("Batch and attendance date are required.", "error")

    saved_count = 0
    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        if not batch:
            return _dashboard_redirect("Selected batch was not found.", "error")
        if not _school_in_scope(scope_school_ids, batch.school_id):
            return _dashboard_redirect("You can mark attendance only for your assigned schools.", "error")
        students = (
            db.query(models.Student)
            .join(models.Enrollment, models.Enrollment.student_id == models.Student.id)
            .filter(models.Enrollment.batch_id == batch.id)
            .distinct()
            .order_by(models.Student.name)
            .all()
        )
        for student in students:
            selected_status = form.get(f"status_{student.id}")
            if selected_status not in {"present", "absent"}:
                continue
            record = (
                db.query(models.AttendanceRecord)
                .filter(
                    models.AttendanceRecord.batch_id == batch.id,
                    models.AttendanceRecord.student_id == student.id,
                    models.AttendanceRecord.attendance_date == selected_date,
                )
                .first()
            )
            if not record:
                record = models.AttendanceRecord(batch_id=batch.id, student_id=student.id, attendance_date=selected_date)
                db.add(record)
            record.status = selected_status
            record.marked_by = marked_by
            record.remarks = remarks
            saved_count += 1
        db.commit()

    if not saved_count:
        return _dashboard_redirect("No students were available to mark for this batch school.", "error")
    return _dashboard_redirect(f"Attendance saved for {saved_count} student(s).")


@app.post("/performance/batch-assessment")
async def save_batch_performance(
    request: Request,
    batch_id: int = Form(...),
    concept_understanding: str = Form("not_assessed"),
    project_understanding: str = Form("not_assessed"),
    design_thinking: str = Form("not_assessed"),
    assessed_by: str = Form("Trainer"),
    remarks: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        if not batch:
            return _dashboard_redirect("Selected batch was not found.", "error")
        if not _school_in_scope(scope_school_ids, batch.school_id):
            return _dashboard_redirect("You can assess only batches from your assigned schools.", "error")
        assessment = (
            db.query(models.BatchPerformanceAssessment)
            .filter(models.BatchPerformanceAssessment.batch_id == batch.id)
            .first()
        )
        if not assessment:
            assessment = models.BatchPerformanceAssessment(batch_id=batch.id)
            db.add(assessment)
        assessment.concept_understanding = _performance_level(concept_understanding)
        assessment.project_understanding = _performance_level(project_understanding)
        assessment.design_thinking = _performance_level(design_thinking)
        assessment.assessed_by = assessed_by.strip() or "Trainer"
        assessment.remarks = remarks.strip() or None
        assessment.assessed_at = datetime.utcnow()
        db.commit()

    return _dashboard_redirect("Batch performance assessment saved successfully.")


@app.post("/performance/teamwork-badges")
async def save_teamwork_badges(request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    form = await request.form()
    batch_id = _parse_int(form.get("batch_id"))
    course_id = _parse_int(form.get("course_id"))
    assigned_by = str(form.get("assigned_by") or "Trainer").strip() or "Trainer"
    remarks = str(form.get("remarks") or "").strip() or None
    if not batch_id:
        return _dashboard_redirect("Select a batch before assigning teamwork badges.", "error")
    if not course_id:
        return _dashboard_redirect("Select a course, experiment, or curriculum item before assigning teamwork badges.", "error")

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        if not batch:
            return _dashboard_redirect("Selected batch was not found.", "error")
        if not _school_in_scope(scope_school_ids, batch.school_id):
            return _dashboard_redirect("You can assign badges only for your assigned schools.", "error")
        course = db.query(models.Course).filter(models.Course.id == course_id).first()
        if not course:
            return _dashboard_redirect("Selected course was not found.", "error")
        saved_count = 0
        for badge_key in TEAMWORK_BADGES:
            student_id = _parse_int(form.get(f"badge_{badge_key}"))
            existing = (
                db.query(models.StudentTeamworkBadge)
                .filter(
                    models.StudentTeamworkBadge.batch_id == batch.id,
                    models.StudentTeamworkBadge.course_id == course.id,
                    models.StudentTeamworkBadge.badge == badge_key,
                )
                .all()
            )
            for badge_record in existing:
                db.delete(badge_record)
            if not student_id:
                continue
            student = db.query(models.Student).filter(models.Student.id == student_id).first()
            if not student or student.school_id != batch.school_id:
                return _dashboard_redirect("Selected badge student does not belong to this batch school.", "error")
            db.add(
                models.StudentTeamworkBadge(
                    batch_id=batch.id,
                    course_id=course.id,
                    student_id=student.id,
                    badge=badge_key,
                    assigned_by=assigned_by,
                    remarks=remarks,
                    assigned_at=datetime.utcnow(),
                )
            )
            saved_count += 1
        db.commit()

    return _dashboard_redirect(f"Teamwork badge assignment saved for {saved_count} badge(s).")


@app.get("/attendance/export")
async def export_attendance(
    request: Request,
    scope: str = "day",
    group: str = "individual",
    school_id: int | None = None,
    batch_id: int | None = None,
    attendance_date: str = "",
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    selected_date = _parse_date(attendance_date) if scope == "day" else None
    if scope == "day" and not selected_date:
        return _dashboard_redirect("Select a date before downloading day-wise attendance.", "error")

    with SessionLocal() as db:
        scope_school_ids = _request_school_scope_ids(request, db)
        if school_id and not _school_in_scope(scope_school_ids, school_id):
            return _dashboard_redirect("You can export attendance only for your assigned schools.", "error")
        if batch_id:
            batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
            if not batch:
                return _dashboard_redirect("Selected batch was not found.", "error")
            if not _school_in_scope(scope_school_ids, batch.school_id):
                return _dashboard_redirect("You can export attendance only for your assigned schools.", "error")
        query = (
            db.query(models.AttendanceRecord, models.Batch, models.Student, models.School)
            .join(models.Batch, models.Batch.id == models.AttendanceRecord.batch_id)
            .join(models.Student, models.Student.id == models.AttendanceRecord.student_id)
            .join(models.School, models.School.id == models.Batch.school_id)
        )
        if school_id:
            query = query.filter(models.Batch.school_id == school_id)
        elif scope_school_ids is not None:
            query = query.filter(models.Batch.school_id.in_(scope_school_ids))
        if batch_id:
            query = query.filter(models.AttendanceRecord.batch_id == batch_id)
        if selected_date:
            query = query.filter(models.AttendanceRecord.attendance_date == selected_date)
        records = query.order_by(models.School.name, models.Batch.name, models.Student.name, models.AttendanceRecord.attendance_date).all()

    if group == "batch":
        summary = {}
        for record, batch, _student, school in records:
            key = (school.name, batch.name)
            bucket = summary.setdefault(
                key,
                {"school": school.name, "batch": batch.name, "trainer": batch.trainer_name or "", "days": set(), "total": 0, "present": 0, "absent": 0},
            )
            bucket["days"].add(record.attendance_date)
            bucket["total"] += 1
            if record.status == "present":
                bucket["present"] += 1
            elif record.status == "absent":
                bucket["absent"] += 1
        rows = [
            {
                "school": item["school"],
                "batch": item["batch"],
                "trainer": item["trainer"],
                "days_marked": len(item["days"]),
                "total_records": item["total"],
                "present": item["present"],
                "absent": item["absent"],
                "present_percent": round((item["present"] / item["total"]) * 100) if item["total"] else 0,
            }
            for item in summary.values()
        ]
        return _export_workbook(
            "Batch Attendance",
            ["school", "batch", "trainer", "days_marked", "total_records", "present", "absent", "present_percent"],
            rows,
            f"attendance-{scope}-batch.xlsx",
        )

    summary = {}
    for record, batch, student, school in records:
        if scope == "day":
            summary[(batch.id, student.id, record.attendance_date)] = {
                "date": record.attendance_date,
                "school": school.name,
                "batch": batch.name,
                "student": student.name,
                "status": record.status,
                "marked_by": record.marked_by or "",
                "remarks": record.remarks or "",
            }
        else:
            key = (batch.id, student.id)
            bucket = summary.setdefault(
                key,
                {"school": school.name, "batch": batch.name, "student": student.name, "total": 0, "present": 0, "absent": 0},
            )
            bucket["total"] += 1
            if record.status == "present":
                bucket["present"] += 1
            elif record.status == "absent":
                bucket["absent"] += 1

    if scope == "day":
        rows = list(summary.values())
        headers = ["date", "school", "batch", "student", "status", "marked_by", "remarks"]
    else:
        rows = [
            {
                **item,
                "present_percent": round((item["present"] / item["total"]) * 100) if item["total"] else 0,
            }
            for item in summary.values()
        ]
        headers = ["school", "batch", "student", "total", "present", "absent", "present_percent"]
    return _export_workbook("Individual Attendance", headers, rows, f"attendance-{scope}-individual.xlsx")


@app.get("/reports/export/{report_type}")
async def export_generated_report(
    report_type: str,
    request: Request,
    school_id: str = "",
    district: str = "",
    taluk: str = "",
    village: str = "",
    trainer: str = "",
    medium: str = "",
    dimension: str = "",
    start_date: str = "",
    end_date: str = "",
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if report_type not in REPORT_TABLE_COLUMNS:
        raise HTTPException(404, "Unknown report type")

    filters = {
        "school_id": school_id,
        "district": district,
        "taluk": taluk,
        "village": village,
        "trainer": trainer,
        "medium": medium,
        "dimension": dimension,
        "gender": request.query_params.get("gender", ""),
        "income_status": request.query_params.get("income_status", ""),
        "physically_challenged": request.query_params.get("physically_challenged", ""),
        "urban_rural": request.query_params.get("urban_rural", ""),
        "religion": request.query_params.get("religion", "") or request.query_params.get("caste", ""),
        "category": request.query_params.get("category", ""),
    }
    parsed_start = _parse_date(start_date)
    parsed_end = _parse_date(end_date)

    with SessionLocal() as db:
        reports_data = _build_reports_payload(db, _current_account(request) or {"role": "student"})
        rows = reports_data["tables"].get(report_type, [])
        detail_source_rows = reports_data["tables"].get("student_detail_list", [])

    filtered_rows = [
        row
        for row in rows
        if _matches_report_filters(row, filters)
        and _date_in_report_range(row.get("start_date") or row.get("date"), parsed_start, parsed_end)
    ]
    headers = [key for key, _label in REPORT_TABLE_COLUMNS[report_type]]
    sheet_name = REPORT_TABLE_TITLES[report_type][:31]
    filename = f"report-{report_type}.xlsx"

    if report_type == "student_combination_summary":
        # Prefer students that belong to the visible summary combinations (same counts as UI).
        # Also keep direct filter matches so location filters (taluk/village/etc.) still apply.
        location_filters = {
            key: filters.get(key, "")
            for key in ("taluk", "village", "trainer", "religion", "category")
            if filters.get(key)
        }
        detail_rows = []
        seen = set()
        for row in detail_source_rows:
            if filtered_rows:
                in_visible_combo = any(_student_matches_combination(row, summary) for summary in filtered_rows)
                if not in_visible_combo:
                    continue
                if location_filters and not _matches_report_filters(row, location_filters):
                    continue
            elif not _matches_report_filters(row, filters):
                continue
            marker = (
                str(row.get("name") or ""),
                str(row.get("email") or ""),
                str(row.get("school_id") or ""),
                str(row.get("phone") or ""),
            )
            if marker in seen:
                continue
            seen.add(marker)
            detail_rows.append(row)
        detail_rows = sorted(detail_rows, key=lambda row: str(row.get("name") or "").lower())
        return _export_student_combination_workbook(filtered_rows, detail_rows, filename)

    return _export_workbook(sheet_name, headers, filtered_rows, filename)


@app.get("/dashboard")
async def dashboard(request: Request):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    current_account = _current_account(request) or {"name": "User", "email": "", "role": "student"}
    is_admin = current_account["role"] == "admin"

    with SessionLocal() as db:
        trainers_query = db.query(models.Trainer)
        if current_account.get("role") == "trainer" and not _is_master_trainer(request):
            trainers_query = trainers_query.filter(func.lower(models.Trainer.email) == current_account.get("email", "").lower())
        trainers = trainers_query.order_by(models.Trainer.role.desc(), models.Trainer.name).all()
        for trainer in trainers:
            trainer.assigned_school_display = _format_assigned_schools(db, trainer.assigned_school)
        trainer_password_map = {}
        if _can_manage_trainers(request):
            trainer_emails = [trainer.email.lower() for trainer in trainers if trainer.email]
            if trainer_emails:
                accounts = (
                    db.query(models.Account)
                    .filter(func.lower(models.Account.email).in_(trainer_emails))
                    .all()
                )
                trainer_password_map = {
                    account.email.lower(): account.plain_password or DEFAULT_TRAINER_PASSWORD
                    for account in accounts
                }
        all_schools = db.query(models.School).order_by(models.School.division, models.School.district, models.School.name).all()
        scope_school_ids = _trainer_school_scope_ids(db, current_account)
        schools = [school for school in all_schools if _school_in_scope(scope_school_ids, school.id)]
        assignment_schools = all_schools if _can_manage_trainers(request) else schools
        school_assignment_map = _active_school_assignment_map(db)
        unassigned_schools = [school for school in assignment_schools if school.id not in school_assignment_map]
        is_atl_trainer = bool(
            current_account.get("role") == "trainer"
            and scope_school_ids is not None
        )
        students_query = (
            db.query(models.Student, models.School.name.label("school_name"), models.School.udise_code.label("school_udise_code"))
            .outerjoin(models.School, models.School.id == models.Student.school_id)
        )
        if scope_school_ids is not None:
            students_query = students_query.filter(models.Student.school_id.in_(scope_school_ids))
        students = students_query.order_by(models.Student.name).all()
        student_rows = [
            {
                "id": student.id,
                "name": student.name,
                "email": student.email,
                "grade": student.grade,
                "father_name": student.father_name,
                "mother_name": student.mother_name,
                "age": student.age,
                "gender": student.gender,
                "religion": normalize_religion(student.caste) or student.caste,
                "category": normalize_category(student.category) or student.category,
                "category_display": category_display(normalize_category(student.category) or student.category),
                "phone": student.phone,
                "urban_rural": student.urban_rural,
                "income_status": student.income_status,
                "physically_challenged": student.physically_challenged,
                "medium": student.medium,
                "state": student.state,
                "district": student.district,
                "taluk": student.taluk,
                "village": student.village,
                "pincode": student.pincode,
                "address": student.address,
                "school_id": student.school_id,
                "school_name": school_name,
                "school_udise_code": school_udise_code,
            }
            for student, school_name, school_udise_code in students
        ]
        students_count = len(student_rows)
        courses_count = db.query(func.count(models.Course.id)).scalar() or 0
        lessons_count = db.query(func.count(models.Lesson.id)).scalar() or 0
        courses = db.query(models.Course).order_by(models.Course.id).all()
        course_rows = [
            {
                "id": course.id,
                "title": course.title,
                "description": course.description,
                "level": course.level,
                "lesson_count": db.query(func.count(models.Lesson.id)).filter(models.Lesson.course_id == course.id).scalar() or 0,
            }
            for course in courses
        ]
        lesson_options_by_course = {
            str(course.id): [
                {"id": lesson.id, "title": lesson.title}
                for lesson in db.query(models.Lesson)
                .filter(models.Lesson.course_id == course.id)
                .order_by(models.Lesson.sort_order, models.Lesson.id)
                .all()
            ]
            for course in courses
        }
        volume_columns = []
        handbook_titles = {volume["title"] for volume in VOLUME_COURSES}
        course_by_id = {course.id: course for course in courses}
        for volume in VOLUME_COURSES:
            course = next((row for row in course_rows if row["title"] == volume["title"]), None)
            if not course:
                continue
            course_model = course_by_id.get(course["id"])
            lessons = (
                db.query(models.Lesson)
                .filter(models.Lesson.course_id == course["id"])
                .order_by(models.Lesson.sort_order, models.Lesson.id)
                .all()
            )
            volume_columns.append(
                {
                    "id": course["id"],
                    "is_custom": False,
                    "can_delete": True,
                    "title": volume["title"],
                    "level": volume["level"],
                    "description": volume["description"],
                    "sector": course_model.sector if course_model else "",
                    "sub_sector": course_model.sub_sector if course_model else "",
                    "occupation": course_model.occupation if course_model else "",
                    "reference_id": course_model.reference_id if course_model else "",
                    "resource_url": volume["resource_url"],
                    "lesson_count": len(lessons),
                    "lessons": [
                        {
                            "id": lesson.id,
                            "title": lesson.title,
                            "content_type": lesson.content_type,
                            "content_body": lesson.content_body,
                            "resource_url": lesson.resource_url,
                        }
                        for lesson in lessons
                    ],
                }
            )
        for course in courses:
            if course.title in handbook_titles or course.title.startswith("Experiment ") or course.title == "ATL Curriculum and Innovation Calendar 2026-27":
                continue
            lessons = (
                db.query(models.Lesson)
                .filter(models.Lesson.course_id == course.id)
                .order_by(models.Lesson.sort_order, models.Lesson.id)
                .all()
            )
            volume_columns.append(
                {
                    "id": course.id,
                    "is_custom": True,
                    "can_delete": True,
                    "title": course.title,
                    "level": course.level or "Custom Course",
                    "description": course.description or "",
                    "sector": course.sector or "",
                    "sub_sector": course.sub_sector or "",
                    "occupation": course.occupation or "",
                    "reference_id": course.reference_id or "",
                    "resource_url": course.resource_url or "",
                    "lesson_count": len(lessons),
                    "lessons": [
                        {
                            "id": lesson.id,
                            "title": lesson.title,
                            "content_type": lesson.content_type,
                            "content_body": lesson.content_body,
                            "resource_url": lesson.resource_url,
                        }
                        for lesson in lessons
                    ],
                }
            )
        content_rows = []
        for course in courses:
            lessons = (
                db.query(models.Lesson)
                .filter(models.Lesson.course_id == course.id)
                .order_by(models.Lesson.sort_order, models.Lesson.id)
                .all()
            )
            for lesson in lessons:
                resources = (
                    db.query(models.LessonResource)
                    .filter(models.LessonResource.lesson_id == lesson.id)
                    .order_by(models.LessonResource.created_at, models.LessonResource.id)
                    .all()
                )
                lesson_resources = [
                    {
                        "id": resource.id,
                        "kind": "resource",
                        "title": resource.title,
                        "content_type": resource.content_type,
                        "content_body": resource.content_body or "",
                        "resource_url": resource.resource_url or "",
                    }
                    for resource in resources
                ]
                if lesson.resource_url or lesson.content_type != "article":
                    lesson_resources.insert(
                        0,
                        {
                            "id": lesson.id,
                            "kind": "lesson",
                            "title": lesson.title,
                            "content_type": lesson.content_type,
                            "content_body": lesson.content_body or "",
                            "resource_url": lesson.resource_url or "",
                        },
                    )
                content_rows.append(
                    {
                        "id": lesson.id,
                        "kind": "lesson",
                        "course_title": course.title,
                        "lesson_title": lesson.title,
                        "content_title": lesson.title,
                        "content_type": "multiple" if lesson_resources else lesson.content_type,
                        "content_body": f"{len(lesson_resources)} uploaded content item(s)." if lesson_resources else (lesson.content_body or ""),
                        "resource_url": "",
                        "resources": lesson_resources,
                        "group": _content_group(course.title, lesson.content_type),
                    }
                )
        content_counts = {
            "all": len(content_rows),
            "volume-1": sum(1 for row in content_rows if row["group"] == "volume-1"),
            "volume-2": sum(1 for row in content_rows if row["group"] == "volume-2"),
            "volume-3": sum(1 for row in content_rows if row["group"] == "volume-3"),
            "uploaded": sum(1 for row in content_rows if row["group"] == "uploaded"),
            "experiments": sum(1 for row in content_rows if row["group"] == "experiments"),
            "curriculum": sum(1 for row in content_rows if row["group"] == "curriculum"),
        }
        enrollments_query = (
            db.query(models.Enrollment, models.Student, models.Course, models.School, models.Batch)
            .join(models.Student, models.Student.id == models.Enrollment.student_id)
            .join(models.Course, models.Course.id == models.Enrollment.course_id)
            .outerjoin(models.School, models.School.id == models.Student.school_id)
            .outerjoin(models.Batch, models.Batch.id == models.Enrollment.batch_id)
        )
        if scope_school_ids is not None:
            enrollments_query = enrollments_query.filter(models.Student.school_id.in_(scope_school_ids))
        enrollments = enrollments_query.order_by(models.Enrollment.assigned_at.desc()).all()
        enrollment_rows = [
            {
                "id": enrollment.id,
                "student_name": student.name,
                "student_email": student.email,
                "school_name": school.name if school else None,
                "batch_name": batch.name if batch else None,
                "course_title": course.title,
                "status": enrollment.status,
                "progress": enrollment.progress or 0,
                "assigned_by": enrollment.assigned_by,
            }
            for enrollment, student, course, school, batch in enrollments
        ]
        batch_student_map = {}
        for enrollment, student, _course, _school, _batch in enrollments:
            if enrollment.batch_id:
                batch_student_map.setdefault(enrollment.batch_id, set()).add(student.id)
        batch_student_map = {batch_id: sorted(student_ids) for batch_id, student_ids in batch_student_map.items()}
        batches_query = (
            db.query(models.Batch, models.School)
            .join(models.School, models.School.id == models.Batch.school_id)
        )
        if scope_school_ids is not None:
            trainer_name_l = (current_account.get("name") or "").strip().lower()
            batch_filters = []
            if scope_school_ids:
                batch_filters.append(models.Batch.school_id.in_(scope_school_ids))
            if trainer_name_l:
                batch_filters.append(func.lower(models.Batch.trainer_name) == trainer_name_l)
            if batch_filters:
                batches_query = batches_query.filter(or_(*batch_filters))
            else:
                batches_query = batches_query.filter(models.Batch.id == -1)
        batches = batches_query.order_by(models.School.name, models.Batch.name).all()
        batch_rows = [
            {
                "id": batch.id,
                "school_id": batch.school_id,
                "school_name": school.name,
                "name": batch.name,
                "trainer_name": batch.trainer_name,
                "schedule": batch.schedule,
                "start_date": batch.start_date,
                "end_date": batch.end_date,
            }
            for batch, school in batches
        ]
        attendance_query = (
            db.query(models.AttendanceRecord, models.Batch, models.Student, models.School)
            .join(models.Batch, models.Batch.id == models.AttendanceRecord.batch_id)
            .join(models.Student, models.Student.id == models.AttendanceRecord.student_id)
            .join(models.School, models.School.id == models.Batch.school_id)
        )
        if scope_school_ids is not None:
            attendance_query = attendance_query.filter(models.Batch.school_id.in_(scope_school_ids))
        attendance_records = attendance_query.order_by(models.AttendanceRecord.attendance_date.desc(), models.School.name, models.Batch.name, models.Student.name).all()
        report_map = {}
        student_report_map = {}
        batch_performance_map = {}
        attendance_status_map = {}
        attendance_rows = []
        for record, batch, student, school in attendance_records:
            attendance_status_map[f"{batch.id}:{record.attendance_date.isoformat()}:{student.id}"] = record.status
            school_bucket = report_map.setdefault(
                school.id,
                {"school_id": school.id, "school_name": school.name, "batches": {}, "total": 0, "present": 0, "absent": 0, "late": 0, "excused": 0},
            )
            batch_bucket = school_bucket["batches"].setdefault(
                batch.id,
                {"batch_id": batch.id, "batch_name": batch.name, "trainer_name": batch.trainer_name, "total": 0, "present": 0, "absent": 0, "late": 0, "excused": 0, "dates": set()},
            )
            student_bucket = student_report_map.setdefault(
                (batch.id, student.id),
                {
                    "school_id": school.id,
                    "batch_id": batch.id,
                    "student_name": student.name,
                    "school_name": school.name,
                    "batch_name": batch.name,
                    "total": 0,
                    "present": 0,
                    "absent": 0,
                    "late": 0,
                    "excused": 0,
                },
            )
            performance_bucket = batch_performance_map.setdefault(
                batch.id,
                {
                    "school_id": school.id,
                    "batch_id": batch.id,
                    "school_name": school.name,
                    "batch_name": batch.name,
                    "trainer_name": batch.trainer_name,
                    "total": 0,
                    "present": 0,
                    "absent": 0,
                    "late": 0,
                    "excused": 0,
                    "dates": set(),
                },
            )
            school_bucket["total"] += 1
            batch_bucket["total"] += 1
            batch_bucket["dates"].add(record.attendance_date)
            student_bucket["total"] += 1
            performance_bucket["total"] += 1
            performance_bucket["dates"].add(record.attendance_date)
            if record.status in {"present", "absent", "late", "excused"}:
                school_bucket[record.status] += 1
                batch_bucket[record.status] += 1
                student_bucket[record.status] += 1
                performance_bucket[record.status] += 1
            attendance_rows.append(
                {
                    "date": record.attendance_date,
                    "school_id": school.id,
                    "batch_id": batch.id,
                    "school_name": school.name,
                    "batch_name": batch.name,
                    "student_name": student.name,
                    "status": record.status,
                    "marked_by": record.marked_by,
                    "remarks": record.remarks,
                }
            )
        attendance_report = [
            {
                **school_data,
                "batches": [
                    {
                        **batch_data,
                        "days_marked": len(batch_data["dates"]),
                        "present_percent": round((batch_data["present"] / batch_data["total"]) * 100) if batch_data["total"] else 0,
                    }
                    for batch_data in school_data["batches"].values()
                ],
            }
            for school_data in report_map.values()
        ]
        individual_attendance_report = [
            {
                **student_data,
                "present_percent": round((student_data["present"] / student_data["total"]) * 100) if student_data["total"] else 0,
            }
            for student_data in student_report_map.values()
        ]
        batch_performance_report = [
            {
                **batch_data,
                "days_marked": len(batch_data["dates"]),
                "present_percent": round((batch_data["present"] / batch_data["total"]) * 100) if batch_data["total"] else 0,
            }
            for batch_data in batch_performance_map.values()
        ]
        batch_assessments_query = db.query(models.BatchPerformanceAssessment).join(models.Batch, models.Batch.id == models.BatchPerformanceAssessment.batch_id)
        if scope_school_ids is not None:
            batch_assessments_query = batch_assessments_query.filter(models.Batch.school_id.in_(scope_school_ids))
        batch_assessments = batch_assessments_query.all()
        batch_assessment_map = {assessment.batch_id: assessment for assessment in batch_assessments}
        teamwork_badges_query = (
            db.query(models.StudentTeamworkBadge, models.Batch, models.Student, models.School, models.Course)
            .join(models.Batch, models.Batch.id == models.StudentTeamworkBadge.batch_id)
            .join(models.Student, models.Student.id == models.StudentTeamworkBadge.student_id)
            .join(models.School, models.School.id == models.Batch.school_id)
            .outerjoin(models.Course, models.Course.id == models.StudentTeamworkBadge.course_id)
        )
        if scope_school_ids is not None:
            teamwork_badges_query = teamwork_badges_query.filter(models.Batch.school_id.in_(scope_school_ids))
        teamwork_badges = teamwork_badges_query.order_by(models.School.name, models.Batch.name, models.Course.title, models.StudentTeamworkBadge.badge).all()
        teamwork_badge_map = {
            f"{badge.batch_id}:{badge.course_id or ''}:{badge.badge}": badge.student_id
            for badge, _batch, _student, _school, _course in teamwork_badges
        }
        batch_assessment_form_map = {
            str(assessment.batch_id): {
                "concept_understanding": assessment.concept_understanding,
                "project_understanding": assessment.project_understanding,
                "design_thinking": assessment.design_thinking,
                "assessed_by": assessment.assessed_by or "",
                "remarks": assessment.remarks or "",
            }
            for assessment in batch_assessments
        }
        teamwork_badge_rows = [
            {
                "school_id": school.id,
                "school_name": school.name,
                "batch_id": batch.id,
                "batch_name": batch.name,
                "course_id": course.id if course else None,
                "course_title": course.title if course else "Not linked",
                "student_name": student.name,
                "badge": badge.badge,
                "badge_label": TEAMWORK_BADGES.get(badge.badge, badge.badge),
                "assigned_by": badge.assigned_by,
                "remarks": badge.remarks,
            }
            for badge, batch, student, school, course in teamwork_badges
        ]
        performance_assessment_rows = [
            {
                "school_id": batch["school_id"],
                "school_name": batch["school_name"],
                "batch_id": batch["id"],
                "batch_name": batch["name"],
                "trainer_name": batch["trainer_name"],
                "concept_understanding": (batch_assessment_map.get(batch["id"]).concept_understanding if batch_assessment_map.get(batch["id"]) else "not_assessed"),
                "project_understanding": (batch_assessment_map.get(batch["id"]).project_understanding if batch_assessment_map.get(batch["id"]) else "not_assessed"),
                "design_thinking": (batch_assessment_map.get(batch["id"]).design_thinking if batch_assessment_map.get(batch["id"]) else "not_assessed"),
                "assessed_by": (batch_assessment_map.get(batch["id"]).assessed_by if batch_assessment_map.get(batch["id"]) else None),
                "remarks": (batch_assessment_map.get(batch["id"]).remarks if batch_assessment_map.get(batch["id"]) else None),
            }
            for batch in batch_rows
        ]
        total_student_count = sum(school.current_students or 0 for school in schools)
        reports_data = _build_reports_payload(db, current_account)
        dashboard_analytics = _build_dashboard_analytics(student_rows, batch_rows, enrollment_rows)

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "app_name": "GTTC Student Dashboard",
            "module_name": "Student + LMS Extension",
            "current_account": current_account,
            "is_admin": is_admin,
            "is_atl_trainer": is_atl_trainer,
            "trainers": trainers,
            "trainer_password_map": trainer_password_map,
            "schools": schools,
            "assignment_schools": assignment_schools,
            "unassigned_schools": unassigned_schools,
            "school_assignment_map": {str(school_id): trainer_id for school_id, trainer_id in school_assignment_map.items()},
            "trainers_count": len(trainers),
            "schools_count": len(schools),
            "students_count": students_count,
            "students": student_rows,
            "courses_count": courses_count,
            "lessons_count": lessons_count,
            "courses": course_rows,
            "lesson_options_by_course": lesson_options_by_course,
            "volume_columns": volume_columns,
            "content_rows": content_rows,
            "content_counts": content_counts,
            "enrollments": enrollment_rows,
            "enrollments_count": len(enrollment_rows),
            "batches": batch_rows,
            "batch_student_map": batch_student_map,
            "attendance_rows": attendance_rows[:30],
            "attendance_report": attendance_report,
            "individual_attendance_report": individual_attendance_report,
            "batch_performance_report": batch_performance_report,
            "performance_levels": [
                ("not_assessed", "Not Assessed"),
                ("needs_support", "Needs Support"),
                ("developing", "Developing"),
                ("proficient", "Proficient"),
                ("excellent", "Excellent"),
            ],
            "teamwork_badges": TEAMWORK_BADGES,
            "teamwork_badge_map": teamwork_badge_map,
            "batch_assessment_form_map": batch_assessment_form_map,
            "teamwork_badge_rows": teamwork_badge_rows,
            "performance_assessment_rows": performance_assessment_rows,
            "atl_forms": _build_atl_forms(db),
            "can_manage_forms": _can_manage_forms(request),
            "can_manage_courses": _can_manage_courses(request),
            "can_manage_trainers": _can_manage_trainers(request),
            "profile_details": _build_profile_details(db, current_account),
            "can_change_password": current_account.get("role") == "trainer",
            "default_trainer_password": DEFAULT_TRAINER_PASSWORD,
            "reports_data": reports_data,
            "religion_options": RELIGION_OPTIONS,
            "category_options": CATEGORY_OPTIONS,
            "category_label_map": {code: f"{code} — {label}" for code, label in CATEGORY_OPTIONS},
            "dashboard_analytics": dashboard_analytics,
            "attendance_status_map": attendance_status_map,
            "today": date.today().isoformat(),
            "total_student_count": total_student_count,
            "notice": request.query_params.get("notice"),
            "notice_kind": request.query_params.get("notice_kind", "success"),
        },
    )


@app.post("/profile/password")
async def change_profile_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    current_account = _current_account(request)
    if not current_account or current_account.get("role") != "trainer":
        return _dashboard_redirect("Password change is only available for trainer accounts.", "error")
    if len(new_password) < 6:
        return _dashboard_redirect("New password must be at least 6 characters.", "error")
    if new_password != confirm_password:
        return _dashboard_redirect("New passwords do not match.", "error")

    with SessionLocal() as db:
        account = (
            db.query(models.Account)
            .filter(func.lower(models.Account.email) == current_account.get("email", "").lower())
            .first()
        )
        if not account or not _verify_password(current_password, account.hashed_password):
            return _dashboard_redirect("Current password is incorrect.", "error")
        account.hashed_password = _hash_password(new_password)
        account.plain_password = new_password
        db.commit()

    return _dashboard_redirect("Password changed successfully.")


@app.post("/forms/upload")
async def upload_atl_form(
    request: Request,
    title: str = Form(...),
    code: str = Form(""),
    description: str = Form(""),
    file: UploadFile = File(...),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_forms(request):
        return _dashboard_redirect("Only admin or master trainers can manage forms.", "error")

    cleaned_title = title.strip()
    cleaned_code = code.strip().upper()
    cleaned_description = _clean_form_description(description)
    if len(cleaned_title) < 3:
        return _dashboard_redirect("Form title must be at least 3 characters.", "error")

    try:
        content = await _read_form_upload(file)
    except ValueError as exc:
        return _dashboard_redirect(str(exc), "error")

    stored_name = _safe_upload_name(file.filename)
    _write_form_file(stored_name, content)
    current_account = _current_account(request) or {"name": "Admin"}

    with SessionLocal() as db:
        existing = None
        if cleaned_code:
            existing = (
                db.query(models.AtlForm)
                .filter(func.upper(models.AtlForm.code) == cleaned_code)
                .first()
            )
        if existing:
            _remove_form_file(existing.stored_name)
            existing.title = cleaned_title
            existing.description = cleaned_description
            existing.filename = Path(file.filename).name
            existing.stored_name = stored_name
            existing.uploaded_by = current_account.get("name")
            existing.is_hidden = False
        else:
            db.add(
                models.AtlForm(
                    code=cleaned_code or None,
                    title=cleaned_title,
                    description=cleaned_description,
                    filename=Path(file.filename).name,
                    stored_name=stored_name,
                    uploaded_by=current_account.get("name"),
                    is_hidden=False,
                )
            )
        db.commit()

    if cleaned_code in _official_form_codes():
        return _dashboard_redirect(f"Official form {cleaned_code} updated successfully.")
    return _dashboard_redirect("Form uploaded successfully.")


@app.post("/forms/official/{form_code}/replace")
async def replace_official_atl_form(
    request: Request,
    form_code: str,
    file: UploadFile = File(...),
    description: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_forms(request):
        return _dashboard_redirect("Only admin or master trainers can manage forms.", "error")

    cleaned_code = form_code.strip().upper()
    catalog = next((item for item in ATL_MER_FORMS if item["code"].upper() == cleaned_code), None)
    if not catalog:
        return _dashboard_redirect("Unknown official form code.", "error")

    try:
        content = await _read_form_upload(file)
    except ValueError as exc:
        return _dashboard_redirect(str(exc), "error")

    stored_name = _safe_upload_name(file.filename)
    _write_form_file(stored_name, content)
    current_account = _current_account(request) or {"name": "Admin"}
    cleaned_description = _clean_form_description(description)

    with SessionLocal() as db:
        existing = (
            db.query(models.AtlForm)
            .filter(func.upper(models.AtlForm.code) == cleaned_code)
            .first()
        )
        if existing:
            _remove_form_file(existing.stored_name)
            existing.title = catalog["title"]
            existing.filename = Path(file.filename).name
            existing.stored_name = stored_name
            existing.uploaded_by = current_account.get("name")
            existing.is_hidden = False
            if cleaned_description is not None:
                existing.description = cleaned_description
        else:
            db.add(
                models.AtlForm(
                    code=cleaned_code,
                    title=catalog["title"],
                    description=cleaned_description,
                    filename=Path(file.filename).name,
                    stored_name=stored_name,
                    uploaded_by=current_account.get("name"),
                    is_hidden=False,
                )
            )
        db.commit()

    return _dashboard_redirect(f"{cleaned_code} replaced successfully.")


@app.post("/forms/{form_id}/replace")
async def replace_atl_form(
    request: Request,
    form_id: int,
    file: UploadFile = File(...),
    title: str = Form(""),
    description: str = Form(""),
):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_forms(request):
        return _dashboard_redirect("Only admin or master trainers can manage forms.", "error")

    try:
        content = await _read_form_upload(file)
    except ValueError as exc:
        return _dashboard_redirect(str(exc), "error")

    stored_name = _safe_upload_name(file.filename)
    current_account = _current_account(request) or {"name": "Admin"}

    with SessionLocal() as db:
        form = _atl_form_record(db, form_id)
        _remove_form_file(form.stored_name)
        _write_form_file(stored_name, content)
        form.filename = Path(file.filename).name
        form.stored_name = stored_name
        form.uploaded_by = current_account.get("name")
        form.is_hidden = False
        cleaned_title = title.strip()
        if cleaned_title:
            form.title = cleaned_title
        cleaned_description = _clean_form_description(description)
        if description.strip() or form.description:
            form.description = cleaned_description
        db.commit()
        form_label = form.code or form.title

    return _dashboard_redirect(f"{form_label} replaced successfully.")


@app.post("/forms/official/{form_code}/delete")
async def hide_official_atl_form(request: Request, form_code: str):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_forms(request):
        return _dashboard_redirect("Only admin or master trainers can manage forms.", "error")

    cleaned_code = form_code.strip().upper()
    catalog = next((item for item in ATL_MER_FORMS if item["code"].upper() == cleaned_code), None)
    if not catalog:
        return _dashboard_redirect("Unknown official form code.", "error")

    current_account = _current_account(request) or {"name": "Admin"}
    with SessionLocal() as db:
        existing = (
            db.query(models.AtlForm)
            .filter(func.upper(models.AtlForm.code) == cleaned_code)
            .first()
        )
        if existing:
            _remove_form_file(existing.stored_name)
            existing.is_hidden = True
            existing.stored_name = ""
            existing.filename = ""
            existing.uploaded_by = current_account.get("name")
        else:
            db.add(
                models.AtlForm(
                    code=cleaned_code,
                    title=catalog["title"],
                    filename="",
                    stored_name="",
                    uploaded_by=current_account.get("name"),
                    is_hidden=True,
                )
            )
        db.commit()

    return _dashboard_redirect(f"{cleaned_code} removed from the forms list.")


@app.post("/forms/{form_id}/delete")
async def delete_atl_form(request: Request, form_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")
    if not _can_manage_forms(request):
        return _dashboard_redirect("Only admin or master trainers can manage forms.", "error")

    with SessionLocal() as db:
        form = _atl_form_record(db, form_id)
        form_label = form.code or form.title
        is_official_override = (form.code or "").strip().upper() in _official_form_codes()
        _remove_form_file(form.stored_name)
        db.delete(form)
        db.commit()

    if is_official_override:
        return _dashboard_redirect(f"{form_label} restored to the default official version.")
    return _dashboard_redirect(f"{form_label} deleted successfully.")


@app.get("/forms/download/{form_id}")
async def download_atl_form(request: Request, form_id: int):
    if not _is_authenticated(request):
        return RedirectResponse("/login")

    with SessionLocal() as db:
        form = db.query(models.AtlForm).filter(models.AtlForm.id == form_id).first()
        if not form:
            raise HTTPException(status_code=404, detail="Form not found.")
        target = FORMS_UPLOAD_DIR / form.stored_name
        if not form.stored_name or not target.exists():
            raise HTTPException(status_code=404, detail="Form file not found.")

    return FileResponse(target, filename=form.filename, media_type="application/octet-stream")


@app.get("/health")
async def health():
    return {"ok": True, "service": "student-dashboard"}
